"""Lightweight integrated source preparation for the one-repository OCL workflow.

This module is intentionally deterministic. It handles common structured OCL
source layouts and publishes the same kind of contract the downstream OCL code
expects: standardized CSVs, metadata, manifest and lineage.

It never changes source workbooks. Ambiguous sheets remain visible in metadata
instead of being silently discarded.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MAX_HEADER_SCAN = 30

LABEL_HINTS = (
    "account description", "description", "account name", "line item", "line_item",
    "account", "name", "label", "category",
)
CODE_HINTS = ("gl account", "gl code", "account code", "account number", "code")
ENTITY_HINTS = ("entity", "company", "business unit", "legal entity")
CURRENCY_HINTS = ("currency", "curr")
PERIOD_HINTS = ("period", "date", "month", "year", "fiscal year", "fiscal_year")
AMOUNT_HINTS = ("amount", "balance", "value", "closing balance", "closing_balance")
MOVEMENT_HINTS = ("movement type", "movement", "event type", "transaction type", "type")
CONTROL_HINTS = ("control", "control line", "tb line", "line", "caption")
REVENUE_HINTS = ("revenue", "sales", "turnover")
PAYROLL_HINTS = ("payroll", "salary cost", "wages", "employee cost", "personnel cost")


@dataclass(frozen=True)
class DataPrepResult:
    output_dir: Path
    manifest_path: Path
    metadata_path: Path
    datasets: tuple[Path, ...]
    warnings: tuple[str, ...]


def prepare_source_package(source_dir: Path, output_dir: Path) -> DataPrepResult:
    """Prepare all supported workbooks under *source_dir* into one publication."""
    source_dir = Path(source_dir).resolve()
    output_dir = Path(output_dir).resolve()
    if not source_dir.is_dir():
        raise FileNotFoundError(f"Source directory does not exist: {source_dir}")
    files = [p for p in sorted(source_dir.rglob("*")) if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() in EXCEL_EXTENSIONS]
    if not files:
        raise ValueError("No supported Excel source files were found in references/source/.")

    output_dir.mkdir(parents=True, exist_ok=True)
    for old in output_dir.iterdir():
        if old.is_file():
            old.unlink()

    buckets: dict[str, list[dict[str, Any]]] = {
        "ocl_annual.csv": [],
        "ocl_monthly.csv": [],
        "ocl_movements.csv": [],
        "tb_control.csv": [],
        "revenue_context.csv": [],
        "payroll_context.csv": [],
    }
    lineage: list[dict[str, str]] = []
    source_snapshot: list[dict[str, Any]] = []
    unclassified: list[dict[str, Any]] = []
    warnings: list[str] = []

    for path in files:
        digest_before = _sha256(path)
        source_snapshot.append({
            "file": path.name,
            "relative_path": path.relative_to(source_dir).as_posix(),
            "sha256": digest_before,
            "size": path.stat().st_size,
        })
        workbook = load_workbook(path, data_only=False, read_only=False)
        try:
            for worksheet in workbook.worksheets:
                result = _extract_sheet(path, worksheet)
                if result is None:
                    if _sheet_has_content(worksheet):
                        unclassified.append({
                            "file": path.name,
                            "worksheet": worksheet.title,
                            "reason": "No supported structured OCL dataset pattern was identified with sufficient deterministic evidence.",
                        })
                    continue
                for dataset_name, rows in result.items():
                    buckets[dataset_name].extend(rows)
                    for row in rows:
                        source_id = str(row.get("Source_Record_ID") or "")
                        if not source_id:
                            continue
                        ref = json.loads(source_id)
                        lineage.append({
                            "Output_File": dataset_name,
                            "Source_Record_ID": source_id,
                            "Source_File": ref["source_id"],
                            "Source_Sheet": ref["worksheet_name"],
                            "Source_Cell": ref["source_cell"],
                        })
        finally:
            workbook.close()
        if _sha256(path) != digest_before:
            raise RuntimeError(f"Source workbook changed while being prepared: {path.name}")

    dataset_paths: list[Path] = []
    for name, rows in buckets.items():
        if not rows:
            continue
        columns = _columns_for(name)
        target = output_dir / name
        _write_csv(target, columns, rows)
        dataset_paths.append(target)

    if not dataset_paths:
        raise ValueError("No usable OCL-related structured dataset could be prepared from the supplied source files.")

    _write_csv(
        output_dir / "lineage.csv",
        ["Output_File", "Source_Record_ID", "Source_File", "Source_Sheet", "Source_Cell"],
        lineage,
    )
    _write_csv(
        output_dir / "field_lineage.csv",
        ["Output_Field", "Meaning"],
        [
            {"Output_Field": "Source_Record_ID", "Meaning": "JSON pointer to original workbook, worksheet, cell and row."},
            {"Output_Field": "Period", "Meaning": "Source reporting period preserved as text."},
            {"Output_Field": "Source_Label", "Meaning": "Source line-item/account description."},
            {"Output_Field": "Source_Code", "Meaning": "Source account/code where available."},
            {"Output_Field": "Amount", "Meaning": "Source-derived financial amount; no balancing plug."},
        ],
    )

    if unclassified:
        warnings.append(f"{len(unclassified)} populated worksheet(s) were not classified into a supported structured dataset and remain listed in metadata.")
    logical_datasets = [
        {
            "file": path.name,
            "purpose": _purpose(path.name),
            "row_count": len(buckets[path.name]),
            "columns": _columns_for(path.name),
        }
        for path in dataset_paths
    ]
    execution_id = _execution_id(source_snapshot)
    metadata = {
        "workflow_run_id": execution_id,
        "logical_datasets": logical_datasets,
        "source_snapshot": source_snapshot,
        "unclassified_material": unclassified,
        "warnings": warnings,
        "integration_mode": "EMBEDDED_DETERMINISTIC_OCL_FAST_PATH",
    }
    metadata_path = output_dir / "databook_metadata.json"
    metadata_path.write_text(json.dumps(metadata, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    manifest = {
        "execution_id": execution_id,
        "final_execution_status": "COMPLETED_WITH_WARNINGS" if warnings else "COMPLETED",
        "outputs_created": [path.name for path in dataset_paths],
        "source_file_count": len(files),
        "source_snapshot": source_snapshot,
        "unclassified_worksheet_count": len(unclassified),
    }
    manifest_path = output_dir / "execution_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return DataPrepResult(output_dir, manifest_path, metadata_path, tuple(dataset_paths), tuple(warnings))


def _extract_sheet(path: Path, ws) -> dict[str, list[dict[str, Any]]] | None:
    header_row = _find_header_row(ws)
    if header_row is None:
        return None
    headers = [_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    normalized = [_norm(value) for value in headers]
    by_name = {name: index + 1 for index, name in enumerate(normalized) if name}

    movement_col = _find_column(normalized, MOVEMENT_HINTS)
    amount_col = _find_column(normalized, AMOUNT_HINTS)
    period_col = _find_column(normalized, PERIOD_HINTS)
    label_col = _find_column(normalized, LABEL_HINTS)
    code_col = _find_column(normalized, CODE_HINTS)
    entity_col = _find_column(normalized, ENTITY_HINTS)
    currency_col = _find_column(normalized, CURRENCY_HINTS)
    control_col = _find_column(normalized, CONTROL_HINTS)
    revenue_col = _find_column(normalized, REVENUE_HINTS)
    payroll_col = _find_column(normalized, PAYROLL_HINTS)

    if movement_col and amount_col and label_col:
        return {"ocl_movements.csv": _extract_long_records(path, ws, header_row, period_col, label_col, code_col, entity_col, currency_col, amount_col, movement_col=movement_col, default_period=_sheet_period(ws.title))}

    if period_col and amount_col and control_col and not label_col:
        return {"tb_control.csv": _extract_control(path, ws, header_row, period_col, control_col, amount_col)}

    if period_col and (revenue_col or payroll_col):
        result: dict[str, list[dict[str, Any]]] = {}
        if revenue_col:
            result["revenue_context.csv"] = _extract_context(path, ws, header_row, period_col, revenue_col, "Revenue")
        if payroll_col:
            result["payroll_context.csv"] = _extract_context(path, ws, header_row, period_col, payroll_col, "Payroll")
        if result:
            return result

    if period_col and amount_col and label_col:
        records = _extract_long_records(path, ws, header_row, period_col, label_col, code_col, entity_col, currency_col, amount_col)
        if not records:
            return None
        monthly_count = sum(_period_kind(str(row["Period"])) == "MONTHLY" for row in records)
        annual_count = sum(_period_kind(str(row["Period"])) == "ANNUAL" for row in records)
        name = "ocl_monthly.csv" if monthly_count > annual_count else "ocl_annual.csv"
        return {name: records}

    period_columns = [(column, headers[column - 1], _period_kind(headers[column - 1])) for column in range(1, ws.max_column + 1)]
    usable_periods = [(column, header, kind) for column, header, kind in period_columns if kind]
    if len(usable_periods) >= 2 and label_col:
        kinds = [kind for _, _, kind in usable_periods]
        name = "ocl_monthly.csv" if kinds.count("MONTHLY") > kinds.count("ANNUAL") else "ocl_annual.csv"
        return {name: _extract_wide_records(path, ws, header_row, usable_periods, label_col, code_col, entity_col, currency_col)}
    return None


def _extract_wide_records(path: Path, ws, header_row: int, period_columns, label_col, code_col, entity_col, currency_col) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        label = _text(ws.cell(row, label_col).value)
        if not label:
            continue
        code = _text(ws.cell(row, code_col).value) if code_col else ""
        entity = _text(ws.cell(row, entity_col).value) if entity_col else ""
        currency = _text(ws.cell(row, currency_col).value) if currency_col else ""
        for column, period, _kind in period_columns:
            value = ws.cell(row, column).value
            amount = _numeric(value)
            if amount is None:
                continue
            rows.append(_record(path, ws.title, row, column, period, label, code, entity, currency, amount))
    return rows


def _extract_long_records(path: Path, ws, header_row: int, period_col, label_col, code_col, entity_col, currency_col, amount_col, *, movement_col=None, default_period: str | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        label = _text(ws.cell(row, label_col).value)
        amount = _numeric(ws.cell(row, amount_col).value)
        if not label or amount is None:
            continue
        period = _text(ws.cell(row, period_col).value) if period_col else default_period
        if not period:
            continue
        result = _record(
            path, ws.title, row, amount_col, period, label,
            _text(ws.cell(row, code_col).value) if code_col else "",
            _text(ws.cell(row, entity_col).value) if entity_col else "",
            _text(ws.cell(row, currency_col).value) if currency_col else "",
            amount,
        )
        if movement_col:
            result["Movement_Type"] = _text(ws.cell(row, movement_col).value)
            if not result["Movement_Type"]:
                continue
        rows.append(result)
    return rows


def _extract_control(path: Path, ws, header_row: int, period_col: int, control_col: int, amount_col: int) -> list[dict[str, Any]]:
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        period = _text(ws.cell(row, period_col).value)
        control = _text(ws.cell(row, control_col).value)
        amount = _numeric(ws.cell(row, amount_col).value)
        if period and control and amount is not None:
            rows.append({"Period": period, "Control": control, "Amount": amount})
    return rows


def _extract_context(path: Path, ws, header_row: int, period_col: int, value_col: int, context_type: str) -> list[dict[str, Any]]:
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        period = _text(ws.cell(row, period_col).value)
        amount = _numeric(ws.cell(row, value_col).value)
        if period and amount is not None:
            rows.append({"Period": period, "Amount": amount})
    return rows


def _record(path: Path, sheet: str, row: int, amount_col: int, period: Any, label: str, code: str, entity: str, currency: str, amount: float | int) -> dict[str, Any]:
    return {
        "Source_Record_ID": json.dumps({
            "source_id": path.name,
            "worksheet_name": sheet,
            "source_cell": f"{get_column_letter(amount_col)}{row}",
            "row": row,
        }, separators=(",", ":")),
        "Period": _period_text(period),
        "Source_Label": label,
        "Source_Code": code,
        "Entity": entity,
        "Currency": currency,
        "Amount": amount,
    }


def _find_header_row(ws) -> int | None:
    best: tuple[int, int] | None = None
    hints = set(_norm(value) for value in (*LABEL_HINTS, *CODE_HINTS, *PERIOD_HINTS, *AMOUNT_HINTS, *MOVEMENT_HINTS, *CONTROL_HINTS, *REVENUE_HINTS, *PAYROLL_HINTS))
    for row in range(1, min(ws.max_row, MAX_HEADER_SCAN) + 1):
        values = [_norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        present = [value for value in values if value]
        if len(present) < 2:
            continue
        explicit = sum(1 for value in present if any(hint == value or hint in value for hint in hints))
        periods = sum(1 for value in present if _period_kind(value))
        score = explicit * 5 + periods * 3 + min(len(present), 8)
        candidate = (score, -row)
        if best is None or candidate > best:
            best = candidate
    return -best[1] if best and best[0] >= 8 else None


def _find_column(normalized_headers: list[str], hints: Iterable[str]) -> int | None:
    normalized_hints = [_norm(hint) for hint in hints]
    exact = [(index + 1, header) for index, header in enumerate(normalized_headers) if header and header in normalized_hints]
    if exact:
        return exact[0][0]
    for index, header in enumerate(normalized_headers, start=1):
        if header and any(hint in header for hint in normalized_hints):
            return index
    return None


def _period_kind(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return "MONTHLY"
    text = _text(value)
    if not text:
        return None
    compact = text.strip().upper()
    if re.fullmatch(r"20\d{2}[-_/](0?[1-9]|1[0-2])", compact):
        return "MONTHLY"
    if re.fullmatch(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[-_ /]?\d{2,4}", compact):
        return "MONTHLY"
    if re.fullmatch(r"FY[-_ ]?\d{2,4}", compact) or re.fullmatch(r"20\d{2}", compact):
        return "ANNUAL"
    return None


def _sheet_period(title: str) -> str | None:
    match = re.search(r"FY[-_ ]?\d{2,4}|20\d{2}", title.upper())
    return match.group(0).replace(" ", "") if match else None


def _period_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _sheet_has_content(ws) -> bool:
    return any(ws.cell(r, c).value not in (None, "") for r in range(1, min(ws.max_row, 50) + 1) for c in range(1, min(ws.max_column, 30) + 1))


def _numeric(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        negative = text.startswith("(") and text.endswith(")")
        if negative:
            text = text[1:-1]
        try:
            number = float(text)
            return -number if negative else number
        except ValueError:
            return None
    return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return " ".join(str(value).strip().split())


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _text(value).casefold()).strip()


def _columns_for(name: str) -> list[str]:
    if name in {"ocl_annual.csv", "ocl_monthly.csv"}:
        return ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Currency", "Amount"]
    if name == "ocl_movements.csv":
        return ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Currency", "Movement_Type", "Amount"]
    if name == "tb_control.csv":
        return ["Period", "Control", "Amount"]
    return ["Period", "Amount"]


def _purpose(name: str) -> str:
    return {
        "ocl_annual.csv": "period-end OCL/current-liability listing",
        "ocl_monthly.csv": "monthly OCL/current-liability balances",
        "ocl_movements.csv": "OCL movement or roll-forward schedule",
        "tb_control.csv": "trial-balance/control totals",
        "revenue_context.csv": "optional revenue context",
        "payroll_context.csv": "optional payroll context",
    }[name]


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _execution_id(snapshot: list[dict[str, Any]]) -> str:
    encoded = json.dumps(snapshot, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return "INTEGRATED_" + hashlib.sha256(encoded).hexdigest()[:16].upper()
