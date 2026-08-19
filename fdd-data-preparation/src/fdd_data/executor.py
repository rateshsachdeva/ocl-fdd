"""Generic deterministic execution of an AI-authored, source-bound processing plan.

The AI host decides what the source means and writes the plan. This module only
executes validated mappings/reshaping and proves source/output/lineage coverage.
"""
from __future__ import annotations

from collections import defaultdict
import csv
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
from pathlib import Path
from time import perf_counter
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .completeness import completeness_checks, evaluate_completeness
from .lineage import SourceRowIdentity, lineage_source_rows, output_record_id, parse_output_record_id, source_row_identity
from .processing_plan import ExecutionNotApprovedError, ProcessingPlanValidationError, SourceIntegrityError, assert_execution_allowed, compare_source_snapshot, validate_processing_plan
from .source_data import read_region

EXECUTION_STATUSES = {"COMPLETED", "COMPLETED_WITH_WARNINGS", "FAILED_VALIDATION", "EXECUTION_BLOCKED", "FAILED"}
SUPPORTED_OPERATIONS = {"SELECT_REGION", "MAP_COLUMNS", "RENAME_FIELDS", "ADD_CONTEXT_FIELD", "UNION", "UNPIVOT", "PARSE_NUMERIC", "PARSE_DATE", "TRIM_TEXT", "FILTER_EXCLUDE", "SOURCE_RECORD_ID", "DERIVE_LINEAGE"}


class UnsupportedOperationError(ValueError):
    pass


def execute_processing_plan(
    plan_path: Path,
    profile_path: Path,
    dataset_map_path: Path,
    source_directory: Path,
    output_root: Path = Path("output"),
    *,
    staging_directory: Path | None = None,
    audit_artifacts: bool = False,
    after_transform_hook: Callable[[], None] | None = None,
) -> dict[str, Any]:
    started = perf_counter()
    plan, profile, dataset_map = _read_json(plan_path), _read_json(profile_path), _read_json(dataset_map_path)
    execution_id, execution_directory = _create_execution_directory(output_root)
    manifest: dict[str, Any] = {
        "execution_id": execution_id,
        "execution_timestamp": datetime.now(timezone.utc).isoformat(),
        "plan_id": plan.get("plan_metadata", {}).get("plan_id"),
        "plan_version": plan.get("plan_metadata", {}).get("plan_version"),
        "plan_hash": plan.get("plan_metadata", {}).get("plan_hash"),
        "source_snapshot_status": None,
        "outputs_attempted": [], "outputs_created": [],
        "input_row_counts": [], "output_row_counts": [], "operations_executed": [],
        "validation_checks": [], "warnings": [], "errors": [],
        "data_completeness": {}, "field_lineage": [], "completeness_summary": [],
        "performance": {}, "final_execution_status": "FAILED",
    }
    issues: list[dict[str, Any]] = []
    exclusions: list[dict[str, Any]] = []
    lineage: list[dict[str, Any]] = []
    rows_by_output: dict[str, list[dict[str, Any]]] = {}
    output_data: list[dict[str, Any]] = []
    field_lineage: list[dict[str, Any]] = []
    try:
        errors = validate_processing_plan(plan, profile, dataset_map)
        if errors:
            raise ProcessingPlanValidationError("\n".join(errors))
        manifest["source_snapshot_status"] = assert_execution_allowed(plan, source_directory, for_stage5=True)["comparison_result"]
        _assert_supported_operations(plan)
        region_index = _profile_index(profile)
        source_lookup = {item["source_id"]: item for item in profile.get("source_files", [])}

        for output in plan.get("proposed_outputs", []):
            manifest["outputs_attempted"].append(output["output_id"])
            rows, out_lineage, out_exclusions, data = _execute_output(
                output, profile, region_index, source_lookup, source_directory, staging_directory
            )
            rows_by_output[output["output_id"]] = rows
            lineage.extend(out_lineage)
            exclusions.extend(out_exclusions)
            output_data.append(data)
            field_lineage.extend(data["field_lineage"])
            manifest["input_row_counts"].extend(data["input_row_counts"])
            manifest["output_row_counts"].append({"output_id": output["output_id"], "expected_rows": data["expected_rows"], "actual_rows": len(rows)})
            manifest["operations_executed"].extend(data["operations"])
            checks = _validate_output(output, rows, data)
            manifest["validation_checks"].extend(checks)
            for check in checks:
                if check["status"] == "FAIL":
                    issues.append(_issue("ERROR", check["check"], check["message"], output["output_id"]))

        non_output_exclusions, non_output_rows = _non_output_exclusions(
            plan, profile, dataset_map, region_index, source_lookup, source_directory, staging_directory
        )
        exclusions.extend(non_output_exclusions)
        completeness = _evaluate_plan_completeness(plan, profile, dataset_map, output_data, lineage, non_output_rows)
        manifest["data_completeness"] = completeness
        manifest["completeness_summary"] = _completeness_summary(completeness)
        manifest["validation_checks"].extend(completeness_checks(completeness))
        manifest["field_lineage"] = _unique_dicts(field_lineage)

        if after_transform_hook:
            after_transform_hook()
        post = compare_source_snapshot(plan["plan_metadata"]["source_snapshot"], source_directory)
        if post["status"] != "SOURCE_UNCHANGED":
            reason = post["findings"][0]["reason"]
            manifest["validation_checks"].append(_check("POST_EXECUTION_SOURCE_INTEGRITY", "FAIL", reason))

        if any(item["status"] == "FAIL" for item in manifest["validation_checks"]):
            manifest["final_execution_status"] = "FAILED_VALIDATION"
        else:
            for output in plan.get("proposed_outputs", []):
                path = execution_directory / output["filename"]
                _write_csv(path, output["output_columns"], rows_by_output[output["output_id"]])
                _validate_reopened_csv(path, [item["Output_Column"] for item in output["output_columns"]], len(rows_by_output[output["output_id"]]), manifest["validation_checks"])
                manifest["outputs_created"].append(output["filename"])
            manifest["final_execution_status"] = "COMPLETED" if not any(item["status"] == "FAIL" for item in manifest["validation_checks"]) else "FAILED_VALIDATION"
        manifest["performance"]["total_active_processing_seconds"] = round(perf_counter() - started, 6)
        _write_artifacts(execution_directory, manifest, lineage, exclusions, issues, field_lineage=manifest["field_lineage"], audit_artifacts=audit_artifacts)
    except (ExecutionNotApprovedError, SourceIntegrityError) as error:
        manifest["final_execution_status"] = "EXECUTION_BLOCKED"
        manifest["errors"].append(str(error))
        _write_artifacts(execution_directory, manifest, lineage, exclusions, issues, field_lineage=field_lineage, audit_artifacts=True)
    except Exception as error:
        manifest["final_execution_status"] = "FAILED"
        manifest["errors"].append(f"{type(error).__name__}: {error}")
        issues.append(_issue("ERROR", type(error).__name__, str(error), ""))
        _write_artifacts(execution_directory, manifest, lineage, exclusions, issues, field_lineage=field_lineage, audit_artifacts=True)
    return {"execution_id": execution_id, "execution_directory": execution_directory, "manifest": manifest}


def _execute_output(output, profile, region_index, source_lookup, source_directory, staging_directory):
    direct_columns = [item for item in output["output_columns"] if item.get("Source_Type") == "DIRECT_COLUMN"]
    source_record_column = _source_record_column(output)
    unpivot = output.get("unpivot") if isinstance(output.get("unpivot"), dict) else None
    unpivot_mapping = dict(zip(unpivot.get("source_field_ids", []), unpivot.get("dimension_values", []))) if unpivot else {}
    records, lineage, exclusions, counts, field_lineage = [], [], [], [], []
    relevant_rows: set[SourceRowIdentity] = set()
    retained_rows: set[SourceRowIdentity] = set()
    excluded_rows: set[SourceRowIdentity] = set()
    included_source_ids, included_region_ids = set(), set()
    expected_rows = 0
    source_totals: dict[tuple[str, str], Decimal] = defaultdict(Decimal)

    for assignment in output.get("source_assignments", []):
        for reference in assignment.get("source_references", []):
            region = region_index[reference["region_id"]]
            source = source_lookup[reference["source_id"]]
            region_data = read_region(source, reference["worksheet_name"], region, profile, source_directory, staging_directory)
            _validate_region_fields(region_data.fields, direct_columns, unpivot_mapping, reference)
            included_source_ids.add(source["source_id"]); included_region_ids.add(reference["region_id"])
            counts.append({"output_id": output["output_id"], "source_id": source["source_id"], "source_file": source["filename"], "source_sheet": reference["worksheet_name"], "source_region": reference["region_id"], "input_record_count": len(region_data.rows)})
            region_unpivot = {field_id: dimension for field_id, dimension in unpivot_mapping.items() if field_id in region_data.fields}
            for physical_row, values in region_data.rows:
                identity = source_row_identity(source["source_id"], reference["worksheet_name"], reference["region_id"], physical_row)
                relevant_rows.add(identity)
                base = _map_columns(output["output_columns"], direct_columns, values, source, reference)
                if _matches_filter_exclusion(base, output.get("operation_steps", [])):
                    excluded_rows.add(identity)
                    exclusions.append({"source_id": source["source_id"], "source_file": source["filename"], "source_sheet": reference["worksheet_name"], "source_region": reference["region_id"], "reason": "FILTER_EXCLUDE condition in approved plan.", "count": 1})
                    continue
                retained_rows.add(identity)
                if unpivot:
                    if not region_unpivot:
                        raise ProcessingPlanValidationError(f"No configured unpivot source field IDs resolve in {reference['worksheet_name']}.")
                    expected_rows += len(region_unpivot)
                    for measure_field_id, dimension in region_unpivot.items():
                        record = dict(base)
                        record[unpivot["dimension_output_column"]] = dimension
                        record[unpivot["measure_output_column"]] = values.get(measure_field_id)
                        if source_record_column:
                            record[source_record_column] = output_record_id(identity, {unpivot["dimension_output_column"]: dimension})
                        _apply_operations(record, output.get("operation_steps", []))
                        records.append(record)
                        amount = record.get(unpivot["measure_output_column"])
                        if amount not in (None, ""):
                            try: source_totals[(source["source_id"], str(dimension))] += _decimal(amount)
                            except InvalidOperation: pass
                        lineage.append(_lineage_record(output, len(records), record, source_record_column, source, reference, physical_row, {unpivot["dimension_output_column"]: dimension}))
                        field_lineage.extend(_field_lineage_rows(output, direct_columns, region_data.fields, measure_field_id))
                else:
                    expected_rows += 1
                    record = dict(base)
                    if source_record_column:
                        record[source_record_column] = output_record_id(identity)
                    _apply_operations(record, output.get("operation_steps", []))
                    records.append(record)
                    lineage.append(_lineage_record(output, len(records), record, source_record_column, source, reference, physical_row, {}))
                    field_lineage.extend(_direct_field_lineage(output, direct_columns, region_data.fields))

    return records, lineage, exclusions, {
        "input_row_counts": counts, "expected_rows": expected_rows,
        "source_totals": source_totals, "unpivot": unpivot,
        "relevant_rows": relevant_rows, "retained_rows": retained_rows, "excluded_rows": excluded_rows,
        "included_source_ids": included_source_ids, "included_region_ids": included_region_ids,
        "field_lineage": field_lineage,
        "operations": [item.get("operation") for item in output.get("transformations", [])],
    }


def _validate_output(output, rows, data):
    checks = [_check("EXPECTED_ROW_COUNT", "PASS" if len(rows) == data["expected_rows"] else "FAIL", f"Expected {data['expected_rows']}; actual {len(rows)}.")]
    record_id = _source_record_column(output)
    if record_id:
        identifiers = [row.get(record_id) for row in rows]
        checks.append(_check("SOURCE_RECORD_ID_UNIQUENESS", "PASS" if len(identifiers) == len(set(identifiers)) and all(identifiers) else "FAIL", "Generated source record IDs are present and unique."))
    if data["unpivot"]:
        period_column = data["unpivot"]["dimension_output_column"]
        amount_column = data["unpivot"]["measure_output_column"]
        output_totals: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
        for row in rows:
            payload = parse_output_record_id(row.get(record_id, "")) if record_id else {}
            amount = row.get(amount_column)
            if amount not in (None, ""):
                try: output_totals[(str(payload.get("source_id", "")), str(row.get(period_column, "")))] += _decimal(amount)
                except InvalidOperation: pass
        comparison = _compare_maps(data["source_totals"], output_totals, _reconciliation_rule(output, "AMOUNT_TOTALS_BY_PERIOD_AND_SOURCE"))
        checks.append(_check("DETAIL_TO_OUTPUT_RECONCILIATION", "PASS" if comparison["passed"] else "FAIL", "Output totals match source detail totals by source partition and generated period/dimension.", **comparison))
    return checks


def compare_reconciliation_values(left, right, rule=None):
    rule = rule or {"comparison_method": "EXACT"}
    method = rule.get("comparison_method", "EXACT")
    left_value, right_value = _decimal(left), _decimal(right)
    raw_difference = left_value - right_value
    parameter = {}
    if method == "EXACT": evaluated_difference, passed = raw_difference, raw_difference == 0
    elif method == "ROUND_TO_PRECISION":
        places = rule.get("decimal_places")
        if not isinstance(places, int) or places < 0: raise ProcessingPlanValidationError("ROUND_TO_PRECISION requires non-negative decimal_places.")
        precision = Decimal(1).scaleb(-places)
        evaluated_difference = left_value.quantize(precision, rounding=ROUND_HALF_UP) - right_value.quantize(precision, rounding=ROUND_HALF_UP)
        passed, parameter = evaluated_difference == 0, {"decimal_places": places}
    elif method == "ABSOLUTE_TOLERANCE":
        tolerance = _decimal(rule["tolerance"])
        if tolerance < 0: raise ProcessingPlanValidationError("ABSOLUTE_TOLERANCE requires non-negative tolerance.")
        evaluated_difference, passed, parameter = abs(raw_difference), abs(raw_difference) <= tolerance, {"tolerance": str(tolerance)}
    else: raise ProcessingPlanValidationError(f"Unsupported reconciliation comparison method: {method}")
    return {"passed": passed, "comparison_method": method, "comparison_parameter": parameter, "raw_difference": str(raw_difference), "evaluated_difference": str(evaluated_difference)}


def _compare_maps(left, right, rule):
    comparisons = [compare_reconciliation_values(left.get(key, Decimal(0)), right.get(key, Decimal(0)), rule) for key in set(left) | set(right)]
    return {"passed": all(item["passed"] for item in comparisons), "compared_key_count": len(comparisons), "mismatch_count": sum(not item["passed"] for item in comparisons), "comparison_method": rule.get("comparison_method", "EXACT")}


def _reconciliation_rule(output, check_name):
    return next((item for item in output.get("validations", []) if item.get("check") == check_name), {"comparison_method": "EXACT"})


def _map_columns(output_columns, direct_columns, values, source, reference):
    record = {}
    for item in direct_columns:
        matches = [field_id for field_id in item.get("Source_Field_IDs", []) if field_id in values]
        if len(matches) != 1:
            raise ProcessingPlanValidationError(f"{item['Output_Column']} must resolve to exactly one source field in {reference['region_id']}.")
        record[item["Output_Column"]] = values.get(matches[0])
    for item in output_columns:
        source_type = item.get("Source_Type")
        if source_type == "CONSTANT": record[item["Output_Column"]] = item.get("Constant_Value")
        elif source_type == "FILENAME_METADATA": record[item["Output_Column"]] = source.get(item.get("Metadata_Key", "filename"))
        elif source_type == "SHEET_METADATA": record[item["Output_Column"]] = reference.get(item.get("Metadata_Key", "worksheet_name"))
        elif source_type == "CONTEXT_METADATA": record[item["Output_Column"]] = item.get("Context_Value")
    return record


def _apply_operations(record, steps):
    for step in steps:
        operation = step.get("operation")
        if operation == "TRIM_TEXT":
            for field in step.get("fields", []):
                if isinstance(record.get(field), str): record[field] = record[field].strip()
        elif operation == "PARSE_NUMERIC":
            for field in step.get("fields", []):
                if isinstance(record.get(field), str): record[field] = _parse_numeric(record[field])
        elif operation == "PARSE_DATE":
            for field in step.get("fields", []):
                if isinstance(record.get(field), str): record[field] = _parse_date(record[field], step.get("formats", []))
        elif operation == "ADD_CONTEXT_FIELD": record[step["field"]] = step.get("value")
        elif operation in {"MAP_COLUMNS", "RENAME_FIELDS"}:
            for source_field, target_field in step.get("mapping", {}).items():
                if source_field in record: record[target_field] = record.pop(source_field)


def _matches_filter_exclusion(record, steps):
    return any(item.get("operation") == "FILTER_EXCLUDE" and record.get(item.get("field")) == item.get("equals") for item in steps)


def _validate_region_fields(fields, direct_columns, unpivot_mapping, reference):
    missing = [item["Output_Column"] for item in direct_columns if len(set(item.get("Source_Field_IDs", [])) & set(fields)) != 1]
    if missing: raise ProcessingPlanValidationError(f"Stable source fields do not resolve in {reference['worksheet_name']} for: {', '.join(missing)}")
    if unpivot_mapping and not set(unpivot_mapping) & set(fields): raise ProcessingPlanValidationError(f"No unpivot source field IDs resolve in {reference['worksheet_name']}.")


def _lineage_record(output, row_number, record, source_record_column, source, reference, physical_row, generated):
    return {"Output_Dataset": output["output_id"], "Output_Row_Number": row_number, "Source_Record_ID": record.get(source_record_column) if source_record_column else None, "Source_ID": source["source_id"], "Source_File": source["filename"], "Source_Sheet": reference["worksheet_name"], "Source_Row": physical_row, "Source_Region": reference["region_id"], "Generated_Context": json.dumps(generated, ensure_ascii=False)}


def _field_lineage_rows(output, direct_columns, region_fields, measure_field_id):
    rows = _direct_field_lineage(output, direct_columns, region_fields)
    if isinstance(output.get("unpivot"), dict):
        for output_name in (output["unpivot"]["dimension_output_column"], output["unpivot"]["measure_output_column"]):
            column = next((item for item in output["output_columns"] if item["Output_Column"] == output_name), None)
            if column and measure_field_id in region_fields: rows.append(_field_lineage_row(output, column, region_fields[measure_field_id]))
    return rows


def _direct_field_lineage(output, direct_columns, region_fields):
    rows = []
    for column in direct_columns:
        field_id = next((item for item in column.get("Source_Field_IDs", []) if item in region_fields), None)
        if field_id: rows.append(_field_lineage_row(output, column, region_fields[field_id]))
    return rows


def _field_lineage_row(output, column, field):
    return {"Output_Dataset": output["output_id"], "Output_Column": column["Output_Column"], "Concept": column.get("Concept"), "Source_Field_ID": field["field_id"], "Exact_Original_Header": field.get("exact_original_header"), "Source_ID": field["source_id"], "Source_Sheet": field["worksheet_name"], "Source_Region": field["region_id"], "Physical_Column": field["physical_column"]}


def _non_output_exclusions(plan, profile, dataset_map, region_index, source_lookup, source_directory, staging_directory):
    datasets = {item["logical_dataset_id"]: item for item in dataset_map.get("logical_datasets", [])}
    rows: set[SourceRowIdentity] = set(); exclusions = []
    for item in plan.get("non_output_logical_datasets", []):
        dataset = datasets.get(item.get("logical_dataset_id"))
        if not dataset: continue
        for reference in dataset.get("contributing_source_regions", []):
            region, source = region_index[reference["region_id"]], source_lookup[reference["source_id"]]
            region_data = read_region(source, reference["worksheet_name"], region, profile, source_directory, staging_directory)
            for physical_row, _ in region_data.rows: rows.add(source_row_identity(source["source_id"], reference["worksheet_name"], reference["region_id"], physical_row))
            exclusions.append({"source_id": source["source_id"], "source_file": source["filename"], "source_sheet": reference["worksheet_name"], "source_region": reference["region_id"], "reason": item.get("reason", "Non-output source region."), "count": len(region_data.rows)})
    return exclusions, rows


def _evaluate_plan_completeness(plan, profile, dataset_map, outputs, lineage, non_output_rows):
    discovered_sources = {item["source_id"] for item in profile.get("source_files", [])}
    discovered_regions = {region["region_id"] for workbook in profile.get("workbook_profiles", []) for worksheet in workbook.get("worksheet_profiles", []) for region in worksheet.get("data_regions", []) if region.get("candidate_confidence") == "TABULAR_CANDIDATE"}
    included_sources = set().union(*(item["included_source_ids"] for item in outputs)) if outputs else set()
    included_regions = set().union(*(item["included_region_ids"] for item in outputs)) if outputs else set()
    retained_rows = set().union(*(item["retained_rows"] for item in outputs)) if outputs else set()
    filtered_rows = set().union(*(item["excluded_rows"] for item in outputs)) if outputs else set()
    relevant_rows = set().union(*(item["relevant_rows"] for item in outputs)) if outputs else set()
    relevant_rows |= non_output_rows
    excluded_rows = filtered_rows | non_output_rows
    datasets = {item.get("logical_dataset_id"): item for item in dataset_map.get("logical_datasets", [])}
    excluded_regions = {ref.get("region_id") for item in plan.get("non_output_logical_datasets", []) for ref in datasets.get(item.get("logical_dataset_id"), {}).get("contributing_source_regions", [])}
    excluded_sources = {ref.get("source_id") for item in plan.get("non_output_logical_datasets", []) for ref in datasets.get(item.get("logical_dataset_id"), {}).get("contributing_source_regions", [])}
    valid_lineage = [item for item in lineage if item.get("Source_Record_ID")]
    return evaluate_completeness(
        discovered_source_ids=discovered_sources, discovered_region_ids=discovered_regions,
        included_source_ids=included_sources, included_region_ids=included_regions,
        excluded_source_ids=excluded_sources, excluded_region_ids=excluded_regions,
        relevant_source_rows=relevant_rows, retained_source_rows=retained_rows, excluded_source_rows=excluded_rows,
        expected_output_records=sum(item["expected_rows"] for item in outputs), actual_output_records=len(lineage),
        output_records_with_valid_lineage=len(valid_lineage), retained_rows_with_output_lineage=lineage_source_rows(item.get("Source_Record_ID") for item in valid_lineage),
    )


def _completeness_summary(summary):
    return ["DATA COMPLETENESS", f"Source files: {summary['source_files']['status']}", f"Source regions: {summary['source_regions']['status']}", f"Source rows: {summary['source_rows']['status']}", f"Output records: {summary['output_records']['status']}", f"Output lineage: {summary['lineage']['status']}", f"OVERALL COMPLETENESS: {summary['overall_status']}"]


def _parse_numeric(value: str) -> Decimal:
    cleaned = value.strip().replace(",", ""); negative = cleaned.startswith("(") and cleaned.endswith(")")
    if negative: cleaned = cleaned[1:-1]
    try: result = Decimal(cleaned.rstrip("%"))
    except InvalidOperation as error: raise ProcessingPlanValidationError("Numeric parse failure for configured value.") from error
    return result / Decimal("100") if value.strip().endswith("%") else (-result if negative else result)


def _parse_date(value: str, formats: list[str]) -> str:
    for date_format in formats:
        try: return datetime.strptime(value, date_format).date().isoformat()
        except ValueError: pass
    raise ProcessingPlanValidationError("Date parse failure for configured value.")


def _decimal(value): return Decimal(str(value))
def _profile_index(profile): return {region["region_id"]: region for workbook in profile.get("workbook_profiles", []) for sheet in workbook.get("worksheet_profiles", []) for region in sheet.get("data_regions", [])}
def _source_record_column(output): return next((item["Output_Column"] for item in output["output_columns"] if item["Output_Column"] == "Source_Record_ID"), None)
def _check(check, status, message, **details): return {"check": check, "status": status, "message": message, **details}
def _issue(severity, issue_type, message, output_id): return {"severity": severity, "issue_type": issue_type, "source": output_id, "location": "", "message": message, "action_status": "OPEN"}


def _assert_supported_operations(plan):
    for output in plan.get("proposed_outputs", []):
        for step in output.get("transformations", []) + output.get("operation_steps", []):
            if step.get("operation") not in SUPPORTED_OPERATIONS: raise UnsupportedOperationError(str(step.get("operation")))


def _create_execution_directory(root: Path):
    base = f"EXEC_{datetime.now().strftime('%Y%m%d_%H%M%S')}"; candidate, suffix = Path(root) / base, 1
    while candidate.exists(): suffix += 1; candidate = Path(root) / f"{base}_{suffix:02d}"
    candidate.mkdir(parents=True, exist_ok=False); return candidate.name, candidate


def _write_artifacts(directory, manifest, lineage, exclusions, issues, *, field_lineage, audit_artifacts):
    (directory / "execution_manifest.json").write_text(json.dumps(manifest, indent=2, default=str), encoding="utf-8")
    if lineage: _write_dict_csv(directory / "lineage.csv", lineage)
    if field_lineage: _write_dict_csv(directory / "field_lineage.csv", field_lineage)
    if exclusions: _write_dict_csv(directory / "exclusions.csv", exclusions)
    if issues: _write_dict_csv(directory / "processing_issues.csv", issues)
    if audit_artifacts: _write_execution_review(directory / "execution_review.xlsx", manifest, exclusions, issues)


def _write_csv(path, columns, rows):
    headers = [item["Output_Column"] for item in columns]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="raise"); writer.writeheader(); writer.writerows([{header: row.get(header) for header in headers} for row in rows])


def _write_dict_csv(path, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0])); writer.writeheader(); writer.writerows(rows)


def _validate_reopened_csv(path, headers, expected, checks):
    with path.open(newline="", encoding="utf-8") as handle: reader = csv.DictReader(handle); rows = list(reader); actual_headers = reader.fieldnames
    checks.extend([_check("REOPENED_CSV_SCHEMA", "PASS" if actual_headers == headers else "FAIL", "Reopened CSV header order matches the approved output contract."), _check("REOPENED_CSV_ROW_COUNT", "PASS" if len(rows) == expected else "FAIL", f"Reopened CSV contains {len(rows)} rows.")])


def _write_execution_review(path, manifest, exclusions, issues):
    sheets = {"Execution_Summary": [{key: value for key, value in manifest.items() if not isinstance(value, (dict, list))}], "Outputs": [{"output": item} for item in manifest.get("outputs_created", [])], "Row_Counts": manifest.get("output_row_counts", []), "Validation_Checks": manifest.get("validation_checks", []), "Exclusions": exclusions, "Issues": issues}
    workbook = Workbook(); active = workbook.active; active.title = "Execution_Summary"
    for name, rows in sheets.items(): _write_review_sheet(active if name == "Execution_Summary" else workbook.create_sheet(name), rows)
    workbook.save(path)


def _write_review_sheet(sheet, rows):
    headers = list(rows[0]) if rows else ["No records"]; sheet.append(headers)
    for row in rows: sheet.append([json.dumps(value, sort_keys=True) if isinstance(value := row.get(header), (dict, list)) else value for header in headers])
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows]; sheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 12), 50)


def _unique_dicts(rows):
    seen, result = set(), []
    for row in rows:
        key = json.dumps(row, sort_keys=True, ensure_ascii=False, default=str)
        if key not in seen: seen.add(key); result.append(row)
    return result


def _read_json(path): return json.loads(Path(path).read_text(encoding="utf-8"))

__all__ = ["EXECUTION_STATUSES", "UnsupportedOperationError", "compare_reconciliation_values", "execute_processing_plan"]
