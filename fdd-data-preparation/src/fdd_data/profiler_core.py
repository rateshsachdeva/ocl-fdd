"""Deterministic, read-only structural profiling of OOXML Excel workbooks."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, is_dataclass
from datetime import date, datetime
from enum import Enum
import json
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from .discovery import discover_source_files, sha256_file
from .lineage import source_field_id
from .models import (
    DataRegionProfile,
    ProfilingCapability,
    ProfilingStatus,
    SourceFile,
    WorkbookProfile,
    WorksheetProfile,
)
from .source_provider import LocalFolderSourceProvider
from .source_data import open_source_workbook, stage_profiled_workbook, write_staging_manifest

MAX_DISTINCT_VALUES = 1_000
MAX_REPRESENTATIVE_VALUES = 5
MAX_REGION_SAMPLES = 8
MAX_SAMPLE_COLUMNS = 25
MAX_CONTEXT_ROWS = 3
MAX_CONTEXT_CELLS = 3
MAX_CONTEXT_BLANK_GAP = 2
MAX_HEADER_SCAN_ROWS = 12
STABLE_RECORD_WINDOW = 4
NUMERIC_TEXT_PATTERN = re.compile(
    r"^[-+]?\s*(?:[$€£¥]\s*)?\(?\s*(?:\d{1,3}(?:[,\s]\d{3})+|\d+)(?:\.\d+)?\s*%?\s*\)?\s*(?:[$€£¥])?$"
)
MONTH_PATTERN = "(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
DATE_TEXT_PATTERNS = (
    re.compile(rf"^{MONTH_PATTERN}[-\s]\d{{2,4}}$", re.IGNORECASE),
    re.compile(rf"^\d{{1,2}}[-\s]{MONTH_PATTERN}[-\s]\d{{2,4}}$", re.IGNORECASE),
    re.compile(r"^\d{4}-\d{1,2}$"),
    re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$"),
    re.compile(rf"^FY\d{{2,4}}\s+{MONTH_PATTERN}$", re.IGNORECASE),
)


@dataclass(slots=True)
class ProfilingRunResult:
    run_id: str
    run_directory: Path
    source_files: list[SourceFile]
    workbook_profiles: list[WorkbookProfile]

    @property
    def fully_profiled_count(self) -> int:
        return sum(profile.profiling_status is ProfilingStatus.PROFILED for profile in self.workbook_profiles)

    @property
    def deferred_count(self) -> int:
        return sum(profile.profiling_status is ProfilingStatus.UNSUPPORTED_DEFERRED for profile in self.workbook_profiles)

    @property
    def failed_count(self) -> int:
        return sum(profile.profiling_status is ProfilingStatus.FAILED for profile in self.workbook_profiles)

    @property
    def worksheet_count(self) -> int:
        return sum(len(profile.worksheet_profiles) for profile in self.workbook_profiles)

    @property
    def region_count(self) -> int:
        return sum(len(worksheet.data_regions) for profile in self.workbook_profiles for worksheet in profile.worksheet_profiles)


def profile_source_file(source_file: SourceFile) -> WorkbookProfile:
    if source_file.profiling_capability is ProfilingCapability.DEFERRED_UNSUPPORTED:
        return WorkbookProfile(
            source_id=source_file.source_id,
            profiling_status=ProfilingStatus.UNSUPPORTED_DEFERRED,
            warnings=[f"Full profiling is deferred for {source_file.extension}; OOXML formats are fully profileable."],
        )
    try:
        workbook = open_source_workbook(source_file.path, read_only=False)
        profile = _profile_open_workbook(source_file, workbook)
        workbook.close()
    except Exception as error:
        return WorkbookProfile(source_id=source_file.source_id, profiling_status=ProfilingStatus.FAILED, errors=[f"{type(error).__name__}: {error}"])
    if sha256_file(source_file.path) != source_file.sha256:
        profile.profiling_status = ProfilingStatus.FAILED
        profile.errors.append("Source SHA-256 changed during profiling; results must be reviewed.")
    return profile


def profile_source_files(source_files: Iterable[SourceFile]) -> list[WorkbookProfile]:
    return [profile_source_file(source_file) for source_file in source_files]


def _profile_open_workbook(source_file: SourceFile, workbook: Workbook) -> WorkbookProfile:
    named_ranges = [{"name": name, "refers_to": getattr(definition, "attr_text", None)} for name, definition in workbook.defined_names.items()]
    worksheets = [_profile_worksheet(source_file, worksheet, index) for index, worksheet in enumerate(workbook.worksheets, start=1)]
    return WorkbookProfile(
        source_id=source_file.source_id,
        worksheet_profiles=worksheets,
        profiling_status=ProfilingStatus.PROFILED,
        worksheet_names=[worksheet.title for worksheet in workbook.worksheets],
        named_ranges=named_ranges,
    )


def _profile_worksheet(source_file: SourceFile, worksheet: Any, index: int) -> WorksheetProfile:
    populated_cells = [cell for cell in worksheet._cells.values() if cell.value is not None]
    meaningful_bounds = _meaningful_bounds(populated_cells)
    row_columns = _row_columns(populated_cells)
    number_formats = Counter(cell.number_format for cell in populated_cells if cell.number_format and cell.number_format != "General")
    indentation_levels = sorted({int(cell.alignment.indent or 0) for cell in populated_cells if (cell.alignment.indent or 0) > 0})
    indentation_examples = [
        {"cell": cell.coordinate, "row": cell.row, "column": get_column_letter(cell.column), "indent": cell.alignment.indent}
        for cell in populated_cells if (cell.alignment.indent or 0) > 0
    ][:MAX_REPRESENTATIVE_VALUES]
    merged_ranges = []
    for merged_range in worksheet.merged_cells.ranges:
        anchor = worksheet.cell(merged_range.min_row, merged_range.min_col)
        merged_ranges.append({"range": str(merged_range), "anchor": anchor.coordinate, "anchor_value": _json_value(anchor.value)})
    hidden_columns = _hidden_columns(worksheet)
    blank_rows, blank_columns = _blank_patterns(meaningful_bounds, row_columns)
    regions = _detect_regions(source_file, worksheet, row_columns)
    return WorksheetProfile(
        source_id=source_file.source_id,
        worksheet_name=worksheet.title,
        worksheet_index=index,
        data_regions=regions,
        visibility=worksheet.sheet_state,
        excel_max_row=worksheet.max_row,
        excel_max_column=worksheet.max_column,
        calculated_dimension=_calculated_dimension(worksheet),
        meaningful_bounds=meaningful_bounds,
        populated_cell_count=len(populated_cells),
        formula_cell_count=sum(cell.data_type == "f" for cell in populated_cells),
        merged_ranges=merged_ranges,
        hidden_rows=sorted(row for row, dimension in worksheet.row_dimensions.items() if dimension.hidden),
        hidden_columns=hidden_columns,
        freeze_panes=str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
        indentation_levels=indentation_levels,
        indentation_examples=indentation_examples,
        number_formats=[name for name, _ in number_formats.most_common(10)],
        blank_row_count=len(blank_rows),
        blank_column_count=len(blank_columns),
        blank_rows=blank_rows,
        blank_columns=[get_column_letter(column) for column in blank_columns],
        observations=_worksheet_observations(meaningful_bounds, blank_rows, blank_columns, populated_cells),
    )


def _meaningful_bounds(cells: list[Cell]) -> str | None:
    if not cells:
        return None
    return f"{get_column_letter(min(cell.column for cell in cells))}{min(cell.row for cell in cells)}:{get_column_letter(max(cell.column for cell in cells))}{max(cell.row for cell in cells)}"


def _row_columns(cells: list[Cell]) -> dict[int, set[int]]:
    result: dict[int, set[int]] = defaultdict(set)
    for cell in cells:
        result[cell.row].add(cell.column)
    return dict(result)


def _calculated_dimension(worksheet: Any) -> str | None:
    try:
        return worksheet.calculate_dimension()
    except (TypeError, ValueError):
        return None


def _hidden_columns(worksheet: Any) -> list[str]:
    hidden: list[str] = []
    for letter, dimension in worksheet.column_dimensions.items():
        if not dimension.hidden:
            continue
        minimum, maximum = dimension.min or 0, dimension.max or 0
        if minimum and maximum:
            hidden.extend(get_column_letter(index) for index in range(minimum, maximum + 1))
        else:
            hidden.append(letter)
    return sorted(set(hidden))


def _blank_patterns(meaningful_bounds: str | None, row_columns: dict[int, set[int]]) -> tuple[list[int], list[int]]:
    if not meaningful_bounds:
        return [], []
    populated_rows = sorted(row_columns)
    min_row, max_row = populated_rows[0], populated_rows[-1]
    populated_columns = {column for columns in row_columns.values() for column in columns}
    min_col, max_col = min(populated_columns), max(populated_columns)
    blank_rows = [row for row in range(min_row, max_row + 1) if row not in row_columns]
    blank_columns = [col for col in range(min_col, max_col + 1) if all(col not in columns for columns in row_columns.values())]
    return blank_rows, blank_columns


def _worksheet_observations(bounds, blank_rows, blank_columns, populated_cells) -> list[str]:
    observations = []
    if bounds is None:
        observations.append("No cells containing values were found.")
    if blank_rows:
        observations.append(f"{len(blank_rows)} blank row(s) occur within the meaningful populated bounds.")
    if blank_columns:
        observations.append(f"{len(blank_columns)} blank column(s) occur within the meaningful populated bounds.")
    indented_count = sum((cell.alignment.indent or 0) > 0 for cell in populated_cells)
    if indented_count:
        observations.append(f"{indented_count} populated cell(s) use non-zero indentation.")
    return observations


def _detect_regions(source_file: SourceFile, worksheet: Any, row_columns: dict[int, set[int]]) -> list[DataRegionProfile]:
    if not row_columns:
        return []
    regions = []
    for row_group in _contiguous_groups(sorted(row_columns)):
        group_columns = sorted({column for row in row_group for column in row_columns[row]})
        for column_group in _contiguous_groups(group_columns):
            region = _build_region(source_file, worksheet, row_group, column_group)
            if region.populated_cell_count:
                regions.append(region)
    return _associate_context_preambles(regions)


def _associate_context_preambles(regions: list[DataRegionProfile]) -> list[DataRegionProfile]:
    remaining = []
    for region in regions:
        if remaining and _is_context_preamble(remaining[-1], region):
            context_region = remaining.pop()
            region.context_preamble_blocks.append(_context_preamble_evidence(context_region))
            region.detection_reason = f"{region.detection_reason} A preceding small context/preamble candidate is associated separately."
        remaining.append(region)
    return remaining


def _is_context_preamble(context_region: DataRegionProfile, table_region: DataRegionProfile) -> bool:
    required = (context_region.start_row, context_region.end_row, context_region.start_column, context_region.column_count, table_region.start_row, table_region.start_column, table_region.row_count, table_region.column_count, table_region.populated_cell_count, table_region.density)
    if any(value is None for value in required):
        return False
    blank_gap = table_region.start_row - context_region.end_row - 1
    values = [value for sample in context_region.samples for value in sample["values"].values()]
    return (
        context_region.row_count <= MAX_CONTEXT_ROWS
        and context_region.column_count == 1
        and context_region.populated_cell_count <= MAX_CONTEXT_CELLS
        and all(isinstance(value, str) for value in values)
        and context_region.start_column == table_region.start_column
        and 1 <= blank_gap <= MAX_CONTEXT_BLANK_GAP
        and table_region.row_count >= 3
        and table_region.column_count >= 3
        and table_region.populated_cell_count >= max(12, context_region.populated_cell_count * 4)
        and table_region.density >= 0.5
    )


def _context_preamble_evidence(context_region: DataRegionProfile) -> dict[str, Any]:
    cells = [{"cell": f"{column}{sample['row']}", "value": value} for sample in context_region.samples for column, value in sample["values"].items()]
    return {"classification": "CONTEXT_PREAMBLE_CANDIDATE", "physical_range": context_region.cell_range, "coordinates": cells, "reason": "Small single-column text block immediately precedes a substantially larger multi-column populated region after a short blank gap."}


def _contiguous_groups(values: list[int]) -> list[list[int]]:
    if not values:
        return []
    groups = [[values[0]]]
    for value in values[1:]:
        if value == groups[-1][-1] + 1:
            groups[-1].append(value)
        else:
            groups.append([value])
    return groups


def _build_region(source_file: SourceFile, worksheet: Any, row_group: list[int], column_group: list[int]) -> DataRegionProfile:
    start_row, end_row = row_group[0], row_group[-1]
    start_column, end_column = column_group[0], column_group[-1]
    cells = [cell for cell in worksheet._cells.values() if cell.value is not None and start_row <= cell.row <= end_row and start_column <= cell.column <= end_column]
    row_count, column_count = end_row - start_row + 1, end_column - start_column + 1
    header_candidates = _header_candidates(worksheet, start_row, end_row, start_column, end_column)
    if not header_candidates and row_count >= 2:
        values = [worksheet.cell(start_row, column).value for column in range(start_column, end_column + 1)]
        present = [value for value in values if value is not None]
        if len(present) >= 2 and all(isinstance(value, str) for value in present):
            header_candidates = [{"row": start_row, "confidence": "PRIMARY", "reason": "Leading text row is the only structurally viable header in a small table."}]
    selected_header_row = max(candidate["row"] for candidate in header_candidates) if header_candidates else None
    repeated_header_rows = _repeated_header_rows(worksheet, selected_header_row, start_row, end_row, start_column, end_column)
    footer_candidates = _footer_candidates(worksheet, start_row, end_row, start_column, end_column)
    footer_rows = {candidate["row"] for candidate in footer_candidates}
    data_rows = [row for row in range((selected_header_row or start_row - 1) + 1, end_row + 1) if row not in footer_rows and row not in repeated_header_rows and _row_has_values(worksheet, row, start_column, end_column)]
    preamble_candidates = [{"row": row, "reason": "Populated row occurs before the first candidate header."} for row in range(start_row, selected_header_row or start_row) if _row_has_values(worksheet, row, start_column, end_column)]
    trailing_note_rows = _trailing_note_rows(worksheet, data_rows, end_row, start_column, end_column)
    region_id = f"{source_file.source_id}:{worksheet.title}:{start_row}-{end_row}:{start_column}-{end_column}"
    return DataRegionProfile(
        region_id=region_id,
        source_id=source_file.source_id,
        worksheet_name=worksheet.title,
        cell_range=f"{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}",
        populated_cell_count=len(cells),
        start_row=start_row, end_row=end_row, start_column=start_column, end_column=end_column,
        row_count=row_count, column_count=column_count,
        density=round(len(cells) / (row_count * column_count), 4),
        detection_reason="Populated rows and columns bounded by blank separator rows or columns.",
        candidate_confidence="TABULAR_CANDIDATE" if row_count >= 2 and column_count >= 2 else "STRUCTURAL_CANDIDATE",
        preamble_candidates=preamble_candidates,
        header_candidates=header_candidates,
        footer_candidates=footer_candidates,
        likely_data_row_ranges=_row_ranges(data_rows),
        repeated_header_rows=repeated_header_rows,
        trailing_note_rows=trailing_note_rows,
        column_profiles=_column_profiles(source_file, worksheet, region_id, data_rows, start_column, end_column, selected_header_row),
        samples=_region_samples(worksheet, start_row, end_row, start_column, end_column, selected_header_row, footer_rows, repeated_header_rows),
    )


def _row_has_values(worksheet: Any, row: int, start_column: int, end_column: int) -> bool:
    return any(worksheet.cell(row, column).value is not None for column in range(start_column, end_column + 1))


def _header_candidates(worksheet: Any, start_row: int, end_row: int, start_column: int, end_column: int) -> list[dict[str, Any]]:
    if end_row - start_row < 2:
        return []
    candidates = []
    for row in range(start_row, min(end_row - 2, start_row + MAX_HEADER_SCAN_ROWS - 1) + 1):
        metrics = _row_metrics(worksheet, row, start_column, end_column)
        stable_rows = _following_stable_rows(worksheet, row + 1, end_row, start_column, end_column)
        if not _is_header_like(metrics, end_column - start_column + 1) or len(stable_rows) < 3:
            continue
        stability = _record_stability(stable_rows)
        if stability < 0.65:
            continue
        type_difference = _signature_difference(metrics["signature"], stable_rows[0]["signature"])
        population_difference = abs(metrics["populated"] - stable_rows[0]["populated"]) / max(1, end_column - start_column + 1)
        score = (stability * 4) + (type_difference * 2) + population_difference + (0.5 if row == start_row else -(row - start_row) * 0.05)
        candidates.append({"row": row, "score": round(score, 3), "reason": f"Header-like row precedes a stable following record pattern; type-pattern difference={type_difference:.2f}, stability={stability:.2f}."})
    if not candidates:
        return []
    primary = max(candidates, key=lambda candidate: (candidate["score"], -candidate["row"]))
    primary["confidence"] = "PRIMARY"
    result = [primary]
    preceding_row = primary["row"] - 1
    if preceding_row >= start_row:
        preceding_metrics = _row_metrics(worksheet, preceding_row, start_column, end_column)
        following_rows = _following_stable_rows(worksheet, primary["row"] + 1, end_row, start_column, end_column)
        if _is_dense_header_like(preceding_metrics, end_column - start_column + 1) and following_rows and _signature_difference(preceding_metrics["signature"], following_rows[0]["signature"]) >= 0.25:
            result.append({"row": preceding_row, "confidence": "SECONDARY", "reason": "Dense header-like row immediately precedes the primary header transition."})
    return sorted(result, key=lambda candidate: candidate["row"])


def _row_metrics(worksheet: Any, row: int, start_column: int, end_column: int) -> dict[str, Any]:
    values = [worksheet.cell(row, column).value for column in range(start_column, end_column + 1)]
    present = [value for value in values if value is not None]
    text_count = sum(isinstance(value, str) and not value.startswith("=") for value in present)
    return {"populated": len(present), "text_ratio": text_count / len(present) if present else 0.0, "signature": tuple(_primitive_characteristic(worksheet.cell(row, column)) if worksheet.cell(row, column).value is not None else "EMPTY" for column in range(start_column, end_column + 1))}


def _following_stable_rows(worksheet: Any, start_row: int, end_row: int, start_column: int, end_column: int) -> list[dict[str, Any]]:
    rows = []
    for row in range(start_row, end_row + 1):
        metrics = _row_metrics(worksheet, row, start_column, end_column)
        if metrics["populated"]:
            rows.append(metrics)
        if len(rows) == STABLE_RECORD_WINDOW:
            break
    return rows


def _is_header_like(metrics: dict[str, Any], column_count: int) -> bool:
    return metrics["populated"] >= 2 and metrics["text_ratio"] >= 0.5 and metrics["populated"] >= column_count / 2


def _is_dense_header_like(metrics: dict[str, Any], column_count: int) -> bool:
    return _is_header_like(metrics, column_count) and metrics["populated"] / column_count >= 0.75


def _record_stability(rows: list[dict[str, Any]]) -> float:
    if len(rows) < 2:
        return 0.0
    reference = rows[0]
    similarities = [(_signature_similarity(reference["signature"], row["signature"]) + (1 - abs(reference["populated"] - row["populated"]) / max(1, len(reference["signature"])))) / 2 for row in rows[1:]]
    return sum(similarities) / len(similarities)


def _signature_similarity(left: tuple[str, ...], right: tuple[str, ...]) -> float:
    return sum(left_value == right_value for left_value, right_value in zip(left, right)) / max(1, len(left))


def _signature_difference(left: tuple[str, ...], right: tuple[str, ...]) -> float:
    return 1 - _signature_similarity(left, right)


def _repeated_header_rows(worksheet: Any, header_row: int | None, start_row: int, end_row: int, start_column: int, end_column: int) -> list[int]:
    if header_row is None:
        return []
    signature = tuple(_normalise_text(worksheet.cell(header_row, column).value) for column in range(start_column, end_column + 1))
    if not any(signature):
        return []
    return [row for row in range(header_row + 1, end_row + 1) if tuple(_normalise_text(worksheet.cell(row, column).value) for column in range(start_column, end_column + 1)) == signature]


def _normalise_text(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    return " ".join(value.strip().casefold().split()) or None


def _footer_candidates(worksheet: Any, start_row: int, end_row: int, start_column: int, end_column: int) -> list[dict[str, Any]]:
    candidates = []
    for row in range(start_row, end_row + 1):
        values = [worksheet.cell(row, column).value for column in range(start_column, end_column + 1)]
        strings = [value for value in values if isinstance(value, str)]
        formula_count = sum(worksheet.cell(row, column).data_type == "f" for column in range(start_column, end_column + 1))
        if any("total" in value.casefold() for value in strings):
            candidates.append({"row": row, "reason": "Row contains a total or subtotal label."})
        elif formula_count and row >= end_row - 1:
            candidates.append({"row": row, "reason": "Trailing row contains formula cells and may be a structural footer."})
    return candidates


def _trailing_note_rows(worksheet: Any, data_rows: list[int], end_row: int, start_column: int, end_column: int) -> list[int]:
    if not data_rows:
        return []
    last_data_row = max(data_rows)
    notes = []
    for row in range(last_data_row + 1, end_row + 1):
        present = [worksheet.cell(row, column).value for column in range(start_column, end_column + 1) if worksheet.cell(row, column).value is not None]
        if present and len(present) <= 2 and all(isinstance(value, str) for value in present):
            notes.append(row)
    return notes


def _row_ranges(rows: list[int]) -> list[dict[str, int]]:
    return [{"start_row": group[0], "end_row": group[-1]} for group in _contiguous_groups(rows)]


def _column_profiles(source_file: SourceFile, worksheet: Any, region_id: str, data_rows: list[int], start_column: int, end_column: int, header_row: int | None) -> list[dict[str, Any]]:
    profiles = []
    for column in range(start_column, end_column + 1):
        header = _json_value(worksheet.cell(header_row, column).value) if header_row else None
        cells = [worksheet.cell(row, column) for row in data_rows]
        populated = [cell for cell in cells if cell.value is not None]
        distinct_values: set[str] = set()
        capped = False
        representatives = []
        characteristics: Counter[str] = Counter()
        whitespace_examples, numeric_text_examples, date_text_examples = [], [], []
        number_formats: Counter[str] = Counter()
        formula_count = indentation_count = 0
        for cell in populated:
            value = cell.value
            if cell.data_type == "f": formula_count += 1
            if (cell.alignment.indent or 0) > 0: indentation_count += 1
            if cell.number_format and cell.number_format != "General": number_formats[cell.number_format] += 1
            characteristics[_primitive_characteristic(cell)] += 1
            key = repr(value)
            if len(distinct_values) < MAX_DISTINCT_VALUES: distinct_values.add(key)
            else: capped = True
            serialised = _json_value(value)
            if serialised not in representatives and len(representatives) < MAX_REPRESENTATIVE_VALUES: representatives.append(serialised)
            if isinstance(value, str):
                stripped = value.strip()
                if value != stripped or "  " in value: _append_bounded(whitespace_examples, value)
                if _is_numeric_text(stripped): _append_bounded(numeric_text_examples, value)
                if _is_date_text(stripped): _append_bounded(date_text_examples, value)
        profiles.append({
            "field_id": source_field_id(source_file.source_id, worksheet.title, region_id, column),
            "physical_column": column,
            "exact_original_header": header,
            "normalized_display_header": _normalise_display_header(header),
            "column_index": column,
            "column_letter": get_column_letter(column),
            "populated_count": len(populated),
            "blank_count": len(cells) - len(populated),
            "distinct_count": len(distinct_values),
            "distinct_count_capped": capped,
            "primitive_characteristic": _overall_characteristic(characteristics),
            "representative_values": representatives,
            "whitespace_anomaly_examples": whitespace_examples,
            "numeric_text_examples": numeric_text_examples,
            "date_text_examples": date_text_examples,
            "formula_presence": formula_count > 0,
            "formula_count": formula_count,
            "indentation_presence": indentation_count > 0,
            "common_number_formats": [name for name, _ in number_formats.most_common(5)],
        })
    return profiles


def _normalise_display_header(value: Any) -> str | None:
    if value is None: return None
    return " ".join(str(value).strip().split())


def _append_bounded(values: list[str], value: str) -> None:
    if value not in values and len(values) < MAX_REPRESENTATIVE_VALUES: values.append(value)


def _primitive_characteristic(cell: Cell) -> str:
    value = cell.value
    if cell.data_type == "f": return "FORMULA"
    if isinstance(value, bool): return "BOOLEAN_LIKE"
    if isinstance(value, (datetime, date)) or cell.is_date: return "DATE_LIKE"
    if isinstance(value, (int, float)): return "NUMERIC_LIKE"
    if isinstance(value, str): return "TEXT_LIKE"
    return "MIXED"


def _overall_characteristic(characteristics: Counter[str]) -> str:
    if not characteristics: return "EMPTY"
    if len(characteristics) == 1: return next(iter(characteristics))
    non_formula = {name for name in characteristics if name != "FORMULA"}
    if len(non_formula) == 1 and non_formula: return next(iter(non_formula))
    return "MIXED"


def _is_numeric_text(value: str) -> bool:
    return bool(NUMERIC_TEXT_PATTERN.fullmatch(value))


def _is_date_text(value: str) -> bool:
    if re.fullmatch(r"\d{6}", value): return 1 <= int(value[4:]) <= 12
    return any(pattern.fullmatch(value) for pattern in DATE_TEXT_PATTERNS)


def _region_samples(worksheet: Any, start_row: int, end_row: int, start_column: int, end_column: int, header_row: int | None, footer_rows: set[int], repeated_header_rows: list[int]) -> list[dict[str, Any]]:
    candidate_rows = {start_row, end_row, (start_row + end_row) // 2}
    if header_row is not None: candidate_rows.add(header_row)
    candidate_rows.update(footer_rows); candidate_rows.update(repeated_header_rows)
    previous_signature = None
    for row in range(start_row, end_row + 1):
        signature = tuple(_primitive_characteristic(worksheet.cell(row, column)) if worksheet.cell(row, column).value is not None else "EMPTY" for column in range(start_column, end_column + 1))
        if previous_signature is not None and signature != previous_signature: candidate_rows.add(row)
        previous_signature = signature
    samples = []
    for row in sorted(candidate_rows)[:MAX_REGION_SAMPLES]:
        values, signals = {}, {}
        for column in range(start_column, min(end_column, start_column + MAX_SAMPLE_COLUMNS - 1) + 1):
            cell = worksheet.cell(row, column)
            if cell.value is None: continue
            letter = get_column_letter(column)
            values[letter] = _json_value(cell.value)
            cell_signals = {}
            if (cell.alignment.indent or 0) > 0: cell_signals["indent"] = cell.alignment.indent
            if cell.number_format and cell.number_format != "General": cell_signals["number_format"] = cell.number_format
            if cell_signals: signals[letter] = cell_signals
        samples.append({"row": row, "values": values, "signals": signals})
    return samples


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)): return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None: return value
    return str(value)


def _json_ready(value: Any) -> Any:
    if is_dataclass(value): return _json_ready(asdict(value))
    if isinstance(value, Enum): return value.value
    if isinstance(value, Path): return str(value)
    if isinstance(value, (datetime, date)): return value.isoformat()
    if isinstance(value, dict): return {str(key): _json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)): return [_json_ready(item) for item in value]
    return value


def profile_directory(source_directory: Path, runs_directory: Path = Path("work/runs"), *, generate_review: bool = False, staging_directory: Path | None = None) -> ProfilingRunResult:
    source_files = discover_source_files(LocalFolderSourceProvider(source_directory))
    run_id, run_directory = _create_run_directory(runs_directory)
    profiles = []
    staging_entries = {}
    for source_file in source_files:
        if staging_directory is None or source_file.profiling_capability is ProfilingCapability.DEFERRED_UNSUPPORTED:
            profiles.append(profile_source_file(source_file)); continue
        try:
            workbook = open_source_workbook(source_file.path, read_only=False)
            profile = _profile_open_workbook(source_file, workbook)
            staging_entries.update(stage_profiled_workbook(_json_ready(source_file), workbook, _json_ready(profile), staging_directory))
            workbook.close()
        except Exception as error:
            profile = WorkbookProfile(source_id=source_file.source_id, profiling_status=ProfilingStatus.FAILED, errors=[f"{type(error).__name__}: {error}"])
        if sha256_file(source_file.path) != source_file.sha256:
            profile.profiling_status = ProfilingStatus.FAILED
            profile.errors.append("Source SHA-256 changed during profiling; results must be reviewed.")
        profiles.append(profile)
    result = ProfilingRunResult(run_id, run_directory, source_files, profiles)
    _write_run_artifacts(result, generate_review=generate_review)
    if staging_directory is not None and not result.failed_count:
        profile_document = json.loads((run_directory / "profile.json").read_text(encoding="utf-8"))
        write_staging_manifest(profile_document, staging_directory, staging_entries)
    return result


def _create_run_directory(runs_directory: Path) -> tuple[str, Path]:
    base_identifier = f"PROFILE_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    candidate = Path(runs_directory) / base_identifier
    suffix = 1
    while candidate.exists():
        suffix += 1
        candidate = Path(runs_directory) / f"{base_identifier}_{suffix:02d}"
    candidate.mkdir(parents=True, exist_ok=False)
    return candidate.name, candidate


def _write_run_artifacts(result: ProfilingRunResult, *, generate_review: bool) -> None:
    by_source_id = {profile.source_id: profile for profile in result.workbook_profiles}
    manifest_rows = []
    for source_file in result.source_files:
        profile = by_source_id[source_file.source_id]
        manifest_rows.append({"source_id": source_file.source_id, "filename": source_file.filename, "path": str(source_file.path), "extension": source_file.extension, "size": source_file.size, "modified_time": source_file.modified_time.isoformat(), "sha256": source_file.sha256, "profiling_capability": source_file.profiling_capability.value, "profiling_status": profile.profiling_status.value, "warnings": " | ".join(profile.warnings), "errors": " | ".join(profile.errors)})
    _write_csv(result.run_directory / "source_manifest.csv", manifest_rows)
    profile_document = {"run_id": result.run_id, "source_files": result.source_files, "workbook_profiles": result.workbook_profiles}
    (result.run_directory / "profile.json").write_text(json.dumps(_json_ready(profile_document), indent=2, ensure_ascii=False), encoding="utf-8")
    if generate_review: _write_review_workbook(result, manifest_rows)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    import csv
    headers = list(rows[0]) if rows else ["source_id", "filename", "path", "extension", "size", "modified_time", "sha256", "profiling_capability", "profiling_status", "warnings", "errors"]
    with path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers); writer.writeheader(); writer.writerows(rows)


def _write_review_workbook(result: ProfilingRunResult, manifest_rows: list[dict[str, Any]]) -> None:
    workbook = Workbook(); files_sheet = workbook.active; files_sheet.title = "Files"
    worksheet_rows, region_rows, column_rows = [], [], []
    source_lookup = {source.source_id: source for source in result.source_files}
    for profile in result.workbook_profiles:
        source = source_lookup[profile.source_id]
        for worksheet in profile.worksheet_profiles:
            worksheet_rows.append({"filename": source.filename, "worksheet": worksheet.worksheet_name, "order": worksheet.worksheet_index, "visibility": worksheet.visibility, "meaningful_bounds": worksheet.meaningful_bounds, "populated_cells": worksheet.populated_cell_count, "formula_cells": worksheet.formula_cell_count, "excel_max_row": worksheet.excel_max_row, "excel_max_column": worksheet.excel_max_column, "merged_range_count": len(worksheet.merged_ranges), "hidden_row_count": len(worksheet.hidden_rows), "hidden_column_count": len(worksheet.hidden_columns), "freeze_panes": worksheet.freeze_panes})
            for region in worksheet.data_regions:
                region_rows.append({"filename": source.filename, "worksheet": worksheet.worksheet_name, "region": region.cell_range, "populated_cells": region.populated_cell_count, "density": region.density, "reason": region.detection_reason, "header_rows": ", ".join(str(item["row"]) for item in region.header_candidates), "footer_rows": ", ".join(str(item["row"]) for item in region.footer_candidates), "repeated_header_rows": ", ".join(map(str, region.repeated_header_rows)), "context_preamble_block_count": len(region.context_preamble_blocks)})
                for column in region.column_profiles:
                    column_rows.append({"field_id": column["field_id"], "filename": source.filename, "worksheet": worksheet.worksheet_name, "region": region.cell_range, "column": column["column_letter"], "exact_original_header": column["exact_original_header"], "populated_count": column["populated_count"], "blank_count": column["blank_count"], "distinct_count": column["distinct_count"], "primitive_characteristic": column["primitive_characteristic"], "formula_presence": column["formula_presence"], "numeric_text_detected": bool(column["numeric_text_examples"]), "date_text_detected": bool(column["date_text_examples"]), "common_number_formats": " | ".join(column["common_number_formats"])})
    _write_review_sheet(files_sheet, manifest_rows)
    _write_review_sheet(workbook.create_sheet("Worksheets"), worksheet_rows)
    _write_review_sheet(workbook.create_sheet("Regions"), region_rows)
    _write_review_sheet(workbook.create_sheet("Columns"), column_rows)
    workbook.save(result.run_directory / "profile_review.xlsx")


def _write_review_sheet(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0]) if rows else ["No records"]
    worksheet.append(headers)
    for row in rows: worksheet.append([row.get(header) for header in headers])
    worksheet.freeze_panes = "A2"; worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows]
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 12), 40)


__all__ = ["ProfilingRunResult", "profile_directory", "profile_source_file", "profile_source_files"]
