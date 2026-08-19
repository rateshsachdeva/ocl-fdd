"""Robust wrapper around the integrated data-preparation fast path.

The fast path handles explicit annual/monthly/amount/period layouts. This module
adds a structural fallback for common FDD schedules where one worksheet is one
reporting period (for example TB_FY2023 / TB_FY2024 / TB_FY2025) and the period
is therefore not repeated in every data row.

The fallback is still deterministic, source-read-only and lineage preserving.
It is deliberately conservative and writes diagnostics when it cannot classify
anything rather than asking the user to reshape the client file.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .integrated import DataPrepResult, prepare_source_package as _fast_prepare

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def prepare_source_package(source_dir: Path, output_dir: Path) -> DataPrepResult:
    """Run the normal parser first, then a structural one-period-per-sheet fallback."""
    try:
        return _fast_prepare(source_dir, output_dir)
    except ValueError as error:
        if "No usable OCL-related structured dataset" not in str(error):
            raise
    return _prepare_structural_fallback(Path(source_dir), Path(output_dir))


def _prepare_structural_fallback(source_dir: Path, output_dir: Path) -> DataPrepResult:
    source_dir = source_dir.resolve()
    output_dir = output_dir.resolve()
    files = [
        path for path in sorted(source_dir.rglob("*"))
        if path.is_file() and not path.name.startswith("~$") and path.suffix.lower() in EXCEL_EXTENSIONS
    ]
    if not files:
        raise ValueError("No supported Excel source files were found in references/source/.")

    output_dir.mkdir(parents=True, exist_ok=True)
    for old in output_dir.iterdir():
        if old.is_file():
            old.unlink()

    annual_rows: list[dict[str, Any]] = []
    lineage: list[dict[str, str]] = []
    diagnostics: list[dict[str, Any]] = []
    snapshot: list[dict[str, Any]] = []

    for path in files:
        before = _sha256(path)
        snapshot.append({
            "file": path.name,
            "relative_path": path.relative_to(source_dir).as_posix(),
            "sha256": before,
            "size": path.stat().st_size,
        })
        wb = load_workbook(path, data_only=False, read_only=False)
        try:
            for ws in wb.worksheets:
                profile = _profile_sheet(ws)
                profile.update({"file": path.name, "worksheet": ws.title})
                diagnostics.append(profile)
                if not profile.get("accepted"):
                    continue
                period = str(profile["period"])
                header_row = int(profile["header_row"])
                label_col = int(profile["label_col"])
                amount_col = int(profile["amount_col"])
                code_col = profile.get("code_col")
                entity_col = profile.get("entity_col")
                currency_col = profile.get("currency_col")
                for row in range(header_row + 1, ws.max_row + 1):
                    label = _text(ws.cell(row, label_col).value)
                    amount = _numeric(ws.cell(row, amount_col).value)
                    if not label or amount is None:
                        continue
                    source_record_id = json.dumps({
                        "source_id": path.name,
                        "worksheet_name": ws.title,
                        "source_cell": f"{get_column_letter(amount_col)}{row}",
                        "row": row,
                    }, separators=(",", ":"))
                    record = {
                        "Source_Record_ID": source_record_id,
                        "Period": period,
                        "Source_Label": label,
                        "Source_Code": _text(ws.cell(row, int(code_col)).value) if code_col else "",
                        "Entity": _text(ws.cell(row, int(entity_col)).value) if entity_col else "",
                        "Currency": _text(ws.cell(row, int(currency_col)).value) if currency_col else "",
                        "Amount": amount,
                    }
                    annual_rows.append(record)
                    lineage.append({
                        "Output_File": "ocl_annual.csv",
                        "Source_Record_ID": source_record_id,
                        "Source_File": path.name,
                        "Source_Sheet": ws.title,
                        "Source_Cell": f"{get_column_letter(amount_col)}{row}",
                    })
        finally:
            wb.close()
        if _sha256(path) != before:
            raise RuntimeError(f"Source workbook changed while being prepared: {path.name}")

    diagnostic_path = output_dir.parent / "source_diagnostic.json"
    diagnostic_path.parent.mkdir(parents=True, exist_ok=True)
    diagnostic_path.write_text(json.dumps({"worksheets": diagnostics}, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    if not annual_rows:
        detail = "; ".join(
            f"{item['file']}::{item['worksheet']} ({item.get('reason','not accepted')})"
            for item in diagnostics[:20]
        )
        raise ValueError(
            "No usable structured dataset could be prepared after the structural fallback. "
            f"Diagnostic written to {diagnostic_path}. Sheets reviewed: {detail}"
        )

    dataset = output_dir / "ocl_annual.csv"
    columns = ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Currency", "Amount"]
    _write_csv(dataset, columns, annual_rows)
    _write_csv(output_dir / "lineage.csv", ["Output_File", "Source_Record_ID", "Source_File", "Source_Sheet", "Source_Cell"], lineage)
    _write_csv(output_dir / "field_lineage.csv", ["Output_Field", "Meaning"], [
        {"Output_Field": "Source_Record_ID", "Meaning": "JSON pointer to original workbook, worksheet, amount cell and row."},
        {"Output_Field": "Period", "Meaning": "Reporting period derived deterministically from the source worksheet title."},
        {"Output_Field": "Source_Label", "Meaning": "Source line-item/account description selected from the structural header profile."},
        {"Output_Field": "Source_Code", "Meaning": "Source account/code where identified."},
        {"Output_Field": "Amount", "Meaning": "Source amount from the numerically dense amount column; no balancing plug."},
    ])

    execution_id = "FALLBACK_" + hashlib.sha256(json.dumps(snapshot, sort_keys=True).encode("utf-8")).hexdigest()[:16].upper()
    warnings = (
        "The explicit-header fast path did not classify the source. A deterministic one-period-per-sheet structural fallback was used; review Mapping and Checks in the final databook.",
    )
    metadata = {
        "workflow_run_id": execution_id,
        "logical_datasets": [{
            "file": "ocl_annual.csv",
            "purpose": "period-end source schedule extracted from one-period-per-sheet workbook structure",
            "row_count": len(annual_rows),
            "columns": columns,
        }],
        "source_snapshot": snapshot,
        "worksheet_profiles": diagnostics,
        "warnings": list(warnings),
        "integration_mode": "STRUCTURAL_ONE_PERIOD_PER_SHEET_FALLBACK",
    }
    metadata_path = output_dir / "databook_metadata.json"
    metadata_path.write_text(json.dumps(metadata, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    manifest = {
        "execution_id": execution_id,
        "final_execution_status": "COMPLETED_WITH_WARNINGS",
        "outputs_created": ["ocl_annual.csv"],
        "source_file_count": len(files),
        "source_snapshot": snapshot,
        "fallback_used": True,
    }
    manifest_path = output_dir / "execution_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return DataPrepResult(output_dir, manifest_path, metadata_path, (dataset,), warnings)


def _profile_sheet(ws) -> dict[str, Any]:
    period = _period_from_title(ws.title)
    if not period:
        return {"accepted": False, "reason": "worksheet title does not contain a recognizable FY/year period"}

    header_row = _best_header_row(ws)
    if not header_row:
        return {"accepted": False, "reason": "no plausible header row found in first 30 rows", "period": period}

    stats = []
    for col in range(1, ws.max_column + 1):
        header = _text(ws.cell(header_row, col).value)
        values = [ws.cell(row, col).value for row in range(header_row + 1, min(ws.max_row, header_row + 250) + 1)]
        nonblank = [value for value in values if value not in (None, "")]
        numeric = sum(_numeric(value) is not None for value in nonblank)
        text = sum(bool(_text(value)) and _numeric(value) is None for value in nonblank)
        stats.append({
            "col": col,
            "header": header,
            "norm": _norm(header),
            "nonblank": len(nonblank),
            "numeric": numeric,
            "text": text,
        })

    label_col = _explicit_column(stats, ("description", "account name", "account description", "long text", "line item", "name", "label"))
    code_col = _explicit_column(stats, ("gl account", "g l account", "account code", "account number", "gl code", "code"))
    entity_col = _explicit_column(stats, ("entity", "company", "business unit", "legal entity"))
    currency_col = _explicit_column(stats, ("currency", "curr"))
    amount_col = _explicit_column(stats, ("amount", "balance", "ending balance", "closing balance", "local currency amount", "lc amount", "value"))

    if not label_col:
        candidates = [item for item in stats if item["text"] >= 3 and item["text"] >= item["numeric"]]
        if candidates:
            label_col = max(candidates, key=lambda item: (item["text"], len(item["header"]), item["col"]))["col"]

    if not amount_col:
        excluded = {value for value in (label_col, code_col, entity_col, currency_col) if value}
        candidates = [
            item for item in stats
            if item["col"] not in excluded and item["numeric"] >= 3 and item["numeric"] >= max(1, int(item["nonblank"] * 0.5))
        ]
        if candidates:
            # Prefer the right-most numerically dense column, which is common in TB/export schedules.
            amount_col = max(candidates, key=lambda item: (item["numeric"] / max(item["nonblank"], 1), item["col"]))["col"]

    if not label_col or not amount_col or label_col == amount_col:
        return {
            "accepted": False,
            "reason": "could not identify distinct description and amount columns",
            "period": period,
            "header_row": header_row,
            "columns": stats,
        }

    usable_rows = 0
    for row in range(header_row + 1, ws.max_row + 1):
        if _text(ws.cell(row, label_col).value) and _numeric(ws.cell(row, amount_col).value) is not None:
            usable_rows += 1
    if usable_rows < 2:
        return {
            "accepted": False,
            "reason": "identified columns but fewer than two usable data rows",
            "period": period,
            "header_row": header_row,
            "label_col": label_col,
            "amount_col": amount_col,
        }

    return {
        "accepted": True,
        "period": period,
        "header_row": header_row,
        "label_col": label_col,
        "amount_col": amount_col,
        "code_col": code_col,
        "entity_col": entity_col,
        "currency_col": currency_col,
        "usable_rows": usable_rows,
        "columns": stats,
    }


def _best_header_row(ws) -> int | None:
    best: tuple[int, int] | None = None
    for row in range(1, min(ws.max_row, 30) + 1):
        present = [_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        present = [value for value in present if value]
        if len(present) < 2:
            continue
        text_count = sum(_numeric(value) is None for value in present)
        score = text_count * 3 + min(len(present), 10) - row // 10
        candidate = (score, -row)
        if best is None or candidate > best:
            best = candidate
    return -best[1] if best else None


def _explicit_column(stats: list[dict[str, Any]], hints: tuple[str, ...]) -> int | None:
    normalized_hints = tuple(_norm(value) for value in hints)
    exact = [item for item in stats if item["norm"] in normalized_hints]
    if exact:
        return exact[0]["col"]
    contains = [item for item in stats if item["norm"] and any(hint in item["norm"] for hint in normalized_hints)]
    return contains[0]["col"] if contains else None


def _period_from_title(title: str) -> str | None:
    text = str(title).upper()
    match = re.search(r"FY\s*[-_ ]?\s*(20\d{2}|\d{2})", text)
    if match:
        value = match.group(1)
        return "FY" + (value if len(value) == 4 else "20" + value)
    match = re.search(r"(?<!\d)(20\d{2})(?!\d)", text)
    return match.group(1) if match else None


def _numeric(value: Any) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        negative = text.startswith("(") and text.endswith(")")
        if negative:
            text = text[1:-1]
        try:
            number = float(text)
            return -number if negative else number
        except ValueError:
            return None
    return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _text(value).casefold()).strip()


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
