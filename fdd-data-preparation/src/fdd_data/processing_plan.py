"""Vendor-neutral processing-plan validation, approval, and review utilities.

This module intentionally does not execute processing plans or create final flat files.
"""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .dataset_map import CONFIDENCES, DATASET_ROLES, _source_index
from .discovery import discover_source_files
from .lineage import profile_field_index
from .source_provider import LocalFolderSourceProvider

PLAN_STATUSES = {"PLAN_DRAFTED", "AWAITING_APPROVAL", "APPROVED", "SUPERSEDED"}
SOURCE_TYPES = {
    "DIRECT_COLUMN", "FILENAME_METADATA", "SHEET_METADATA", "CONTEXT_METADATA",
    "CONSTANT", "DERIVED", "SUPPORTING_JOIN",
}
JOIN_TYPES = {"ENRICHMENT_JOIN", "RECONCILIATION_RELATIONSHIP"}
CARDINALITIES = {"ONE_TO_ONE", "ONE_TO_MANY", "MANY_TO_ONE", "MANY_TO_MANY", "UNKNOWN"}
RECONCILIATION_COMPARISON_METHODS = {"EXACT", "ROUND_TO_PRECISION", "ABSOLUTE_TOLERANCE"}


class ProcessingPlanValidationError(ValueError):
    """Raised when a plan is structurally invalid."""


class ApprovalError(ValueError):
    """Raised when approval inputs do not exactly match the plan."""


class ExecutionNotApprovedError(PermissionError):
    """Raised by the Stage 4 execution guard."""


class SourceIntegrityError(PermissionError):
    """Raised when the current package differs from the plan's bound snapshot."""


def build_source_snapshot(profile: dict[str, Any], source_directory: Path) -> dict[str, Any]:
    """Bind the discovered/profiled source package to a plan without reading cell values."""
    root = Path(source_directory).resolve()
    status_by_source = {
        workbook["source_id"]: workbook.get("profiling_status", "DISCOVERED")
        for workbook in profile.get("workbook_profiles", [])
    }
    source_files = []
    for source in profile.get("source_files", []):
        source_path = Path(source["path"]).resolve()
        try:
            relative_path = source_path.relative_to(root).as_posix()
        except ValueError as error:
            raise ProcessingPlanValidationError(
                f"Profile source is outside the bound source directory: {source_path}"
            ) from error
        source_files.append({
            "source_id": source["source_id"],
            "filename": source["filename"],
            "relative_path": relative_path,
            "extension": source["extension"],
            "sha256": source["sha256"],
            "profiling_status": status_by_source.get(source["source_id"], "DISCOVERED"),
        })
    return {
        "expected_source_file_count": len(source_files),
        "source_files": sorted(source_files, key=lambda item: item["relative_path"]),
    }


def compare_source_snapshot(source_snapshot: dict[str, Any], source_directory: Path) -> dict[str, Any]:
    """Compare a bound source package to the current local package using SHA-256."""
    root = Path(source_directory).resolve()
    expected = {item["relative_path"]: item for item in source_snapshot.get("source_files", [])}
    current_files = discover_source_files(LocalFolderSourceProvider(root))
    current = {
        path.relative_to(root).as_posix(): source
        for source in current_files
        for path in [source.path.resolve()]
    }
    findings: list[dict[str, str]] = []
    for relative_path in sorted(set(expected) - set(current)):
        findings.append({"reason": "SOURCE_FILE_DELETED", "file": relative_path})
    for relative_path in sorted(set(current) - set(expected)):
        findings.append({"reason": "SOURCE_FILE_ADDED", "file": relative_path})
    for relative_path in sorted(set(expected) & set(current)):
        snapshot_file = expected[relative_path]
        current_file = current[relative_path]
        if snapshot_file["sha256"] != current_file.sha256:
            findings.append({
                "reason": "SOURCE_FILE_REPLACED",
                "file": relative_path,
                "expected_sha256": snapshot_file["sha256"],
                "current_sha256": current_file.sha256,
            })
    expected_count = source_snapshot.get("expected_source_file_count")
    if expected_count != len(expected):
        findings.append({"reason": "SOURCE_SNAPSHOT_INVALID", "file": "source_snapshot"})
    if not findings and expected_count == len(current):
        return {
            "status": "SOURCE_UNCHANGED",
            "comparison_result": "SOURCE_UNCHANGED",
            "findings": [],
            "expected_source_file_count": expected_count,
            "current_source_file_count": len(current),
        }
    primary_reason = findings[0]["reason"] if findings else "SOURCE_FILE_CHANGED"
    return {
        "status": "STALE_SOURCE",
        "comparison_result": "SOURCE_FILE_CHANGED",
        "primary_reason": primary_reason,
        "findings": findings,
        "expected_source_file_count": expected_count,
        "current_source_file_count": len(current),
    }


def compute_plan_hash(plan: dict[str, Any]) -> str:
    """Hash execution-relevant plan content while excluding approval bookkeeping."""
    canonical = deepcopy(plan)
    metadata = canonical.get("plan_metadata", {})
    for key in ("plan_hash", "status", "approval"):
        metadata.pop(key, None)
    encoded = json.dumps(canonical, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def prepare_plan(plan: dict[str, Any]) -> dict[str, Any]:
    """Return a new plan version awaiting approval with its current canonical hash."""
    prepared = deepcopy(plan)
    metadata = prepared.setdefault("plan_metadata", {})
    metadata["status"] = "AWAITING_APPROVAL"
    metadata.pop("approval", None)
    metadata["plan_hash"] = compute_plan_hash(prepared)
    return prepared


def approve_plan(
    plan: dict[str, Any], *, plan_id: str, plan_version: str, plan_hash: str,
    approver_note: str | None = None, approver_type: str = "HUMAN",
    approval_policy_version: str | None = None, approval_audit: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Approve only the exact draft/version/hash presented for user review."""
    approved = deepcopy(plan)
    metadata = approved.get("plan_metadata", {})
    current_hash = compute_plan_hash(approved)
    if metadata.get("status") != "AWAITING_APPROVAL":
        raise ApprovalError("Only an AWAITING_APPROVAL plan can be approved.")
    if metadata.get("plan_id") != plan_id or metadata.get("plan_version") != plan_version:
        raise ApprovalError("Approval plan ID or version does not match the presented plan.")
    if metadata.get("plan_hash") != current_hash or plan_hash != current_hash:
        raise ApprovalError("Approval hash does not match the current execution-relevant plan content.")
    metadata["status"] = "APPROVED"
    metadata["approval"] = {
        "approved_plan_id": plan_id,
        "approved_plan_version": plan_version,
        "approved_plan_hash": current_hash,
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "approver_type": approver_type,
        "approver_note": approver_note,
    }
    if approval_policy_version is not None:
        metadata["approval"]["approval_policy_version"] = approval_policy_version
    if approval_audit is not None:
        metadata["approval"]["approval_audit"] = approval_audit
    return approved


def refresh_plan_after_change(plan: dict[str, Any]) -> dict[str, Any]:
    """Invalidate any approval if execution-relevant plan content has changed."""
    refreshed = deepcopy(plan)
    metadata = refreshed.setdefault("plan_metadata", {})
    new_hash = compute_plan_hash(refreshed)
    approved_hash = metadata.get("approval", {}).get("approved_plan_hash")
    if metadata.get("status") == "APPROVED" and approved_hash == new_hash:
        return refreshed
    metadata["status"] = "AWAITING_APPROVAL"
    metadata.pop("approval", None)
    metadata["plan_hash"] = new_hash
    return refreshed


def assert_execution_allowed(
    plan: dict[str, Any], source_directory: Path | None = None, *, for_stage5: bool = False
) -> dict[str, Any] | None:
    """Require exact approval and an unchanged bound package before Stage 5 can begin."""
    metadata = plan.get("plan_metadata", {})
    if metadata.get("status") != "APPROVED":
        raise ExecutionNotApprovedError("Processing plan is not approved; execution is blocked.")
    approval = metadata.get("approval", {})
    if (
        approval.get("approved_plan_id") != metadata.get("plan_id")
        or approval.get("approved_plan_version") != metadata.get("plan_version")
        or approval.get("approved_plan_hash") != compute_plan_hash(plan)
    ):
        raise ExecutionNotApprovedError("Approved plan hash no longer matches; execution is blocked.")
    if not metadata.get("source_snapshot"):
        raise SourceIntegrityError("EXECUTION_BLOCKED\nReason: SOURCE_SNAPSHOT_REQUIRED")
    if source_directory is None:
        raise SourceIntegrityError("EXECUTION_BLOCKED\nReason: SOURCE_DIRECTORY_REQUIRED")
    comparison = compare_source_snapshot(metadata["source_snapshot"], source_directory)
    if comparison["status"] != "SOURCE_UNCHANGED":
        finding = comparison["findings"][0]
        raise SourceIntegrityError(
            f"EXECUTION_BLOCKED\nReason: {finding['reason']}\nFile: {finding['file']}"
        )
    if for_stage5:
        return comparison
    raise NotImplementedError("Stage 5 execution is intentionally not implemented.")


def validate_processing_plan(plan: dict[str, Any], profile: dict[str, Any], dataset_map: dict[str, Any]) -> list[str]:
    """Validate plan contracts/references only; never judge business correctness."""
    errors: list[str] = []
    metadata = plan.get("plan_metadata", {})
    _require(metadata, {"plan_id", "plan_version", "profile_run_id", "dataset_map_reference", "created_at", "status", "plan_hash", "source_snapshot"}, "plan_metadata", errors)
    if metadata.get("profile_run_id") != profile.get("run_id"):
        errors.append("plan_metadata.profile_run_id does not match the profile.")
    if metadata.get("status") not in PLAN_STATUSES:
        errors.append("plan_metadata.status is invalid.")
    if metadata.get("plan_hash") != compute_plan_hash(plan):
        errors.append("plan_metadata.plan_hash does not match execution-relevant content.")
    _validate_source_snapshot(metadata.get("source_snapshot"), profile, errors)
    dataset_ids = {item.get("logical_dataset_id") for item in dataset_map.get("logical_datasets", [])}
    source_index = _source_index(profile)
    physical_fields = profile_field_index(profile)
    output_ids: set[str] = set()
    for index, output in enumerate(plan.get("proposed_outputs", [])):
        path = f"proposed_outputs[{index}]"
        _require(output, {"output_id", "filename", "source_logical_datasets", "output_grain", "output_columns", "unions", "joins", "period_handling", "formula_handling", "exclusions", "validations", "unresolved_issues", "confidence"}, path, errors)
        output_id = output.get("output_id")
        if not isinstance(output_id, str) or not output_id:
            errors.append(f"{path}.output_id is required.")
        elif output_id in output_ids:
            errors.append(f"Duplicate output_id: {output_id}.")
        else:
            output_ids.add(output_id)
        _enum(output.get("confidence"), CONFIDENCES, f"{path}.confidence", errors)
        assignment_ids: set[str] = set()
        for assignment_index, assignment in enumerate(output.get("source_assignments", [])):
            assignment_path = f"{path}.source_assignments[{assignment_index}]"
            assignment_id = assignment.get("assignment_id")
            if not isinstance(assignment_id, str) or not assignment_id:
                errors.append(f"{assignment_path}.assignment_id is required.")
            elif assignment_id in assignment_ids:
                errors.append(f"Duplicate source assignment ID in {path}: {assignment_id}.")
            else:
                assignment_ids.add(assignment_id)
            for reference in assignment.get("source_references", []):
                _validate_source_reference(reference, source_index, assignment_path, errors, require_region=True)
        for dataset_id in output.get("source_logical_datasets", []):
            if dataset_id not in dataset_ids:
                errors.append(f"{path} references unknown logical dataset: {dataset_id}.")
        for column_index, column in enumerate(output.get("output_columns", [])):
            column_path = f"{path}.output_columns[{column_index}]"
            _require(column, {"Output_Column", "Concept", "Data_Type", "Source_Type", "Transformation", "Required", "Blank_Handling", "Confidence"}, column_path, errors)
            _enum(column.get("Source_Type"), SOURCE_TYPES, f"{column_path}.Source_Type", errors)
            _enum(column.get("Confidence"), CONFIDENCES, f"{column_path}.Confidence", errors)
            if column.get("Source_Type") == "DIRECT_COLUMN":
                _require(column, {"Source_Field_IDs", "Source_Dataset"}, column_path, errors)
                _validate_plan_field_ids(
                    column.get("Source_Field_IDs"), physical_fields, column_path, errors,
                    output.get("source_assignments", []), column.get("Source_References", []),
                )
                if not column.get("Source_References") and column.get("Source_Assignment_ID") not in assignment_ids:
                    errors.append(f"{column_path} requires Source_References or a valid Source_Assignment_ID.")
                for reference in column.get("Source_References", []):
                    _validate_source_reference(reference, source_index, column_path, errors, require_region=True)
        for join_index, join in enumerate(output.get("joins", [])):
            join_path = f"{path}.joins[{join_index}]"
            _enum(join.get("join_type"), JOIN_TYPES, f"{join_path}.join_type", errors)
            _enum(join.get("expected_cardinality"), CARDINALITIES, f"{join_path}.expected_cardinality", errors)
            if join.get("expected_cardinality") == "MANY_TO_MANY" and not join.get("duplicate_key_action"):
                errors.append(f"{join_path} unsafe MANY_TO_MANY join requires duplicate_key_action.")
        for validation_index, validation in enumerate(output.get("validations", [])):
            _validate_reconciliation_rule(
                validation, f"{path}.validations[{validation_index}]", errors
            )
        _validate_unpivot_config(output, path, errors, physical_fields)
        _validate_blank_handling(output, path, errors)
    for index, item in enumerate(plan.get("non_output_logical_datasets", [])):
        dataset_id = item.get("logical_dataset_id")
        if dataset_id not in dataset_ids:
            errors.append(f"non_output_logical_datasets[{index}] references unknown logical dataset: {dataset_id}.")
    return errors


def validate_processing_plan_file(plan_path: Path, profile_path: Path, dataset_map_path: Path) -> None:
    plan = json.loads(Path(plan_path).read_text(encoding="utf-8"))
    profile = json.loads(Path(profile_path).read_text(encoding="utf-8"))
    dataset_map = json.loads(Path(dataset_map_path).read_text(encoding="utf-8"))
    errors = validate_processing_plan(plan, profile, dataset_map)
    if errors:
        raise ProcessingPlanValidationError("\n".join(errors))


def write_processing_plan_review(plan: dict[str, Any], output_path: Path) -> None:
    """Write a human-review workbook; it does not create an output dataset."""
    sheets: dict[str, list[dict[str, Any]]] = {name: [] for name in (
        "Plan_Summary", "Source_Snapshot", "Outputs", "Output_Columns", "Source_Mappings", "Transformations", "Reconciliation", "Exclusions", "Unresolved"
    )}
    metadata = plan["plan_metadata"]
    summary_metadata = {
        key: json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value
        for key, value in metadata.items()
        if key not in {"approval", "source_snapshot"}
    }
    sheets["Plan_Summary"].append(
        summary_metadata | {"bound_source_file_count": metadata["source_snapshot"]["expected_source_file_count"]}
    )
    sheets["Source_Snapshot"].extend(metadata["source_snapshot"]["source_files"])
    for output in plan.get("proposed_outputs", []):
        sheets["Outputs"].append({
            "output_id": output.get("output_id"), "filename": output.get("filename"), "description": output.get("description"),
            "source_logical_datasets": " | ".join(output.get("source_logical_datasets", [])),
            "output_grain": output.get("output_grain"), "confidence": output.get("confidence"),
        })
        for assignment in output.get("source_assignments", []):
            for reference in assignment.get("source_references", []):
                sheets["Source_Mappings"].append({
                    "output_id": output.get("output_id"),
                    "output_column": "ALL_DIRECT_SOURCE_COLUMNS",
                    "assignment_id": assignment.get("assignment_id"),
                    **reference,
                })
        for column in output.get("output_columns", []):
            sheets["Output_Columns"].append({"output_id": output.get("output_id"), **column})
            for reference in column.get("Source_References", []):
                sheets["Source_Mappings"].append({"output_id": output.get("output_id"), "output_column": column.get("Output_Column"), **reference})
        for transformation in output.get("transformations", []):
            sheets["Transformations"].append({"output_id": output.get("output_id"), **transformation})
        for validation in output.get("validations", []):
            sheets["Reconciliation"].append({"output_id": output.get("output_id"), **validation})
        for exclusion in output.get("exclusions", []):
            sheets["Exclusions"].append({"output_id": output.get("output_id"), **exclusion})
        for item in output.get("unresolved_issues", []):
            sheets["Unresolved"].append({"output_id": output.get("output_id"), **item})
    for item in plan.get("non_output_logical_datasets", []):
        sheets["Outputs"].append({"output_id": "NON_OUTPUT", **item})
    workbook = Workbook()
    active = workbook.active
    active.title = "Plan_Summary"
    for name, rows in sheets.items():
        worksheet = active if name == "Plan_Summary" else workbook.create_sheet(name)
        _write_sheet(worksheet, rows)
    workbook.save(output_path)


def write_processing_plan_preview(previews: dict[str, list[dict[str, Any]]], output_path: Path) -> None:
    """Write bounded illustrative plan previews, not final output files."""
    workbook = Workbook()
    active = workbook.active
    first = True
    for output_id, rows in previews.items():
        worksheet = active if first else workbook.create_sheet()
        first = False
        worksheet.title = output_id[:31]
        worksheet.append(["ILLUSTRATIVE PREVIEW - NOT FINAL OUTPUT"])
        headers = list(rows[0]) if rows else ["No preview rows"]
        worksheet.append(headers)
        for row in rows[:10]:
            worksheet.append([row.get(header) for header in headers])
        worksheet.freeze_panes = "A3"
        worksheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{worksheet.max_row}"
        for index, header in enumerate(headers, start=1):
            values = [str(header)] + [str(row.get(header, "")) for row in rows[:10]]
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 12), 45)
    workbook.save(output_path)


def _validate_blank_handling(output: dict[str, Any], path: str, errors: list[str]) -> None:
    for column in output.get("output_columns", []):
        treatment = str(column.get("Blank_Handling", "")).casefold()
        if "zero" in treatment and "do not" not in treatment and "preserve" not in treatment:
            errors.append(f"{path} converts blank to zero without an explicit preservation statement.")


def _validate_reconciliation_rule(validation: Any, path: str, errors: list[str]) -> None:
    if not isinstance(validation, dict) or "comparison_method" not in validation:
        return
    method = validation.get("comparison_method")
    if method not in RECONCILIATION_COMPARISON_METHODS:
        errors.append(f"{path}.comparison_method is invalid.")
        return
    if method == "ROUND_TO_PRECISION":
        decimal_places = validation.get("decimal_places")
        if not isinstance(decimal_places, int) or decimal_places < 0:
            errors.append(f"{path}.decimal_places must be a non-negative integer.")
    elif method == "ABSOLUTE_TOLERANCE":
        try:
            if Decimal(str(validation.get("tolerance"))) < 0:
                errors.append(f"{path}.tolerance must be non-negative.")
        except (InvalidOperation, ValueError):
            errors.append(f"{path}.tolerance must be numeric.")


def _validate_unpivot_config(
    output: dict[str, Any],
    path: str,
    errors: list[str],
    physical_fields: dict[str, dict[str, Any]],
) -> None:
    unpivot = output.get("unpivot")
    if unpivot is None:
        return
    _require(unpivot, {"source_field_ids", "dimension_output_column", "dimension_values", "measure_output_column"}, f"{path}.unpivot", errors)
    if not isinstance(unpivot, dict):
        return
    source_fields, values = unpivot.get("source_field_ids"), unpivot.get("dimension_values")
    if not isinstance(source_fields, list) or not source_fields:
        errors.append(f"{path}.unpivot.source_field_ids must be a non-empty list.")
    if not isinstance(values, list) or len(values) != len(source_fields or []):
        errors.append(f"{path}.unpivot.dimension_values must align with source_field_ids.")
    for field_id in source_fields or []:
        if field_id not in physical_fields:
            errors.append(f"{path}.unpivot references unknown source_field_id: {field_id}.")
    output_columns = {item.get("Output_Column") for item in output.get("output_columns", [])}
    for key in ("dimension_output_column", "measure_output_column"):
        if unpivot.get(key) not in output_columns:
            errors.append(f"{path}.unpivot.{key} must reference an output column.")


def _validate_plan_field_ids(
    field_ids: Any,
    physical_fields: dict[str, dict[str, Any]],
    path: str,
    errors: list[str],
    assignments: list[dict[str, Any]],
    direct_references: list[dict[str, Any]],
) -> None:
    if not isinstance(field_ids, list) or not field_ids:
        errors.append(f"{path}.Source_Field_IDs must be a non-empty list.")
        return
    assigned_regions = {
        reference.get("region_id")
        for assignment in assignments
        for reference in assignment.get("source_references", [])
    }
    assigned_regions.update(reference.get("region_id") for reference in direct_references)
    assigned_regions.discard(None)
    seen_regions: set[str] = set()
    for field_id in field_ids:
        field = physical_fields.get(field_id)
        if field is None:
            errors.append(f"{path} references unknown source_field_id: {field_id}.")
            continue
        region_id = field.get("region_id")
        if region_id not in assigned_regions:
            errors.append(f"{path} source_field_id is outside its assigned source regions: {field_id}.")
        if region_id in seen_regions:
            errors.append(f"{path} has multiple source_field_ids for one region: {region_id}.")
        seen_regions.add(region_id)
    missing_regions = assigned_regions - seen_regions
    if missing_regions:
        errors.append(f"{path}.Source_Field_IDs does not cover assigned region(s): {', '.join(sorted(missing_regions))}.")


def _validate_source_snapshot(source_snapshot: Any, profile: dict[str, Any], errors: list[str]) -> None:
    if not isinstance(source_snapshot, dict):
        errors.append("plan_metadata.source_snapshot must be an object.")
        return
    source_files = source_snapshot.get("source_files")
    if not isinstance(source_files, list):
        errors.append("plan_metadata.source_snapshot.source_files must be a list.")
        return
    if source_snapshot.get("expected_source_file_count") != len(source_files):
        errors.append("plan_metadata.source_snapshot expected file count does not match snapshot entries.")
    profile_statuses = {
        workbook["source_id"]: workbook.get("profiling_status")
        for workbook in profile.get("workbook_profiles", [])
    }
    profile_sources = {source["source_id"]: source for source in profile.get("source_files", [])}
    seen_paths: set[str] = set()
    for index, source in enumerate(source_files):
        path = f"plan_metadata.source_snapshot.source_files[{index}]"
        _require(source, {"source_id", "filename", "relative_path", "extension", "sha256", "profiling_status"}, path, errors)
        source_id = source.get("source_id")
        profile_source = profile_sources.get(source_id)
        if profile_source is None:
            errors.append(f"{path} references unknown profile source_id: {source_id}.")
            continue
        for key in ("filename", "extension", "sha256"):
            if source.get(key) != profile_source.get(key):
                errors.append(f"{path}.{key} does not match the accepted profile source.")
        if source.get("profiling_status") != profile_statuses.get(source_id):
            errors.append(f"{path}.profiling_status does not match the accepted profile.")
        relative_path = source.get("relative_path")
        if relative_path in seen_paths:
            errors.append(f"Duplicate source snapshot relative_path: {relative_path}.")
        seen_paths.add(relative_path)


def _validate_source_reference(reference: Any, source_index: dict[str, dict[str, set[str]]], path: str, errors: list[str], *, require_region: bool) -> None:
    if not isinstance(reference, dict):
        errors.append(f"{path} source reference must be an object.")
        return
    source_id, worksheet_name, region_id = reference.get("source_id"), reference.get("worksheet_name"), reference.get("region_id")
    if source_id not in source_index:
        errors.append(f"{path} references unknown source_id: {source_id}.")
        return
    if worksheet_name not in source_index[source_id]:
        errors.append(f"{path} references unknown worksheet: {worksheet_name}.")
        return
    if require_region and not region_id:
        errors.append(f"{path} requires region_id.")
    if region_id and region_id not in source_index[source_id][worksheet_name]:
        errors.append(f"{path} references unknown region_id: {region_id}.")


def _require(value: Any, keys: set[str], path: str, errors: list[str]) -> None:
    if not isinstance(value, dict):
        errors.append(f"{path} must be an object.")
        return
    for key in keys:
        if key not in value:
            errors.append(f"{path}.{key} is required.")


def _enum(value: Any, allowed: set[str], path: str, errors: list[str]) -> None:
    if value not in allowed:
        errors.append(f"{path} must be one of {sorted(allowed)}.")


def _write_sheet(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = list(dict.fromkeys(key for row in rows for key in row)) if rows else ["No records"]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([
            json.dumps(value, sort_keys=True, ensure_ascii=False)
            if isinstance(value := row.get(header), (dict, list)) else value
            for header in headers
        ])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows]
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 12), 50)
