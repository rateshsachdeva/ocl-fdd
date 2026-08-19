"""Vendor-neutral Dataset Map validation and review-artifact utilities."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .metadata import METADATA_STATUSES, METADATA_TYPES, REPORTING_FREQUENCIES, SCENARIOS
from .lineage import profile_field_index


CONFIDENCES = {"HIGH", "MEDIUM", "LOW"}
DATASET_ROLES = {"PRIMARY_DATA", "SUPPORTING_DATA", "CONTEXT", "PRESENTATION_IGNORE", "UNKNOWN"}
FIELD_ROLES = {
    "IDENTIFIER", "DIMENSION", "MEASURE", "PERIOD", "DATE", "SCENARIO", "CURRENCY",
    "DESCRIPTION", "CLASSIFICATION", "METADATA", "UNKNOWN",
}
INTERPRETATION_STATUSES = {"DRAFT", "REVIEWED", "CONFIRMED"}


class DatasetMapValidationError(ValueError):
    """Raised when a Dataset Map has invalid structure or profile references."""


def validate_dataset_map(dataset_map: dict[str, Any], profile: dict[str, Any]) -> list[str]:
    """Validate Dataset Map structure and references without judging interpretation."""
    errors: list[str] = []
    _require_keys(dataset_map, {"run_info", "source_understanding", "logical_datasets", "global_observations"}, "root", errors)
    run_info = dataset_map.get("run_info", {})
    _require_keys(run_info, {"profile_run_id", "dataset_map_version", "created_at", "interpretation_status"}, "run_info", errors)
    if run_info.get("profile_run_id") != profile.get("run_id"):
        errors.append("run_info.profile_run_id does not match the supplied profile.")
    if run_info.get("interpretation_status") not in INTERPRETATION_STATUSES:
        errors.append("run_info.interpretation_status is invalid.")

    source_index = _source_index(profile)
    physical_field_index = profile_field_index(profile)
    for index, understanding in enumerate(dataset_map.get("source_understanding", [])):
        path = f"source_understanding[{index}]"
        _validate_source_reference(understanding.get("source_reference"), source_index, path, errors, field_index=physical_field_index)
        _validate_enum(understanding.get("role"), DATASET_ROLES, f"{path}.role", errors)
        _validate_enum(understanding.get("confidence"), CONFIDENCES, f"{path}.confidence", errors)

    dataset_ids: set[str] = set()
    fields_by_dataset: dict[str, set[str]] = {}
    for index, dataset in enumerate(dataset_map.get("logical_datasets", [])):
        path = f"logical_datasets[{index}]"
        _require_keys(dataset, {"logical_dataset_id", "name", "role", "contributing_source_regions", "interpretation_confidence", "unresolved_matters"}, path, errors)
        dataset_id = dataset.get("logical_dataset_id")
        if not isinstance(dataset_id, str) or not dataset_id:
            errors.append(f"{path}.logical_dataset_id is required.")
        elif dataset_id in dataset_ids:
            errors.append(f"Duplicate logical_dataset_id: {dataset_id}.")
        else:
            dataset_ids.add(dataset_id)
        _validate_enum(dataset.get("role"), DATASET_ROLES, f"{path}.role", errors)
        _validate_enum(dataset.get("interpretation_confidence"), CONFIDENCES, f"{path}.interpretation_confidence", errors)
        _validate_enum(dataset.get("grain_confidence"), CONFIDENCES, f"{path}.grain_confidence", errors, optional=True)
        for reference in dataset.get("contributing_source_regions", []):
            _validate_source_reference(reference, source_index, f"{path}.contributing_source_regions", errors, require_region=True, field_index=physical_field_index)
        for reference in dataset.get("supporting_sources", []):
            _validate_source_reference(reference, source_index, f"{path}.supporting_sources", errors, field_index=physical_field_index)
        for metadata_index, metadata in enumerate(dataset.get("metadata", [])):
            _validate_metadata_record(metadata, f"{path}.metadata[{metadata_index}]", errors)
        field_ids: set[str] = set()
        for field_index, field in enumerate(dataset.get("conceptual_fields", [])):
            field_path = f"{path}.conceptual_fields[{field_index}]"
            field_id = field.get("canonical_name")
            if not isinstance(field_id, str) or not field_id:
                errors.append(f"{field_path}.canonical_name is required.")
            elif field_id in field_ids:
                errors.append(f"Duplicate field name in {dataset_id}: {field_id}.")
            else:
                field_ids.add(field_id)
            physical_ids = field.get("source_field_ids")
            if not isinstance(physical_ids, list):
                errors.append(f"{field_path}.source_field_ids must be a list.")
            elif field.get("directly_present") and not physical_ids:
                errors.append(f"{field_path}.source_field_ids is required for directly present fields.")
            for physical_id in physical_ids or []:
                if physical_id not in physical_field_index:
                    errors.append(f"{field_path} references unknown source_field_id: {physical_id}.")
            _validate_enum(field.get("proposed_role"), FIELD_ROLES, f"{field_path}.proposed_role", errors)
            _validate_enum(field.get("confidence"), CONFIDENCES, f"{field_path}.confidence", errors)
            for reference in field.get("source_references", []):
                _validate_source_reference(reference, source_index, f"{field_path}.source_references", errors, require_region=True, field_index=physical_field_index)
        fields_by_dataset[dataset_id] = field_ids

    for dataset in dataset_map.get("logical_datasets", []):
        for relationship in dataset.get("relationships", []):
            _validate_relationship(relationship, dataset_ids, fields_by_dataset, errors)
    for relationship in dataset_map.get("global_observations", {}).get("cross_file_relationships", []):
        _validate_relationship(relationship, dataset_ids, fields_by_dataset, errors)
    return errors


def validate_dataset_map_file(dataset_map_path: Path, profile_path: Path) -> None:
    """Raise a concise exception if a Dataset Map file is invalid for its profile."""
    dataset_map = json.loads(Path(dataset_map_path).read_text(encoding="utf-8"))
    profile = json.loads(Path(profile_path).read_text(encoding="utf-8"))
    errors = validate_dataset_map(dataset_map, profile)
    if errors:
        raise DatasetMapValidationError("\n".join(errors))


def write_dataset_map_review(dataset_map: dict[str, Any], output_path: Path) -> None:
    """Create a simple human-review workbook from a validated Dataset Map."""
    workbook = Workbook()
    sheets = {
        "Logical_Datasets": [],
        "Source_Assignments": [],
        "Fields": [],
        "Relationships": [],
        "Metadata": [],
        "Unresolved": [],
    }
    for dataset in dataset_map.get("logical_datasets", []):
        sheets["Logical_Datasets"].append({
            "logical_dataset_id": dataset.get("logical_dataset_id"),
            "name": dataset.get("name"),
            "role": dataset.get("role"),
            "proposed_grain": dataset.get("proposed_grain"),
            "grain_confidence": dataset.get("grain_confidence"),
            "interpretation_confidence": dataset.get("interpretation_confidence"),
            "description": dataset.get("description"),
        })
        for reference in dataset.get("contributing_source_regions", []):
            sheets["Source_Assignments"].append({
                "logical_dataset_id": dataset.get("logical_dataset_id"),
                "assignment_type": "CONTRIBUTING_REGION",
                **reference,
            })
        for reference in dataset.get("supporting_sources", []):
            sheets["Source_Assignments"].append({
                "logical_dataset_id": dataset.get("logical_dataset_id"),
                "assignment_type": "SUPPORTING_SOURCE",
                **reference,
            })
        for field in dataset.get("conceptual_fields", []) + dataset.get("context_derived_fields", []):
            sheets["Fields"].append({
                "logical_dataset_id": dataset.get("logical_dataset_id"),
                "canonical_name": field.get("canonical_name"),
                "proposed_role": field.get("proposed_role"),
                "directly_present": field.get("directly_present"),
                "confidence": field.get("confidence"),
                "interpretation": field.get("interpretation"),
                "source_field_names": " | ".join(field.get("source_field_names", [])),
                "source_field_ids": " | ".join(field.get("source_field_ids", [])),
            })
        for relationship in dataset.get("relationships", []):
            sheets["Relationships"].append({"logical_dataset_id": dataset.get("logical_dataset_id"), **relationship})
        for metadata in dataset.get("metadata", []):
            sheets["Metadata"].append({
                "Dataset": dataset.get("name"),
                "Metadata_Type": metadata.get("metadata_type"),
                "Value": _metadata_value_display(metadata.get("value")),
                "Confidence": metadata.get("confidence"),
                "Evidence": metadata.get("evidence"),
                "Source_Context": metadata.get("source_context"),
                "Status": metadata.get("status"),
            })
        for matter in dataset.get("unresolved_matters", []):
            sheets["Unresolved"].append({"logical_dataset_id": dataset.get("logical_dataset_id"), "matter": matter})
    for matter in dataset_map.get("global_observations", {}).get("unresolved_questions", []):
        sheets["Unresolved"].append({"logical_dataset_id": "GLOBAL", "matter": matter})
    active = workbook.active
    active.title = "Logical_Datasets"
    for sheet_name, rows in sheets.items():
        worksheet = active if sheet_name == "Logical_Datasets" else workbook.create_sheet(sheet_name)
        _write_review_sheet(worksheet, rows)
    workbook.save(output_path)


def _source_index(profile: dict[str, Any]) -> dict[str, dict[str, set[str]]]:
    index: dict[str, dict[str, set[str]]] = {}
    for workbook in profile.get("workbook_profiles", []):
        worksheets: dict[str, set[str]] = {}
        for worksheet in workbook.get("worksheet_profiles", []):
            worksheets[worksheet["worksheet_name"]] = {region["region_id"] for region in worksheet.get("data_regions", [])}
        index[workbook["source_id"]] = worksheets
    return index


def _validate_source_reference(
    reference: Any,
    source_index: dict[str, dict[str, set[str]]],
    path: str,
    errors: list[str],
    *,
    require_region: bool = False,
    field_index: dict[str, dict[str, Any]] | None = None,
) -> None:
    if not isinstance(reference, dict):
        errors.append(f"{path} must be an object.")
        return
    source_id = reference.get("source_id")
    worksheet_name = reference.get("worksheet_name")
    region_id = reference.get("region_id")
    if source_id not in source_index:
        errors.append(f"{path} references unknown source_id: {source_id}.")
        return
    if worksheet_name not in source_index[source_id]:
        errors.append(f"{path} references unknown worksheet: {worksheet_name}.")
        return
    if require_region and not region_id:
        errors.append(f"{path} requires region_id.")
    if region_id and region_id not in source_index[source_id][worksheet_name]:
        errors.append(f"{path} references unknown region_id: {region_id}.")
    physical_field_id = reference.get("source_field_id")
    if physical_field_id:
        physical = (field_index or {}).get(physical_field_id)
        if physical is None:
            errors.append(f"{path} references unknown source_field_id: {physical_field_id}.")
        elif any((
            physical.get("source_id") != source_id,
            physical.get("worksheet_name") != worksheet_name,
            region_id is not None and physical.get("region_id") != region_id,
        )):
            errors.append(f"{path}.source_field_id does not match its source context.")


def _validate_relationship(
    relationship: Any,
    dataset_ids: set[str],
    fields_by_dataset: dict[str, set[str]],
    errors: list[str],
) -> None:
    if not isinstance(relationship, dict):
        errors.append("Relationship must be an object.")
        return
    source_dataset = relationship.get("source_dataset_id")
    target_dataset = relationship.get("target_dataset_id")
    if source_dataset not in dataset_ids:
        errors.append(f"Relationship references missing source dataset: {source_dataset}.")
    if target_dataset not in dataset_ids:
        errors.append(f"Relationship references missing target dataset: {target_dataset}.")
    if source_dataset in fields_by_dataset and relationship.get("source_field") not in fields_by_dataset[source_dataset]:
        errors.append(f"Relationship references missing source field: {relationship.get('source_field')}.")
    if target_dataset in fields_by_dataset and relationship.get("target_field") not in fields_by_dataset[target_dataset]:
        errors.append(f"Relationship references missing target field: {relationship.get('target_field')}.")
    _validate_enum(relationship.get("confidence"), CONFIDENCES, "relationship.confidence", errors)


def _validate_metadata_record(metadata: Any, path: str, errors: list[str]) -> None:
    _require_keys(metadata, {"metadata_type", "value", "confidence", "evidence", "source_context", "status"}, path, errors)
    if not isinstance(metadata, dict):
        return
    _validate_enum(metadata.get("metadata_type"), METADATA_TYPES, f"{path}.metadata_type", errors)
    _validate_enum(metadata.get("confidence"), CONFIDENCES, f"{path}.confidence", errors)
    _validate_enum(metadata.get("status"), METADATA_STATUSES, f"{path}.status", errors)
    if metadata.get("metadata_type") == "REPORTING_FREQUENCY":
        _validate_values(metadata.get("value"), REPORTING_FREQUENCIES, f"{path}.value", errors)
    if metadata.get("metadata_type") == "SCENARIO":
        _validate_values(metadata.get("value"), SCENARIOS, f"{path}.value", errors)


def _validate_values(value: Any, allowed: set[str], path: str, errors: list[str]) -> None:
    values = value if isinstance(value, list) else [value]
    for item in values:
        if item is not None and item not in allowed:
            errors.append(f"{path} contains an unsupported value: {item}.")


def _require_keys(value: Any, keys: set[str], path: str, errors: list[str]) -> None:
    if not isinstance(value, dict):
        errors.append(f"{path} must be an object.")
        return
    for key in keys:
        if key not in value:
            errors.append(f"{path}.{key} is required.")


def _validate_enum(value: Any, allowed: set[str], path: str, errors: list[str], *, optional: bool = False) -> None:
    if optional and value is None:
        return
    if value not in allowed:
        errors.append(f"{path} must be one of {sorted(allowed)}.")


def _write_review_sheet(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0]) if rows else ["No records"]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows]
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 12), 45)


def _metadata_value_display(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, list):
        return " | ".join(str(item) for item in value)
    return str(value)
