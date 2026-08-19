"""Canonical immutable source reading and run-local staging."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

STAGING_VERSION = 1


class SourceDataError(ValueError):
    """Raised when physical source data cannot be resolved safely."""


@dataclass(slots=True)
class RegionData:
    rows: list[tuple[int, dict[str, Any]]]
    fields: dict[str, dict[str, Any]]
    from_staging: bool


def open_source_workbook(path: Path, *, read_only: bool) -> Any:
    """Open one Excel source using the repository's canonical safe options."""
    path = Path(path)
    return load_workbook(
        path,
        read_only=read_only,
        data_only=False,
        keep_vba=path.suffix.lower() in {".xlsm", ".xltm"},
    )


def staging_manifest_path(staging_directory: Path) -> Path:
    return Path(staging_directory) / "staging_manifest.json"


def create_staging(
    profile: dict[str, Any],
    source_directory: Path,
    staging_directory: Path,
) -> dict[str, Any]:
    """Create a source-hash-bound row staging layer from a completed profile."""
    staging_directory = Path(staging_directory)
    staging_directory.mkdir(parents=True, exist_ok=True)
    source_lookup = {item["source_id"]: item for item in profile.get("source_files", [])}
    manifest: dict[str, Any] = {
        "staging_version": STAGING_VERSION,
        "profile_run_id": profile.get("run_id"),
        "source_hashes": {item["source_id"]: item["sha256"] for item in profile.get("source_files", [])},
        "regions": {},
    }
    for workbook_profile in profile.get("workbook_profiles", []):
        source = source_lookup[workbook_profile["source_id"]]
        source_path = _resolve_source_path(source, source_directory)
        workbook = open_source_workbook(source_path, read_only=True)
        try:
            manifest["regions"].update(
                stage_profiled_workbook(source, workbook, workbook_profile, staging_directory)
            )
        finally:
            workbook.close()
    return write_staging_manifest(profile, staging_directory, manifest["regions"])


def stage_profiled_workbook(
    source: dict[str, Any],
    workbook: Any,
    workbook_profile: dict[str, Any],
    staging_directory: Path,
) -> dict[str, dict[str, Any]]:
    """Stage regions while the profiler already has the workbook open."""
    entries: dict[str, dict[str, Any]] = {}
    for worksheet_profile in workbook_profile.get("worksheet_profiles", []):
        worksheet = workbook[worksheet_profile["worksheet_name"]]
        for region in worksheet_profile.get("data_regions", []):
            entry, rows = _stage_region(worksheet, region)
            relative_path = Path("regions") / f"{_region_key(region['region_id'])}.jsonl"
            target = Path(staging_directory) / relative_path
            target.parent.mkdir(parents=True, exist_ok=True)
            with target.open("w", encoding="utf-8", newline="\n") as handle:
                for row in rows:
                    handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
            entries[region["region_id"]] = {**entry, "path": relative_path.as_posix()}
    return entries


def write_staging_manifest(
    profile: dict[str, Any],
    staging_directory: Path,
    region_entries: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    manifest = {
        "staging_version": STAGING_VERSION,
        "profile_run_id": profile.get("run_id"),
        "source_hashes": {item["source_id"]: item["sha256"] for item in profile.get("source_files", [])},
        "regions": region_entries,
    }
    Path(staging_directory).mkdir(parents=True, exist_ok=True)
    staging_manifest_path(staging_directory).write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    return manifest


def validate_staging(staging_directory: Path, profile: dict[str, Any]) -> tuple[bool, str]:
    path = staging_manifest_path(staging_directory)
    if not path.is_file():
        return False, "STAGING_MANIFEST_MISSING"
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False, "STAGING_MANIFEST_INVALID"
    if manifest.get("staging_version") != STAGING_VERSION:
        return False, "STAGING_VERSION_MISMATCH"
    expected = {item["source_id"]: item["sha256"] for item in profile.get("source_files", [])}
    if manifest.get("source_hashes") != expected:
        return False, "STAGING_SOURCE_HASH_MISMATCH"
    if manifest.get("profile_run_id") != profile.get("run_id"):
        return False, "STAGING_PROFILE_MISMATCH"
    for entry in manifest.get("regions", {}).values():
        if not (Path(staging_directory) / entry.get("path", "")).is_file():
            return False, "STAGING_REGION_MISSING"
    return True, "STAGING_VALID"


def read_region(
    source: dict[str, Any],
    worksheet_name: str,
    region: dict[str, Any],
    profile: dict[str, Any],
    source_directory: Path,
    staging_directory: Path | None = None,
) -> RegionData:
    """Read one profiled region by stable field ID, preferring valid staging."""
    fields = {
        item["field_id"]: {
            **item,
            "source_id": source["source_id"],
            "worksheet_name": worksheet_name,
            "region_id": region["region_id"],
        }
        for item in region.get("column_profiles", [])
        if item.get("field_id")
    }
    if not fields:
        raise SourceDataError(f"Region has no stable source fields: {region['region_id']}")
    if staging_directory is not None:
        valid, reason = validate_staging(staging_directory, profile)
        if not valid:
            raise SourceDataError(f"Staging cannot be reused: {reason}")
        manifest = json.loads(staging_manifest_path(staging_directory).read_text(encoding="utf-8"))
        entry = manifest.get("regions", {}).get(region["region_id"])
        if entry is None:
            raise SourceDataError(f"Staging does not contain region: {region['region_id']}")
        rows = []
        with (Path(staging_directory) / entry["path"]).open(encoding="utf-8") as handle:
            for line in handle:
                record = json.loads(line)
                rows.append((int(record["physical_row"]), record["values_by_field_id"]))
        return RegionData(rows, fields, True)
    return _read_region_from_workbook(source, worksheet_name, region, fields, source_directory)


def _read_region_from_workbook(
    source: dict[str, Any],
    worksheet_name: str,
    region: dict[str, Any],
    fields: dict[str, dict[str, Any]],
    source_directory: Path,
) -> RegionData:
    source_path = _resolve_source_path(source, source_directory)
    workbook = open_source_workbook(source_path, read_only=True)
    try:
        worksheet = workbook[worksheet_name]
        header_row = _header_row(region)
        _validate_exact_headers(worksheet, header_row, fields)
        rows: list[tuple[int, dict[str, Any]]] = []
        allowed_rows = _allowed_data_rows(region)
        for physical_row, values in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                max_row=region["end_row"],
                min_col=region["start_column"],
                max_col=region["end_column"],
                values_only=True,
            ),
            start=header_row + 1,
        ):
            if allowed_rows is not None and physical_row not in allowed_rows:
                continue
            if not any(value is not None for value in values):
                continue
            values_by_column = dict(zip(range(region["start_column"], region["end_column"] + 1), values))
            rows.append((physical_row, {
                field_id: _json_value(values_by_column.get(field["physical_column"]))
                for field_id, field in fields.items()
            }))
        return RegionData(rows, fields, False)
    finally:
        workbook.close()


def _stage_region(worksheet: Any, region: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    fields = {item["field_id"]: item for item in region.get("column_profiles", []) if item.get("field_id")}
    header_row = _header_row(region)
    _validate_exact_headers(worksheet, header_row, fields)
    allowed_rows = _allowed_data_rows(region)
    rows = []
    for physical_row, values in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=region["end_row"],
            min_col=region["start_column"],
            max_col=region["end_column"],
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if allowed_rows is not None and physical_row not in allowed_rows:
            continue
        if not any(value is not None for value in values):
            continue
        values_by_column = dict(zip(range(region["start_column"], region["end_column"] + 1), values))
        rows.append({
            "physical_row": physical_row,
            "values_by_field_id": {
                field_id: _json_value(values_by_column.get(field["physical_column"]))
                for field_id, field in fields.items()
            },
        })
    return {
        "source_id": region["source_id"],
        "worksheet_name": region["worksheet_name"],
        "region_id": region["region_id"],
        "header_row": header_row,
        "row_count": len(rows),
        "field_ids": sorted(fields),
    }, rows


def _validate_exact_headers(worksheet: Any, header_row: int, fields: dict[str, dict[str, Any]]) -> None:
    for field_id, field in fields.items():
        actual = worksheet.cell(header_row, int(field["physical_column"])).value
        if actual != field.get("exact_original_header"):
            raise SourceDataError(
                f"Exact source header changed for {field_id} at {worksheet.title}!{field['physical_column']}."
            )


def _header_row(region: dict[str, Any]) -> int:
    primary = [item["row"] for item in region.get("header_candidates", []) if item.get("confidence") == "PRIMARY"]
    if primary:
        return primary[0]
    candidates = [item["row"] for item in region.get("header_candidates", [])]
    return max(candidates) if candidates else int(region["start_row"])


def _allowed_data_rows(region: dict[str, Any]) -> set[int] | None:
    ranges = region.get("likely_data_row_ranges")
    if not ranges:
        return None
    return {row for item in ranges for row in range(int(item["start_row"]), int(item["end_row"]) + 1)}


def _resolve_source_path(source: dict[str, Any], source_directory: Path) -> Path:
    candidate = Path(source_directory) / source["filename"]
    return candidate if candidate.exists() else Path(source["path"])


def _region_key(region_id: str) -> str:
    return hashlib.sha256(region_id.encode("utf-8")).hexdigest()


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


__all__ = [
    "RegionData",
    "SourceDataError",
    "create_staging",
    "open_source_workbook",
    "read_region",
    "stage_profiled_workbook",
    "staging_manifest_path",
    "validate_staging",
    "write_staging_manifest",
]
