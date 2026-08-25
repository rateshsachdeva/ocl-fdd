"""Final deterministic QA for the published OCL workbook."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


class FinalQAError(RuntimeError):
    pass


def validate_final_databook(databook: Path, qa_output: Path | None = None) -> dict:
    databook = Path(databook)
    if not databook.exists():
        raise FinalQAError(f"Databook does not exist: {databook}")
    workbook = load_workbook(databook, read_only=False, data_only=False)
    issues: list[str] = []
    metrics: dict[str, object] = {"sheet_count": len(workbook.sheetnames), "sheets": list(workbook.sheetnames)}
    try:
        mandatory = {"Checks", "Mapping", "UNMAPPED", "SCOPE_EXCLUDED", "Flat File"}
        missing = sorted(mandatory - set(workbook.sheetnames))
        if missing:
            issues.append("Missing mandatory sheets: " + ", ".join(missing))

        formula_count = 0
        ref_error_count = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        if "#REF!" in cell.value.upper():
                            ref_error_count += 1
        metrics["formula_count"] = formula_count
        metrics["ref_error_count"] = ref_error_count
        if ref_error_count:
            issues.append(f"Workbook contains {ref_error_count} formula(s) with #REF!.")

        if "Checks" in workbook.sheetnames:
            sheet = workbook["Checks"]
            header_row, headers = _find_headers(sheet, {"Control_ID", "Python_Status"})
            status_col = headers.get("Python_Status")
            blocking = []
            if status_col and header_row:
                for row in range(header_row + 1, sheet.max_row + 1):
                    status = str(sheet.cell(row, status_col).value or "").upper()
                    if status in {"FAIL", "REVIEW_REQUIRED"}:
                        blocking.append(str(sheet.cell(row, headers.get("Control_ID", 1)).value or f"row {row}"))
            metrics["blocking_controls"] = blocking
            if blocking:
                issues.append("Blocking controls remain: " + ", ".join(blocking))

        if "Flat File" in workbook.sheetnames:
            sheet = workbook["Flat File"]
            required = ["Source_Record_ID", "Amount", "Scope", "Review_Status"]
            header_row, headers = _find_headers(sheet, set(required))
            missing_headers = [name for name in required if name not in headers]
            if missing_headers:
                issues.append("Flat File missing required columns: " + ", ".join(missing_headers))
            incomplete = 0
            if not missing_headers and header_row:
                for row in range(header_row + 1, sheet.max_row + 1):
                    if not sheet.cell(row, headers["Source_Record_ID"]).value:
                        incomplete += 1
                    if sheet.cell(row, headers["Amount"]).value in (None, ""):
                        incomplete += 1
                    if not sheet.cell(row, headers["Scope"]).value:
                        incomplete += 1
                    if not sheet.cell(row, headers["Review_Status"]).value:
                        incomplete += 1
            metrics["flat_file_incomplete_cells"] = incomplete
            if incomplete:
                issues.append(f"Flat File contains {incomplete} missing required lineage/judgment value(s).")

        source_tabs = [name for name in workbook.sheetnames if name.startswith("SRC_")]
        metrics["source_tab_count"] = len(source_tabs)
        unprotected = [name for name in source_tabs if not workbook[name].protection.sheet]
        if unprotected:
            issues.append("Source-copy tabs are not protected: " + ", ".join(unprotected))

        # In a full analysis run these sheets are AI-host authored from a
        # validated evidence package. If they exist, READY must never publish
        # them as empty placeholders. Explicit "no material issue/question"
        # conclusions are substantive and therefore pass this check.
        narrative_checks = {
            "Deal Issues": 4,
            "Key Findings": 8,
            "Q&A": 8,
        }
        narrative_status: dict[str, bool] = {}
        for name, start_row in narrative_checks.items():
            if name not in workbook.sheetnames:
                continue
            populated = _has_substantive_content(workbook[name], start_row)
            narrative_status[name] = populated
            if not populated:
                issues.append(f"{name} exists but contains no substantive analysis content.")
        metrics["narrative_sections_populated"] = narrative_status

        metrics["status"] = "PASS" if not issues else "FAIL"
        metrics["issues"] = issues
    finally:
        workbook.close()

    if qa_output:
        qa_output = Path(qa_output)
        qa_output.parent.mkdir(parents=True, exist_ok=True)
        qa_output.write_text(json.dumps(metrics, indent=2) + "\n", encoding="utf-8")
    if issues:
        raise FinalQAError("Final databook QA failed: " + " | ".join(issues))
    return metrics


def _find_headers(sheet, required: set[str]) -> tuple[int | None, dict[str, int]]:
    for row in range(1, min(sheet.max_row, 12) + 1):
        headers = {str(sheet.cell(row, col).value): col for col in range(1, sheet.max_column + 1) if sheet.cell(row, col).value not in (None, "")}
        if required <= set(headers):
            return row, headers
    return None, {}


def _has_substantive_content(sheet, start_row: int) -> bool:
    for row in range(start_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row, col).value
            if value not in (None, ""):
                return True
    return False
