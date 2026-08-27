"""Lean review workbooks for the standardized-data and OCL semantic handoff stages."""
from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ocl_agent.part1_databook.input_contract import DatasetProfile, StandardizedPackage
from ocl_agent.part1_databook.record_builder import RecordBuildResult
from ocl_agent.part1_databook.semantic_handoff import SemanticHandoff
from ocl_agent.schemas import ControlResult, OCLRecord

MAX_SAMPLE_VALUE_LENGTH = 160


def write_input_review(package: StandardizedPackage, profiles: tuple[DatasetProfile, ...], output_path: Path) -> Path:
    workbook = _new_workbook()
    sheet = workbook.create_sheet("Input_Datasets")
    sheet.append(["Dataset_File", "Rows", "Columns", "Manifest", "Metadata", "Lineage", "Field_Lineage"])
    for profile in profiles:
        sheet.append([profile.path.name, profile.row_count, len(profile.columns), "AVAILABLE" if package.execution_manifest else "MISSING", "AVAILABLE" if package.metadata else "MISSING", "AVAILABLE" if package.lineage else "MISSING", "AVAILABLE" if package.field_lineage else "MISSING"])
    _finish_sheet(sheet)
    fields = workbook.create_sheet("Fields")
    fields.append(["Dataset_File", "Column", "Sample_1", "Sample_2", "Sample_3"])
    for profile in profiles:
        for column in profile.columns:
            samples = [_clip(row.get(column)) for row in profile.sample_rows[:3]]
            samples.extend([None] * (3 - len(samples)))
            fields.append([profile.path.name, column, *samples])
    _finish_sheet(fields)
    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(["Logical_Dataset_ID", "Name", "Role", "Grain", "Metadata_Type", "Value", "Status", "Confidence"])
    metadata = package.metadata_payload()
    for dataset in metadata.get("logical_datasets", []):
        records = dataset.get("metadata", []) or [None]
        for item in records:
            metadata_sheet.append([dataset.get("logical_dataset_id"), dataset.get("name"), dataset.get("role"), dataset.get("dataset_grain"), item.get("metadata_type") if item else None, _clip(item.get("value")) if item else None, item.get("status") if item else None, item.get("confidence") if item else None])
    _finish_sheet(metadata_sheet)
    warnings = workbook.create_sheet("Input_Warnings")
    warnings.append(["Severity", "Message"])
    for message in package.warnings or ("No package-level handoff warnings.",):
        warnings.append(["WARNING" if package.warnings else "INFO", message])
    _finish_sheet(warnings)
    return _save(workbook, output_path)


def write_semantic_review(package: StandardizedPackage, profiles: tuple[DatasetProfile, ...], handoff: SemanticHandoff, build: RecordBuildResult, output_path: Path, *, judgment_issues=(), controls: tuple[ControlResult, ...] = ()) -> Path:
    workbook = _new_workbook()
    _write_dataset_summary(workbook, profiles, handoff, build)
    _write_handoff(workbook, handoff)
    _write_scope_review(workbook, build.records)
    _write_mapping_review(workbook, build.records)
    _write_wc_review(workbook, build.records)
    _write_unresolved(workbook, build, judgment_issues)
    _write_checks(workbook, build, controls)
    return _save(workbook, output_path)


def _write_dataset_summary(workbook: Workbook, profiles: tuple[DatasetProfile, ...], handoff: SemanticHandoff, build: RecordBuildResult) -> None:
    bindings = {item.file: item for item in handoff.datasets}
    sheet = workbook.create_sheet("Input_Datasets")
    sheet.append(["Dataset_File", "Rows", "Columns", "Usages", "Rows_Read_For_OCL", "Rows_Excluded_By_Usage_Filter", "Package_ID"])
    for profile in profiles:
        binding = bindings.get(profile.path.name)
        sheet.append([profile.path.name, profile.row_count, len(profile.columns), ", ".join(item.value for item in binding.usages) if binding else "UNBOUND", build.input_rows_by_dataset.get(profile.path.name, 0), build.excluded_rows_by_dataset.get(profile.path.name, 0), handoff.package_id])
    _finish_sheet(sheet)


def _write_handoff(workbook: Workbook, handoff: SemanticHandoff) -> None:
    sheet = workbook.create_sheet("Semantic_Handoff")
    sheet.append(["Dataset_File", "Usages", "Source_Record_ID", "Period", "Amount", "Source_Label", "Source_Code", "Entity", "Currency", "Movement_Type", "Dimensions", "Notes"])
    for item in handoff.datasets:
        fields = item.fields
        sheet.append([item.file, ", ".join(value.value for value in item.usages), fields.source_record_id, fields.period, fields.amount, fields.source_label, fields.source_code, fields.entity, fields.currency, fields.movement_type, ", ".join(item.dimensions), item.notes])
    _finish_sheet(sheet)


def _label_period_matrix(records: tuple[OCLRecord, ...]):
    periods = tuple(sorted({row.period for row in records}))
    values = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    exemplar = {}
    counts = defaultdict(int)
    for row in records:
        key = (row.source_label, str(row.dimensions.get("source_code") or ""), str(row.dimensions.get("entity") or ""))
        exemplar.setdefault(key, row)
        counts[key] += 1
        values[key][row.period] += row.amount
    return periods, values, exemplar, counts


def _write_scope_review(workbook: Workbook, records: tuple[OCLRecord, ...]) -> None:
    periods, values, exemplar, counts = _label_period_matrix(records)
    sheet = workbook.create_sheet("OCL_Scope_Review")
    sheet.append(["Source_Label", "Source_Code", "Entity", "Scope", "Review_Status", "Reason", "Record_Count", *periods])
    for key in sorted(exemplar, key=lambda value: tuple(str(part).casefold() for part in value)):
        row = exemplar[key]
        sheet.append([row.source_label, row.dimensions.get("source_code"), row.dimensions.get("entity"), row.judgment.scope.value, row.judgment.review_status.value, row.judgment.reason, counts[key], *[values[key].get(period, Decimal("0")) for period in periods]])
    _finish_sheet(sheet)


def _write_mapping_review(workbook: Workbook, records: tuple[OCLRecord, ...]) -> None:
    sheet = workbook.create_sheet("Mapping_Review")
    sheet.append(["Source_Label", "Source_Code", "Entity", "Scope", "Category", "Parent_Category", "Review_Status", "Reason"])
    for row in _unique_labels(records):
        judgment = row.judgment
        sheet.append([row.source_label, row.dimensions.get("source_code"), row.dimensions.get("entity"), judgment.scope.value, judgment.category, judgment.parent_category, judgment.review_status.value, judgment.reason])
    _finish_sheet(sheet)


def _write_wc_review(workbook: Workbook, records: tuple[OCLRecord, ...]) -> None:
    sheet = workbook.create_sheet("WC_Debt_Review")
    sheet.append(["Source_Label", "Source_Code", "Entity", "Scope", "Management_View", "FDD_View", "Normality", "Review_Status"])
    for row in _unique_labels(records):
        judgment = row.judgment
        sheet.append([row.source_label, row.dimensions.get("source_code"), row.dimensions.get("entity"), judgment.scope.value, judgment.management_view, judgment.fdd_view, judgment.normality, judgment.review_status.value])
    _finish_sheet(sheet)


def _write_unresolved(workbook: Workbook, build: RecordBuildResult, judgment_issues=()) -> None:
    sheet = workbook.create_sheet("Unresolved_Items")
    sheet.append(["Type", "Dataset_File", "CSV_Row", "Source_Record_ID", "Source_Label", "Message"])
    for issue in build.issues:
        sheet.append([issue.issue_type, issue.dataset_file, issue.csv_row, issue.source_record_id, None, issue.message])
    for issue in judgment_issues:
        sheet.append([issue.issue_type, None, None, None, issue.source_label, issue.message])
    if sheet.max_row == 1:
        sheet.append(["NONE", None, None, None, None, "No unresolved Stage 2 items."])
    _finish_sheet(sheet)


def _write_checks(workbook: Workbook, build: RecordBuildResult, controls: tuple[ControlResult, ...] = ()) -> None:
    sheet = workbook.create_sheet("Checks")
    sheet.append(["Check", "Status", "Actual", "Expected", "Message"])
    input_rows = sum(build.input_rows_by_dataset.values())
    excluded_rows = sum(build.excluded_rows_by_dataset.values())
    accounted = len(build.records) + len(build.issues) + excluded_rows
    sheet.append(["record_row_coverage", "PASS" if accounted == input_rows else "FAIL", accounted, input_rows, "Every row read for OCL records must become an OCLRecord, appear in Unresolved_Items, or be explicitly excluded by its usage filter."])
    ids = [row.source.source_record_id for row in build.records]
    duplicate_count = len(ids) - len(set(ids))
    sheet.append(["source_record_id_uniqueness", "PASS" if duplicate_count == 0 else "REVIEW_REQUIRED", duplicate_count, 0, "Duplicate Source_Record_ID values require review; no rows are dropped."])
    sheet.append(["semantic_build_issues", "PASS" if not build.issues else "REVIEW_REQUIRED", len(build.issues), 0, "Blank/invalid required standardized values remain visible in Unresolved_Items."])
    existing = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}
    for control in controls:
        if control.control_id not in existing:
            sheet.append([control.control_id, control.status.value, control.actual, control.expected, control.message])
    _finish_sheet(sheet)


def _unique_labels(records: tuple[OCLRecord, ...]) -> tuple[OCLRecord, ...]:
    seen = set()
    result = []
    for row in sorted(records, key=lambda item: (item.source_label.casefold(), str(item.dimensions.get("source_code") or "").casefold(), str(item.dimensions.get("entity") or "").casefold())):
        key = (" ".join(row.source_label.casefold().split()), str(row.dimensions.get("source_code") or "").strip().casefold(), str(row.dimensions.get("entity") or "").strip().casefold())
        if key not in seen:
            seen.add(key)
            result.append(row)
    return tuple(result)


def _new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def _finish_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        width = min(42, max(10, max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 100) + 1)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def _save(workbook: Workbook, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _clip(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    return text if len(text) <= MAX_SAMPLE_VALUE_LENGTH else text[: MAX_SAMPLE_VALUE_LENGTH - 1] + "…"
