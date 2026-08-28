"""Explicit movement support for OCL roll-forward schedules.

Movement arithmetic is never guessed. The package-specific semantic handoff may
include a `movement_rules` object on a MOVEMENT_RECORDS dataset. Each exact
source movement value maps to a role and multiplier. The top-level
`movement_to_annual` list explicitly aligns movement periods to annual periods.
"""
from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ocl_agent.part1_databook.input_contract import StandardizedPackage
from ocl_agent.part1_databook.judgments import JudgmentStore
from ocl_agent.part1_databook.reconciliation import DEFAULT_TOLERANCE
from ocl_agent.part1_databook.semantic_handoff import DatasetUsage, SemanticHandoff
from ocl_agent.schemas import CheckStatus, ControlResult, MovementRecord, OCLRecord, ReviewStatus, Scope, SourceReference
from ocl_agent.project_title import resolve_project_title
from ocl_agent.workbook_style import style_generated_support_cell

ALLOWED_ROLES = {"OPENING", "FLOW", "CLOSING"}


@dataclass(frozen=True)
class MovementAlignment:
    movement_period: str
    annual_period: str


@dataclass(frozen=True)
class MovementBuildResult:
    records: tuple[MovementRecord, ...]
    issues: tuple[str, ...]
    alignments: tuple[MovementAlignment, ...]
    population_coverage: str = "NONE"


def build_movements(package: StandardizedPackage, handoff: SemanticHandoff, judgments: JudgmentStore, handoff_path: Path) -> MovementBuildResult:
    movement_bindings = [binding for binding in handoff.datasets if DatasetUsage.MOVEMENT_RECORDS in binding.usages]
    if not movement_bindings:
        return MovementBuildResult((), (), (), "NONE")
    payload = json.loads(Path(handoff_path).read_text(encoding="utf-8"))
    raw_datasets = {str(item.get("file")): item for item in payload.get("datasets", []) if isinstance(item, dict)}
    alignments = tuple(MovementAlignment(str(item.get("movement_period", "")).strip(), str(item.get("annual_period", "")).strip()) for item in payload.get("movement_to_annual", []) if isinstance(item, dict))
    issues: list[str] = []
    records: list[MovementRecord] = []
    population_coverage = "PARTIAL" if any(binding.population_coverage == "PARTIAL" for binding in movement_bindings) else "FULL"
    for binding in movement_bindings:
        item = raw_datasets.get(binding.file, {})
        rules = item.get("movement_rules") or {}
        if not isinstance(rules, dict) or not rules:
            issues.append(f"{binding.file}: movement_rules are missing.")
            continue
        normalized_rules: dict[str, tuple[str, Decimal]] = {}
        for source_value, rule in rules.items():
            if isinstance(rule, str):
                role, multiplier = rule.strip().upper(), Decimal("1")
            elif isinstance(rule, dict):
                role = str(rule.get("role", "")).strip().upper()
                try:
                    multiplier = Decimal(str(rule.get("multiplier", "1")))
                except InvalidOperation:
                    issues.append(f"{binding.file}: invalid multiplier for movement value {source_value!r}.")
                    continue
            else:
                issues.append(f"{binding.file}: invalid movement rule for {source_value!r}.")
                continue
            if role not in ALLOWED_ROLES:
                issues.append(f"{binding.file}: unsupported movement role {role!r} for {source_value!r}.")
                continue
            normalized_rules[str(source_value).strip().casefold()] = (role, multiplier)
        fields = binding.fields
        assert fields.source_record_id and fields.period and fields.amount and fields.source_label and fields.movement_type
        with (package.root / binding.file).open(newline="", encoding="utf-8-sig") as handle:
            for csv_row, row in enumerate(csv.DictReader(handle), start=2):
                raw_type = str(row.get(fields.movement_type, "") or "").strip()
                rule = normalized_rules.get(raw_type.casefold())
                if rule is None:
                    issues.append(f"{binding.file}:{csv_row}: movement type {raw_type!r} is not explicitly mapped.")
                    continue
                role, rule_multiplier = rule
                source_record_id = str(row.get(fields.source_record_id, "") or "").strip()
                period = str(row.get(fields.period, "") or "").strip()
                label = str(row.get(fields.source_label, "") or "").strip()
                try:
                    amount = Decimal(str(row.get(fields.amount, "") or "").strip().replace(",", ""))
                except InvalidOperation:
                    issues.append(f"{binding.file}:{csv_row}: movement amount is not numeric.")
                    continue
                row_multiplier = rule_multiplier
                if fields.movement_multiplier:
                    try:
                        row_multiplier = Decimal(str(row.get(fields.movement_multiplier, "") or "").strip().replace(",", ""))
                    except InvalidOperation:
                        issues.append(f"{binding.file}:{csv_row}: movement multiplier is not numeric.")
                        continue
                    if row_multiplier not in {Decimal("-1"), Decimal("1")}:
                        issues.append(f"{binding.file}:{csv_row}: movement multiplier must be -1 or 1.")
                        continue
                if role in {"OPENING", "CLOSING"} and row_multiplier != Decimal("1"):
                    issues.append(f"{binding.file}:{csv_row}: {role} movement multiplier must be 1.")
                    continue
                source_code = str(row.get(fields.source_code, "") or "").strip() if fields.source_code else None
                entity = str(row.get(fields.entity, "") or "").strip() if fields.entity else None
                judgment = judgments.get(label, source_code, entity)
                if judgment.scope == Scope.REVIEW_REQUIRED or judgment.review_status != ReviewStatus.REVIEWED:
                    issues.append(f"{binding.file}:{csv_row}: movement label {label!r} does not have reviewed OCL judgment.")
                    continue
                dimensions = {
                    "dataset_file": binding.file,
                    "standardized_csv_row": csv_row,
                    "source_code": source_code,
                    "entity": entity,
                    "raw_movement_type": raw_type,
                    "population_coverage": binding.population_coverage,
                }
                records.append(MovementRecord(SourceReference(source_record_id), period, amount, label, role, row_multiplier, judgment, dimensions))
    if movement_bindings and not alignments and population_coverage == "FULL":
        issues.append("movement_to_annual alignment is missing.")
    if any(not item.movement_period or not item.annual_period for item in alignments):
        issues.append("movement_to_annual contains blank period values.")
    return MovementBuildResult(tuple(records), tuple(issues), alignments, population_coverage)


def rollforward_control(
    movements: tuple[MovementRecord, ...],
    annual_records: tuple[OCLRecord, ...],
    alignments: tuple[MovementAlignment, ...],
    issues: tuple[str, ...],
    population_coverage: str = "FULL",
) -> ControlResult:
    if not movements and not issues:
        return ControlResult("chk_rollforward", CheckStatus.NOT_APPLICABLE, message="No movement dataset is available.")
    if issues:
        return ControlResult("chk_rollforward", CheckStatus.REVIEW_REQUIRED, message="Movement data requires explicit reviewed rules/alignment before roll-forward publication.", evidence={"issues": list(issues)[:100]})
    grouped = _group_movements(movements)
    breaks: list[dict[str, str]] = []
    for (period, category), values in sorted(grouped.items()):
        expected_closing = values["OPENING"] + values["FLOW"]
        difference = expected_closing - values["CLOSING"]
        if abs(difference) >= DEFAULT_TOLERANCE:
            breaks.append({"type": "ROLLFORWARD", "period": period, "category": category, "opening": str(values["OPENING"]), "flow": str(values["FLOW"]), "closing": str(values["CLOSING"]), "difference": str(difference)})
    if population_coverage == "FULL":
        annual: dict[tuple[str, str], Decimal] = {}
        for row in annual_records:
            if row.dimensions.get("record_usage") == "MONTHLY_RECORDS" or row.judgment.scope != Scope.IN_SCOPE or not row.judgment.category:
                continue
            annual[(row.period, str(row.judgment.category))] = annual.get((row.period, str(row.judgment.category)), Decimal("0")) + row.amount
        for alignment in alignments:
            categories = {category for period, category in grouped if period == alignment.movement_period}
            for category in sorted(categories):
                closing = grouped[(alignment.movement_period, category)]["CLOSING"]
                listing = annual.get((alignment.annual_period, category), Decimal("0"))
                difference = closing - listing
                if abs(difference) >= DEFAULT_TOLERANCE:
                    breaks.append({"type": "CLOSING_TO_LISTING", "movement_period": alignment.movement_period, "annual_period": alignment.annual_period, "category": category, "closing": str(closing), "listing": str(listing), "difference": str(difference)})
    passed = not breaks
    if population_coverage == "PARTIAL":
        message = (
            "Explicit movement roll-forward arithmetic for the selected movement population; "
            "full OCL population completeness and closing-to-listing reconciliation are not assessed."
            if passed else "Selected-population movement roll-forward contains arithmetic breaks."
        )
    else:
        message = "Explicit movement roll-forward and closing-to-listing reconciliation." if passed else "Movement roll-forward contains reconciliation breaks."
    return ControlResult(
        "chk_rollforward",
        CheckStatus.PASS if passed else CheckStatus.FAIL,
        Decimal(len(breaks)),
        Decimal("0"),
        Decimal(len(breaks)),
        message=message,
        evidence={
            "breaks": breaks[:100],
            "population_coverage": population_coverage,
            "population_completeness_assessed": population_coverage == "FULL",
        },
    )


def embed_rollforward(
    databook_path: Path,
    movements: tuple[MovementRecord, ...],
    population_coverage: str = "FULL",
) -> None:
    if not movements:
        return
    workbook = load_workbook(databook_path)
    support = workbook["Movements"] if "Movements" in workbook.sheetnames else workbook.create_sheet("Movements")
    sheet = workbook["Roll-forward"] if "Roll-forward" in workbook.sheetnames else workbook.create_sheet("Roll-forward")
    if support.max_row:
        support.delete_rows(1, support.max_row)
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    grouped = _group_movements(movements)

    support_headers = ["Source_Record_ID", "Source_Dataset", "Source_Row", "Period", "Category", "Movement_Role", "Raw_Amount", "Multiplier", "Signed_Amount"]
    for column, value in enumerate(support_headers, start=1):
        support.cell(1, column, value)
    support_row = 2
    for movement in movements:
        if movement.judgment.scope != Scope.IN_SCOPE or not movement.judgment.category:
            continue
        values = [
            movement.source.source_record_id,
            movement.dimensions.get("dataset_file"),
            movement.dimensions.get("standardized_csv_row"),
            movement.period,
            str(movement.judgment.category),
            movement.movement_role,
            movement.amount,
            movement.multiplier,
        ]
        for column, value in enumerate(values, start=1):
            cell = support.cell(support_row, column, value)
            style_generated_support_cell(cell, role="model")
        signed_amount = support.cell(support_row, 9, f"=G{support_row}*H{support_row}")
        style_generated_support_cell(signed_amount, role="model")
        support_row += 1
    support.freeze_panes = "A2"
    support.sheet_view.showGridLines = False

    sheet["A1"] = resolve_project_title(workbook=workbook)
    sheet["A2"] = "Roll-forward"
    if population_coverage == "PARTIAL":
        sheet["A3"] = "Coverage: selected movement records only; this schedule does not evidence full OCL population completeness."
    periods = sorted({period for period, _category in grouped})
    categories = sorted({category for _period, category in grouped})
    current_row = 6
    for category in categories:
        sheet.cell(current_row, 2, category)
        header_row = current_row + 1
        sheet.cell(header_row, 2, "Movement")
        for column, period in enumerate(periods, start=3):
            cell = sheet.cell(header_row, column, _display_period(period))
            if cell.value != period:
                cell.number_format = "mmmyy"
        opening_row = header_row + 1
        flow_row = header_row + 2
        closing_row = header_row + 3
        calculated_row = header_row + 4
        for row_number, label in (
            (opening_row, "Opening"),
            (flow_row, "Net movement"),
            (closing_row, "Closing"),
            (calculated_row, "Calculated closing"),
        ):
            sheet.cell(row_number, 2, label)
        category_criteria = category.replace('"', '""')
        for column, period in enumerate(periods, start=3):
            period_criteria = period.replace('"', '""')
            letter = get_column_letter(column)
            if column == 3:
                opening_formula = _movement_sumifs(category_criteria, period_criteria, "OPENING")
            else:
                opening_formula = f"={get_column_letter(column - 1)}{closing_row}"
            sheet.cell(opening_row, column, opening_formula)
            sheet.cell(flow_row, column, _movement_sumifs(category_criteria, period_criteria, "FLOW"))
            sheet.cell(closing_row, column, _movement_sumifs(category_criteria, period_criteria, "CLOSING"))
            sheet.cell(calculated_row, column, f"={letter}{opening_row}+{letter}{flow_row}")
        current_row = calculated_row + 3
    sheet.freeze_panes = "C8"
    sheet.sheet_view.showGridLines = False
    for column in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    workbook.save(databook_path)


def _movement_sumifs(category: str, period: str, role: str) -> str:
    return (
        f'=SUMIFS(\'Movements\'!$I:$I,\'Movements\'!$E:$E,"{category}",'
        f'\'Movements\'!$D:$D,"{period}",\'Movements\'!$F:$F,"{role}")'
    )


def _display_period(value: str):
    for fmt in ("%Y-%m-%d", "%Y-%m", "%d-%b-%Y", "%d-%b-%y", "%b-%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return value


def _group_movements(movements: tuple[MovementRecord, ...]) -> dict[tuple[str, str], dict[str, Decimal]]:
    grouped: dict[tuple[str, str], dict[str, Decimal]] = {}
    for row in movements:
        if row.judgment.scope != Scope.IN_SCOPE or not row.judgment.category:
            continue
        bucket = grouped.setdefault((row.period, str(row.judgment.category)), {"OPENING": Decimal("0"), "FLOW": Decimal("0"), "CLOSING": Decimal("0")})
        if row.movement_role == "FLOW":
            bucket["FLOW"] += row.signed_amount
        else:
            bucket[row.movement_role] += row.signed_amount
    return grouped
