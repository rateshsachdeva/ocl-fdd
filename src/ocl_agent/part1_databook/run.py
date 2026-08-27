"""Part 1 orchestration: standardized package -> review -> dynamic OCL databook."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook

from ocl_agent.auto_semantics import ensure_semantic_handoff
from ocl_agent.part1_databook.controls import build_core_controls
from ocl_agent.part1_databook.input_contract import StandardizedPackage, discover_standardized_package, profile_package
from ocl_agent.part1_databook.judgment_validation import JudgmentIssue, validate_judgment_completion
from ocl_agent.part1_databook.judgments import JudgmentStore, load_judgments
from ocl_agent.part1_databook.movements import MovementBuildResult, build_movements, embed_rollforward, rollforward_control
from ocl_agent.part1_databook.periods import continuity_control
from ocl_agent.part1_databook.record_builder import RecordBuildResult, build_ocl_records
from ocl_agent.part1_databook.renderer import render_workbook
from ocl_agent.part1_databook.review_context import write_review_context
from ocl_agent.part1_databook.review_workbook import write_input_review, write_semantic_review
from ocl_agent.part1_databook.semantic_handoff import SemanticHandoff, load_semantic_handoff, package_id as semantic_package_id, write_semantic_handoff_draft
from ocl_agent.part1_databook.workbook_blueprint import WorkbookBlueprint, build_blueprint
from ocl_agent.schemas import CheckStatus, ControlResult


@dataclass(frozen=True)
class Part1Result:
    state: str
    package: StandardizedPackage
    judgments: JudgmentStore
    input_review: Path
    handoff_draft: Path | None = None
    handoff: SemanticHandoff | None = None
    build: RecordBuildResult | None = None
    movement_build: MovementBuildResult | None = None
    semantic_review: Path | None = None
    review_context: Path | None = None
    judgment_issues: tuple[JudgmentIssue, ...] = ()
    controls: tuple[ControlResult, ...] = ()
    blueprint: WorkbookBlueprint | None = None
    databook: Path | None = None


Part1Stage2Result = Part1Result


def run_stage2(standardized_output: Path, config_dir: Path, output_dir: Path) -> Part1Result:
    return run_part1(standardized_output, config_dir, output_dir)


def run_part1(
    standardized_output: Path,
    config_dir: Path,
    output_dir: Path,
    *,
    working_databook: Path | None = None,
    support_dir: Path | None = None,
) -> Part1Result:
    package = discover_standardized_package(standardized_output)
    profiles = profile_package(package)
    judgments = load_judgments(config_dir)
    output_dir = Path(output_dir)
    safe_package_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", semantic_package_id(package)).strip("._")[:100] or "package"
    working_databook = (
        Path(working_databook)
        if working_databook is not None
        else output_dir.parent / "work" / "ocl_runtime" / safe_package_id / "OCL_Databook_working.xlsx"
    )
    runtime_dir = working_databook.parent
    support_dir = Path(support_dir) if support_dir is not None else output_dir / "support working" / safe_package_id
    runtime_dir.mkdir(parents=True, exist_ok=True)
    support_dir.mkdir(parents=True, exist_ok=True)
    input_review = write_input_review(package, profiles, support_dir / "OCL_Input_Review.xlsx")

    # The upstream AI has already established meaning when it creates the
    # standardized canonical publication. Carry those semantics forward
    # deterministically instead of asking AI to reinterpret the same package.
    # If the package is non-canonical, this returns None and the explicit
    # semantic-review checkpoint remains the safe fallback.
    ensure_semantic_handoff(standardized_output, config_dir)

    handoff_path = Path(config_dir) / "semantic_handoff.json"
    if not handoff_path.exists():
        draft = write_semantic_handoff_draft(package, profiles, runtime_dir / "semantic_handoff_draft.json")
        return Part1Result("AWAITING_SEMANTIC_HANDOFF", package, judgments, input_review, handoff_draft=draft)
    handoff = load_semantic_handoff(handoff_path, package, profiles, require_confirmed=True)
    build = build_ocl_records(package, handoff, judgments)
    movement_build = build_movements(package, handoff, judgments, handoff_path)
    judgment_issues = validate_judgment_completion(build.records)
    movement_check = rollforward_control(movement_build.records, build.records, movement_build.alignments, movement_build.issues)
    period_check = continuity_control(build.records, handoff_path)
    controls = build_core_controls(build.records, build, handoff, judgment_issues, package, movement_control=movement_check, continuity_control=period_check)
    semantic_review = write_semantic_review(package, profiles, handoff, build, support_dir / "OCL_Stage2_Review.xlsx", judgment_issues=judgment_issues, controls=controls)
    review_context = write_review_context(package, handoff, build.records, runtime_dir / "OCL_Review_Context.json")
    if build.issues or judgment_issues:
        return Part1Result("AWAITING_JUDGMENT_REVIEW", package, judgments, input_review, handoff=handoff, build=build, movement_build=movement_build, semantic_review=semantic_review, review_context=review_context, judgment_issues=judgment_issues, controls=controls)
    blocking_controls = tuple(control for control in controls if control.status in {CheckStatus.FAIL, CheckStatus.REVIEW_REQUIRED})
    if blocking_controls:
        return Part1Result("AWAITING_CONTROL_ALIGNMENT", package, judgments, input_review, handoff=handoff, build=build, movement_build=movement_build, semantic_review=semantic_review, review_context=review_context, controls=controls)
    blueprint = build_blueprint(build.records, source_dataset_files=[path.name for path in package.datasets], has_rollforward_data=bool(movement_build.records), supported_analyses=())
    databook = render_workbook(blueprint, build.records, controls, working_databook, package=package, handoff=handoff)
    embed_rollforward(databook, movement_build.records)
    reopened = load_workbook(databook, read_only=True, data_only=False)
    expected = [sheet.title for sheet in blueprint.sheets]
    if reopened.sheetnames != expected:
        raise RuntimeError("Rendered OCL workbook sheet structure changed after reopen validation.")
    reopened.close()
    return Part1Result("DATABOOK_READY", package, judgments, input_review, handoff=handoff, build=build, movement_build=movement_build, semantic_review=semantic_review, review_context=review_context, controls=controls, blueprint=blueprint, databook=databook)
