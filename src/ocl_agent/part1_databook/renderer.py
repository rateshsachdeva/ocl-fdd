"""Deterministic Excel renderer for a dynamic OCL workbook blueprint.

This module renders reviewed decisions. It does not decide accounting meaning,
create categories, or force unsupported analyses into the workbook. Workbook
layout follows the OCL databook formatting contract so downstream formulas can
reference stable foundation ranges.
"""
from __future__ import annotations

import csv
import hashlib
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ocl_agent.part1_databook.input_contract import StandardizedPackage
from ocl_agent.part1_databook.semantic_handoff import SemanticHandoff
from ocl_agent.part1_databook.workbook_blueprint import WorkbookBlueprint
from ocl_agent.schemas import ControlResult, OCLRecord, Scope

EXCEL_MAX_DATA_ROWS = 1_048_575
PROJECT_LABEL = "TargetCo - Other Current Liabilities"
CORE_FLAT_COLUMNS = (
    "Source_Dataset", "Source_Record_ID", "Source_ID", "Source_Sheet", "Source_Cell", "Raw_Source_Sheet",
    "Entity", "Period", "Record_Usage", "Parent_Category", "Category", "Source_Label", "Source_Code", "Amount",
    "Scope", "Management_View", "FDD_View", "Normality", "Review_Status", "Judgment_Key",
)


class WorkbookRenderError(ValueError):
    pass


def render_workbook(blueprint: WorkbookBlueprint, records: Iterable[OCLRecord], controls: Iterable[ControlResult], output_path: Path, *, package: StandardizedPackage | None = None, handoff: SemanticHandoff | None = None) -> Path:
    rows = tuple(records)
    checks = tuple(controls)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_map = {spec.key: workbook.create_sheet(spec.title) for spec in blueprint.sheets}
    source_sheet_by_file = {spec.dataset_file: spec.title for spec in blueprint.sheets if spec.dataset_file}
    source_headers_by_file = _source_headers(package) if package is not None else {}
    if package is not None:
        _write_source_copies(sheet_map, blueprint, package, handoff)
    flat_sheet = sheet_map.get("flat_file")
    if flat_sheet is not None:
        _write_flat(flat_sheet, rows, source_sheet_by_file, source_headers_by_file, handoff)
    monthly_sheet = sheet_map.get("monthly_flat")
    if monthly_sheet is not None:
        monthly_rows = tuple(row for row in rows if row.dimensions.get("record_usage") == "MONTHLY_RECORDS")
        _write_flat(monthly_sheet, monthly_rows, source_sheet_by_file, source_headers_by_file, handoff)
    if "balance_by_category" in sheet_map:
        _write_balance(sheet_map["balance_by_category"], blueprint, monthly=False)
    if "monthly_balance" in sheet_map:
        _write_balance(sheet_map["monthly_balance"], blueprint, monthly=True)
    if "checks" in sheet_map:
        _write_checks(sheet_map["checks"], checks, flat_sheet)
    if "mapping" in sheet_map:
        _write_mapping(sheet_map["mapping"], rows)
    if "unmapped" in sheet_map:
        _write_unmapped(sheet_map["unmapped"], rows)
    if "scope_excluded" in sheet_map:
        _write_scope_excluded(sheet_map["scope_excluded"], rows)
    populated_keys = {"flat_file", "monthly_flat", "balance_by_category", "monthly_balance", "checks", "mapping", "unmapped", "scope_excluded"}
    for spec in blueprint.sheets:
        if spec.key not in populated_keys and not spec.dataset_file:
            sheet_map[spec.key].append(["Analysis", "Status"])
            sheet_map[spec.key].append([spec.title, "SUPPORTED_BY_BLUEPRINT"])
    for sheet in workbook.worksheets:
        _finish_sheet(sheet)
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _write_source_copies(sheet_map, blueprint: WorkbookBlueprint, package: StandardizedPackage, handoff: SemanticHandoff | None) -> None:
    amount_by_file = {binding.file: binding.fields.amount for binding in handoff.datasets if binding.fields.amount} if handoff else {}
    for spec in blueprint.sheets:
        if not spec.dataset_file:
            continue
        sheet = sheet_map[spec.key]
        path = package.root / spec.dataset_file
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            header = next(reader, None)
            if header is None:
                raise WorkbookRenderError(f"Source dataset is empty: {spec.dataset_file}")
            sheet.append(header)
            amount_name = amount_by_file.get(spec.dataset_file)
            amount_index = header.index(amount_name) if amount_name in header else None
            for data_row, values in enumerate(reader, start=1):
                if data_row > EXCEL_MAX_DATA_ROWS:
                    raise WorkbookRenderError(f"{spec.dataset_file} exceeds one Excel sheet; source copy cannot be written without partitioning.")
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(data_row + 1, column)
                    if amount_index is not None and column - 1 == amount_index and str(value).strip():
                        try:
                            cell.value = Decimal(str(value).strip().replace(",", ""))
                        except InvalidOperation:
                            cell.value = value
                            cell.data_type = "s"
                    else:
                        cell.value = value
                        if isinstance(value, str) and value.startswith("="):
                            cell.data_type = "s"
        sheet.protection.sheet = True


def _write_flat(sheet, rows: tuple[OCLRecord, ...], source_sheet_by_file: dict[str, str], source_headers_by_file: dict[str, list[str]], handoff: SemanticHandoff | None) -> None:
    excluded_dimensions = {"dataset_file", "record_usage", "standardized_csv_row", "source_code", "entity", "currency"}
    dynamic_dimensions = sorted({key for row in rows for key in row.dimensions if key not in excluded_dimensions})
    sheet.append([sheet.title])
    sheet.append([*CORE_FLAT_COLUMNS, *dynamic_dimensions])
    amount_column_by_file = {binding.file: binding.fields.amount for binding in handoff.record_bindings()} if handoff else {}
    for row in rows:
        dataset = str(row.dimensions.get("dataset_file") or "")
        source_tab = None
        source_cell = None
        amount_value: Decimal | str = row.amount
        standardized_csv_row = row.dimensions.get("standardized_csv_row")
        amount_name = amount_column_by_file.get(dataset)
        header = source_headers_by_file.get(dataset, [])
        if dataset in source_sheet_by_file and amount_name in header and standardized_csv_row:
            source_tab = source_sheet_by_file[dataset]
            amount_column = get_column_letter(header.index(amount_name) + 1)
            source_cell = f"{amount_column}{int(standardized_csv_row)}"
            source_title = source_tab.replace("'", "''")
            amount_value = f"='{source_title}'!{source_cell}"
        judgment = row.judgment
        source_code = _text(row.dimensions.get("source_code"))
        entity = _text(row.dimensions.get("entity"))
        values = [dataset, row.source.source_record_id, row.source.source_file, source_tab, source_cell, row.source.source_sheet, entity, row.period, row.dimensions.get("record_usage"), judgment.parent_category, judgment.category, row.source_label, source_code, amount_value, judgment.scope.value, judgment.management_view, judgment.fdd_view, judgment.normality, judgment.review_status.value, _judgment_key(entity, source_code, row.source_label)]
        values.extend(row.dimensions.get(key) for key in dynamic_dimensions)
        sheet.append(values)


def _source_headers(package: StandardizedPackage) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for path in package.datasets:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            result[path.name] = next(csv.reader(handle), [])
    return result


def _write_balance(sheet, blueprint: WorkbookBlueprint, *, monthly: bool) -> None:
    periods = blueprint.monthly_periods if monthly else blueprint.periods
    usage = "MONTHLY_RECORDS" if monthly else "OCL_RECORDS"
    flat_sheet = "Monthly Flat" if monthly else "Flat File"
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = sheet.title
    sheet["B6"] = "Monthly balance by category" if monthly else "Balance by category"
    for column, value in enumerate(("Category", *periods), start=2):
        sheet.cell(7, column, value)
    rows_for_total: list[int] = []
    child_set = {child for children in blueprint.hierarchy.values() for child in children}
    current_row = 8
    for parent, children in blueprint.hierarchy.items():
        child_rows: list[int] = []
        for child in children:
            sheet.cell(current_row, 2, child)
            for column, period in enumerate(periods, start=3):
                sheet.cell(current_row, column, _sumifs_formula(flat_sheet, child, period, usage))
            child_rows.append(current_row)
            rows_for_total.append(current_row)
            current_row += 1
        if child_rows:
            sheet.cell(current_row, 2, parent)
            for column in range(3, 3 + len(periods)):
                letter = get_column_letter(column)
                sheet.cell(current_row, column, f"=SUM({letter}{child_rows[0]}:{letter}{child_rows[-1]})")
            current_row += 1
    for category in blueprint.categories:
        if category in child_set or category in blueprint.hierarchy:
            continue
        sheet.cell(current_row, 2, category)
        for column, period in enumerate(periods, start=3):
            sheet.cell(current_row, column, _sumifs_formula(flat_sheet, category, period, usage))
        rows_for_total.append(current_row)
        current_row += 1
    if rows_for_total:
        sheet.cell(current_row, 2, "Total OCL")
        for column in range(3, 3 + len(periods)):
            letter = get_column_letter(column)
            sheet.cell(current_row, column, "=" + "+".join(f"{letter}{row_number}" for row_number in rows_for_total))


def _sumifs_formula(flat_sheet: str, category: str, period: str, usage: str) -> str:
    category_text = str(category).replace('"', '""')
    period_text = str(period).replace('"', '""')
    amount = _flat_col("Amount")
    category_column = _flat_col("Category")
    period_column = _flat_col("Period")
    usage_column = _flat_col("Record_Usage")
    scope_column = _flat_col("Scope")
    return (f"=SUMIFS('{flat_sheet}'!${amount}:${amount}," f"'{flat_sheet}'!${category_column}:${category_column},\"{category_text}\"," f"'{flat_sheet}'!${period_column}:${period_column},\"{period_text}\"," f"'{flat_sheet}'!${usage_column}:${usage_column},\"{usage}\"," f"'{flat_sheet}'!${scope_column}:${scope_column},\"IN_SCOPE\")")


def _write_checks(sheet, checks: tuple[ControlResult, ...], flat_sheet) -> None:
    sheet["A1"] = "All check values must be zero"
    sheet["A2"] = "Green means tie; red means break. Unsupported checks remain NOT_APPLICABLE rather than being forced."
    headers = ["Control_ID", "Python_Status", "Python_Actual", "Python_Expected", "Python_Difference", "Workbook_Difference", "Workbook_Status", "Message"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)
    for row_number, check in enumerate(checks, start=5):
        workbook_difference = None
        workbook_status = None
        if check.control_id == "chk_categories_sum" and flat_sheet is not None:
            amount = _flat_col("Amount")
            scope = _flat_col("Scope")
            category = _flat_col("Category")
            workbook_difference = (f'=SUMIFS(\'Flat File\'!${amount}:${amount},\'Flat File\'!${scope}:${scope},"IN_SCOPE",' f'\'Flat File\'!${category}:${category},"<>")-SUMIFS(\'Flat File\'!${amount}:${amount},' f'\'Flat File\'!${scope}:${scope},"IN_SCOPE")')
            workbook_status = f'=IF(ABS(F{row_number})<0.5,"PASS","FAIL")'
        values = [check.control_id, check.status.value, check.actual, check.expected, check.difference, workbook_difference, workbook_status, check.message]
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value)


def _write_mapping(sheet, rows: tuple[OCLRecord, ...]) -> None:
    sheet["A1"] = "Original source label to OCL category / scope mapping"
    headers = ["Source_Label", "Source_Code", "Entity", "Scope", "Category", "Parent_Category", "Management_View", "FDD_View", "Normality", "Review_Status", "Reason"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(2, column, value)
    seen: set[tuple[str, str, str]] = set()
    current_row = 3
    for row in rows:
        source_code = _text(row.dimensions.get("source_code"))
        entity = _text(row.dimensions.get("entity"))
        key = (row.source_label.casefold(), (source_code or "").casefold(), (entity or "").casefold())
        if key in seen:
            continue
        seen.add(key)
        judgment = row.judgment
        values = [row.source_label, source_code, entity, judgment.scope.value, judgment.category or "UNMAPPED", judgment.parent_category, judgment.management_view, judgment.fdd_view, judgment.normality, judgment.review_status.value, judgment.reason]
        for column, value in enumerate(values, start=1):
            sheet.cell(current_row, column, value)
        current_row += 1


def _write_unmapped(sheet, rows: tuple[OCLRecord, ...]) -> None:
    sheet["A1"] = "Unmapped rows are retained and excluded from mapped-category subtotals"
    sheet["A2"] = "These records remain visible for review; Total OCL must not hide an unresolved mapping gap."
    headers = ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Amount", "Source_Dataset"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(3, column, value)
    current_row = 4
    for row in rows:
        if row.judgment.scope == Scope.IN_SCOPE and not row.judgment.category:
            values = [row.source.source_record_id, row.period, row.source_label, row.dimensions.get("source_code"), row.dimensions.get("entity"), row.amount, row.dimensions.get("dataset_file")]
            for column, value in enumerate(values, start=1):
                sheet.cell(current_row, column, value)
            current_row += 1


def _write_scope_excluded(sheet, rows: tuple[OCLRecord, ...]) -> None:
    sheet["A1"] = "Out-of-scope rows are retained and excluded from OCL"
    headers = ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Scope", "Amount", "Source_Dataset"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(2, column, value)
    current_row = 3
    for row in rows:
        if row.judgment.scope != Scope.IN_SCOPE:
            values = [row.source.source_record_id, row.period, row.source_label, row.dimensions.get("source_code"), row.dimensions.get("entity"), row.judgment.scope.value, row.amount, row.dimensions.get("dataset_file")]
            for column, value in enumerate(values, start=1):
                sheet.cell(current_row, column, value)
            current_row += 1


def _flat_col(name: str) -> str:
    return get_column_letter(CORE_FLAT_COLUMNS.index(name) + 1)


def _finish_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    if sheet.title in {"Flat File", "Monthly Flat", "Mapping", "SCOPE_EXCLUDED"}:
        sheet.freeze_panes = "A3"
    elif sheet.title == "UNMAPPED":
        sheet.freeze_panes = "A4"
    elif sheet.title == "Checks":
        sheet.freeze_panes = "A5"
    elif sheet.title in {"Balance by Category", "Monthly Balance"}:
        sheet.freeze_panes = "B8"
    elif sheet.title.startswith("SRC_"):
        sheet.freeze_panes = "A2"
    else:
        sheet.freeze_panes = "A2"
    for column in range(1, sheet.max_column + 1):
        width = min(40, max(10, max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 100) + 1)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def _judgment_key(entity: str | None, code: str | None, label: str) -> str:
    text = "|".join([entity or "", code or "", label])
    return hashlib.sha256(text.casefold().encode("utf-8")).hexdigest()[:16].upper()


def _text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
