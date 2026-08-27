"""Part 3 - management Q&A grouped by commercial theme."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ocl_agent.part3_qanda.engine import build_questions
from ocl_agent.schemas import AnalysisResult, ManagementQuestion

PROJECT_LABEL = "TargetCo - Other Current Liabilities"


def run_qanda(analysis: AnalysisResult, databook_path: Path) -> tuple[ManagementQuestion, ...]:
    questions = build_questions(analysis)
    _embed_questions(Path(databook_path), questions, analysis)
    return questions


def _embed_questions(path: Path, questions: tuple[ManagementQuestion, ...], analysis: AnalysisResult) -> None:
    workbook = load_workbook(path)
    for name in ("Management Questions", "Q&A"):
        if name in workbook.sheetnames:
            del workbook[name]
    sheet = workbook.create_sheet("Q&A")
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = "Management Q&A"
    sheet["B6"] = "Material management questions arising from the OCL analysis"
    headers = ["#", "FDD Lens", "Theme", "Question", "Why it matters", "Evidence trigger", "Management Response"]
    for column, value in enumerate(headers, start=2):
        sheet.cell(7, column, value)

    finding_by_id = {finding.finding_id: finding for finding in analysis.findings}
    ordered = sorted(questions, key=lambda q: (_theme(finding_by_id.get(q.linked_finding_id)), q.question))
    for number, item in enumerate(ordered, start=1):
        finding = finding_by_id.get(item.linked_finding_id)
        row = number + 7
        theme = _theme(finding)
        values = [number, theme, theme, item.question, item.rationale, _evidence(finding, item), ""]
        for column, value in enumerate(values, start=2):
            sheet.cell(row, column, value)

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B8"
    sheet.column_dimensions["A"].width = 5
    workbook.save(path)


def _theme(finding) -> str:
    if finding is None:
        return "Other OCL matters"
    return {
        "DEBT_LIKE": "Net debt & equity value",
        "DEBT_LIKE_GAP": "Net debt & equity value",
        "ONE_OFF": "Quality of earnings",
        "SEASONALITY": "Seasonality & phasing",
        "MONTHLY_VARIABILITY": "Seasonality & phasing",
        "STALE_BALANCE": "Working capital & balance validity",
        "NEW_ITEM": "Completeness & balance validity",
        "CLIFF": "Quality of earnings",
        "CATEGORY_MOVEMENT": "Balance movements",
        "TOTAL_CHANGE": "Balance movements",
        "CONCENTRATION": "Balance composition",
    }.get(finding.finding_type, "Other OCL matters")


def _evidence(finding, question: ManagementQuestion) -> str:
    if finding is None:
        return question.rationale
    references = ", ".join(finding.evidence_references)
    parts = [finding.text, question.rationale]
    if references:
        parts.append(f"References: {references}")
    return " | ".join(part for part in parts if part)
