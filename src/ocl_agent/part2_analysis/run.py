"""Part 2 - Python metrics plus formula-linked Excel analysis layer.

Python computes analytical numbers independently from the reconciled Part 1
model. Excel analysis tabs point back to formula-driven foundation schedules so
every displayed figure remains traceable. Presentation follows the OCL databook
formatting guide; AI may improve explanation but never alter calculated amounts
or materiality decisions.
"""
from __future__ import annotations

from datetime import date, datetime
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ocl_agent.part1_databook.input_contract import StandardizedPackage
from ocl_agent.part1_databook.semantic_handoff import SemanticHandoff
from ocl_agent.part2_analysis.context import enrich_with_context, load_context
from ocl_agent.part2_analysis.diagnostics import diagnostic_findings
from ocl_agent.part2_analysis.engine import DATABOOK_PERCENT_THRESHOLD, analyse_records
from ocl_agent.schemas import AnalysisResult, Finding, OCLRecord
from ocl_agent.workbook_hierarchy import copy_row_outline

PROJECT_LABEL = "TargetCo - Other Current Liabilities"
ANALYSIS_SHEETS = ("Analysis Summary", "Seasonality", "Item Monthly Charts", "Deal Issues", "Key Findings")


def run_analysis(records: Iterable[OCLRecord], databook_path: Path, *, package: StandardizedPackage | None = None, handoff: SemanticHandoff | None = None) -> AnalysisResult:
    rows = tuple(records)
    base = analyse_records(rows)
    extra = diagnostic_findings(rows)
    seen = {item.finding_id for item in base.findings}
    result = AnalysisResult(
        (*base.findings, *(item for item in extra if item.finding_id not in seen)),
        base.tables,
        base.annual_periods,
        base.monthly_periods,
        base.latest_annual_period,
    )
    if package is not None and handoff is not None:
        result = enrich_with_context(result, rows, load_context(package, handoff))
    _embed_analysis(Path(databook_path), result, handoff)
    return result


def _embed_analysis(path: Path, result: AnalysisResult, handoff: SemanticHandoff | None = None) -> None:
    workbook = load_workbook(path)
    for name in ANALYSIS_SHEETS:
        if name in workbook.sheetnames:
            del workbook[name]

    summary = _analysis_sheet(workbook, "Analysis Summary", "Formula-linked monthly OCL statistics")
    _write_formula_linked_monthly_stats(summary, workbook, 6, handoff)

    if "Monthly Balance" in workbook.sheetnames:
        layout = _balance_layout(workbook["Monthly Balance"])
        if layout and len(layout[2]) >= 12:
            _write_seasonality(workbook)
            _write_monthly_charts(workbook)

    _write_deal_issues(workbook, result.findings)
    _write_key_findings(workbook, result.findings)

    for name in ANALYSIS_SHEETS:
        if name in workbook.sheetnames:
            _finish_sheet(workbook[name])
    workbook.save(path)


def _analysis_sheet(workbook, name: str, purpose: str):
    sheet = workbook.create_sheet(name)
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = name
    sheet["B6"] = purpose
    sheet.column_dimensions["A"].width = 5
    return sheet


def _write_formula_linked_annual(summary, workbook, title_row: int) -> int:
    if "Balance by Category" not in workbook.sheetnames:
        return title_row
    source = workbook["Balance by Category"]
    layout = _balance_layout(source)
    if layout is None:
        return title_row
    header_row, category_col, period_cols = layout
    summary.cell(title_row, 2, "Annual OCL balance by category")
    periods = [source.cell(header_row, col).value for col in period_cols]
    add_movement = len(period_cols) >= 2
    header = ["Category", *periods]
    if add_movement:
        header.extend(["Movement", "Movement %", "Review Flag"])
    header_row_target = title_row + 1
    for col, value in enumerate(header, start=2):
        summary.cell(header_row_target, col, value)
    target_row = header_row_target + 1
    for source_row in range(header_row + 1, source.max_row + 1):
        label = source.cell(source_row, category_col).value
        if label in (None, ""):
            continue
        summary.cell(target_row, 2, f"='Balance by Category'!{get_column_letter(category_col)}{source_row}")
        for offset, source_col in enumerate(period_cols, start=3):
            summary.cell(target_row, offset, f"='Balance by Category'!{get_column_letter(source_col)}{source_row}")
        if add_movement:
            previous_source_col, latest_source_col = period_cols[-2], period_cols[-1]
            movement_col = 3 + len(period_cols)
            pct_col = movement_col + 1
            flag_col = pct_col + 1
            summary.cell(target_row, movement_col, f"='Balance by Category'!{get_column_letter(latest_source_col)}{source_row}-'Balance by Category'!{get_column_letter(previous_source_col)}{source_row}")
            summary.cell(target_row, pct_col, f'=IFERROR({get_column_letter(movement_col)}{target_row}/ABS(\'Balance by Category\'!{get_column_letter(previous_source_col)}{source_row}),"")')
            summary.cell(target_row, flag_col, f'=IF(OR(ABS({get_column_letter(movement_col)}{target_row})>=100000,ABS({get_column_letter(pct_col)}{target_row})>={DATABOOK_PERCENT_THRESHOLD/100}),"REVIEW","")')
        target_row += 1
    return target_row


def _write_formula_linked_monthly_stats(summary, workbook, title_row: int, handoff: SemanticHandoff | None = None) -> int:
    if "Monthly Balance" not in workbook.sheetnames:
        return title_row
    source = workbook["Monthly Balance"]
    layout = _balance_layout(source)
    if layout is None:
        return title_row
    header_row, category_col, period_cols = layout
    summary.cell(title_row, 2, "Monthly OCL statistics by category")
    blocks = _monthly_year_blocks(source, header_row, period_cols, handoff)
    if not blocks:
        return title_row
    year_row = title_row + 1
    metric_row = title_row + 2
    summary.cell(year_row, 2, "Category")
    summary.merge_cells(start_row=year_row, start_column=2, end_row=metric_row, end_column=2)
    target_column = 3
    for label, _columns in blocks:
        summary.cell(year_row, target_column, label)
        summary.merge_cells(start_row=year_row, start_column=target_column, end_row=year_row, end_column=target_column + 3)
        for offset, metric in enumerate(("Average", "Minimum", "Maximum", "Latest")):
            summary.cell(metric_row, target_column + offset, metric)
        target_column += 4
    target_row = title_row + 3
    for source_row in range(header_row + 1, source.max_row + 1):
        label = source.cell(source_row, category_col).value
        if label in (None, ""):
            continue
        summary.cell(target_row, 2, f"='Monthly Balance'!{get_column_letter(category_col)}{source_row}")
        target_column = 3
        for _block_label, columns in blocks:
            first_letter = get_column_letter(columns[0])
            last_letter = get_column_letter(columns[-1])
            source_range = f"'Monthly Balance'!{first_letter}{source_row}:{last_letter}{source_row}"
            summary.cell(target_row, target_column, f"=AVERAGE({source_range})")
            summary.cell(target_row, target_column + 1, f"=MIN({source_range})")
            summary.cell(target_row, target_column + 2, f"=MAX({source_range})")
            summary.cell(target_row, target_column + 3, f"='Monthly Balance'!{last_letter}{source_row}")
            target_column += 4
        copy_row_outline(source, source_row, summary, target_row)
        target_row += 1
    return target_row


def _monthly_year_blocks(source, header_row: int, period_cols: list[int], handoff: SemanticHandoff | None):
    """Return up to three source-backed fiscal/calendar period blocks."""
    column_by_text = {str(source.cell(header_row, column).value): column for column in period_cols}
    aligned: list[tuple[str, int]] = []
    if handoff is not None:
        for item in handoff.monthly_to_annual:
            column = column_by_text.get(str(item.monthly_period))
            if column is not None:
                aligned.append((str(item.annual_period), column))
    if aligned:
        aligned.sort(key=lambda item: period_cols.index(item[1]))
        blocks = []
        prior_index = -1
        for label, end_column in aligned:
            end_index = period_cols.index(end_column)
            start_index = prior_index + 1 if prior_index >= 0 else max(0, end_index - 11)
            columns = period_cols[start_index : end_index + 1]
            if columns:
                blocks.append((label, columns))
            prior_index = end_index
        return blocks[-3:]

    by_year: dict[int, list[int]] = {}
    for column in period_cols:
        key = _period_date(source.cell(header_row, column).value)
        if key is not None:
            by_year.setdefault(key.year, []).append(column)
    return [(str(year), by_year[year]) for year in sorted(by_year)[-3:]]


def _period_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%b-%y", "%b-%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _write_seasonality(workbook) -> None:
    source = workbook["Monthly Balance"]
    layout = _balance_layout(source)
    if layout is None:
        return
    source_header_row, category_col, period_cols = layout
    period_cols = period_cols[-12:]
    sheet = _analysis_sheet(workbook, "Seasonality", "Year-end balance compared with the trailing 12-month average")
    headers = ["Category", *[source.cell(source_header_row, col).value for col in period_cols], "12M Average", "Year End", "YE vs Avg", "Peak Month", "Flag"]
    for col, value in enumerate(headers, start=2):
        sheet.cell(7, col, value)
    target_row = 8
    for source_row in range(source_header_row + 1, source.max_row + 1):
        label = source.cell(source_row, category_col).value
        if label in (None, ""):
            continue
        sheet.cell(target_row, 2, f"='Monthly Balance'!{get_column_letter(category_col)}{source_row}")
        month_start_col = 3
        for offset, source_col in enumerate(period_cols, start=month_start_col):
            sheet.cell(target_row, offset, f"='Monthly Balance'!{get_column_letter(source_col)}{source_row}")
        first_month = get_column_letter(month_start_col)
        last_month = get_column_letter(month_start_col + len(period_cols) - 1)
        avg_col = month_start_col + len(period_cols)
        year_end_col = avg_col + 1
        deviation_col = year_end_col + 1
        peak_col = deviation_col + 1
        flag_col = peak_col + 1
        sheet.cell(target_row, avg_col, f"=AVERAGE({first_month}{target_row}:{last_month}{target_row})")
        sheet.cell(target_row, year_end_col, f"={last_month}{target_row}")
        sheet.cell(target_row, deviation_col, f'=IFERROR({get_column_letter(year_end_col)}{target_row}/{get_column_letter(avg_col)}{target_row}-1,"")')
        sheet.cell(target_row, peak_col, f"=INDEX({first_month}$7:{last_month}$7,1,MATCH(MAX({first_month}{target_row}:{last_month}{target_row}),{first_month}{target_row}:{last_month}{target_row},0))")
        sheet.cell(target_row, flag_col, f'=IF({get_column_letter(deviation_col)}{target_row}>15%,"YEAR-END SPIKE",IF({get_column_letter(deviation_col)}{target_row}<-15%,"YEAR-END DIP",""))')
        copy_row_outline(source, source_row, sheet, target_row)
        target_row += 1


def _write_monthly_charts(workbook) -> None:
    source = workbook["Monthly Balance"]
    layout = _balance_layout(source)
    if layout is None:
        return
    source_header_row, category_col, period_cols = layout
    sheet = _analysis_sheet(workbook, "Item Monthly Charts", "Monthly balance and LTM 12-month average by category")
    periods = [source.cell(source_header_row, col).value for col in period_cols]
    sheet["B6"] = "Monthly amounts"
    for col, value in enumerate(("Category", *periods), start=2):
        sheet.cell(7, col, value)
    amount_rows: list[tuple[str, int, int]] = []
    target_row = 8
    for source_row in range(source_header_row + 1, source.max_row + 1):
        label = source.cell(source_row, category_col).value
        if label in (None, ""):
            continue
        sheet.cell(target_row, 2, f"='Monthly Balance'!{get_column_letter(category_col)}{source_row}")
        for offset, source_col in enumerate(period_cols, start=3):
            sheet.cell(target_row, offset, f"='Monthly Balance'!{get_column_letter(source_col)}{source_row}")
        amount_rows.append((str(label), source_row, target_row))
        target_row += 1

    ltm_title_row = target_row + 2
    sheet.cell(ltm_title_row, 2, "LTM 12-month average")
    for col, value in enumerate(("Category", *periods), start=2):
        sheet.cell(ltm_title_row + 1, col, value)
    ltm_rows: dict[str, int] = {}
    target_row = ltm_title_row + 2
    for label, source_row, _amount_row in amount_rows:
        sheet.cell(target_row, 2, f"='Monthly Balance'!{get_column_letter(category_col)}{source_row}")
        for index, source_col in enumerate(period_cols, start=1):
            target_col = index + 2
            if index < 12:
                sheet.cell(target_row, target_col, "")
            else:
                first = get_column_letter(source_col - 11)
                last = get_column_letter(source_col)
                sheet.cell(target_row, target_col, f"=AVERAGE('Monthly Balance'!{first}{source_row}:{last}{source_row})")
        ltm_rows[label] = target_row
        target_row += 1

    chart_anchor_row = target_row + 3
    for index, (label, _source_row, amount_row) in enumerate(amount_rows):
        ltm_row = ltm_rows[label]
        bar = BarChart()
        bar.type = "col"
        bar.height = 6.35
        bar.width = 30.48
        bar.title = label
        bar.legend.position = "b"
        bar.y_axis.delete = True
        bar.y_axis.majorGridlines = None
        bar.display_blanks = "gap"
        categories = Reference(sheet, min_col=3, max_col=2 + len(periods), min_row=7)
        amount_data = Reference(sheet, min_col=3, max_col=2 + len(periods), min_row=amount_row, max_row=amount_row)
        bar.add_data(amount_data, from_rows=True, titles_from_data=False)
        bar.set_categories(categories)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = "00338D"
            bar.series[0].graphicalProperties.line.solidFill = "00338D"
        line = LineChart()
        line_data = Reference(sheet, min_col=3, max_col=2 + len(periods), min_row=ltm_row, max_row=ltm_row)
        line.add_data(line_data, from_rows=True, titles_from_data=False)
        if line.series:
            line.series[0].graphicalProperties.line.solidFill = "FFC000"
        bar += line
        sheet.add_chart(bar, f"B{chart_anchor_row + index * 16}")


def _write_deal_issues(workbook, findings: tuple[Finding, ...]) -> None:
    sheet = workbook.create_sheet("Deal Issues")
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = "Key deal issues this block answers"
    if not findings:
        sheet["A4"] = "No material deal issue identified from the available evidence"
        sheet["A4"].font = Font(bold=True)
        sheet["A5"] = "The deterministic analysis did not identify a material OCL deal issue from the evidence supplied."
        return
    row = 4
    for finding in findings:
        sheet.cell(row, 1, _deal_issue_title(finding))
        sheet.cell(row, 1).font = Font(bold=True)
        sheet.cell(row + 1, 1, f"FDD implication / So what: {_so_what(finding)}")
        sheet.cell(row + 2, 1, f"Evidence: {finding.text}")
        sheet.cell(row + 3, 1, "Evidence limitation: The deterministic evidence does not establish facts beyond the stated analysis.")
        sheet.cell(row + 4, 1, f"Fact to establish: {_ask_management(finding)}")
        row += 6


def _write_key_findings(workbook, findings: tuple[Finding, ...]) -> None:
    sheet = _analysis_sheet(workbook, "Key Findings", "Material evidence-led OCL findings")
    headers = ["ID", "FDD Lens", "Area", "Metric", "FY periods / Item", "FDD implication / So what", "Evidence", "Evidence limitation", "Fact to establish", "Materiality"]
    for col, value in enumerate(headers, start=2):
        sheet.cell(7, col, value)
    for row, finding in enumerate(findings, start=8):
        values = [
            finding.finding_id,
            _theme_for_finding(finding),
            _theme_for_finding(finding),
            finding.finding_type,
            ", ".join(finding.evidence_references),
            _so_what(finding),
            finding.text,
            "The deterministic evidence does not establish facts beyond the stated analysis.",
            _ask_management(finding),
            str(finding.metrics.get("materiality") or "MATERIAL"),
        ]
        for col, value in enumerate(values, start=2):
            sheet.cell(row, col, value)


def _finding_formula(workbook, finding: Finding):
    metrics = finding.metrics
    category = metrics.get("category")
    latest = metrics.get("latest_period") or metrics.get("period")
    previous = metrics.get("previous_period")
    if "Balance by Category" in workbook.sheetnames:
        source = workbook["Balance by Category"]
        layout = _balance_layout(source)
        if layout:
            header_row, category_col, _period_cols = layout
            row = _find_row(source, category if category else "Total OCL", column=category_col, start_row=header_row + 1)
            latest_col = _find_col(source, latest, row=header_row)
            previous_col = _find_col(source, previous, row=header_row)
            if row and latest_col and previous_col and finding.finding_type in {"TOTAL_CHANGE", "CATEGORY_MOVEMENT"}:
                return f"='Balance by Category'!{get_column_letter(latest_col)}{row}-'Balance by Category'!{get_column_letter(previous_col)}{row}"
            if row and latest_col and finding.finding_type == "CONCENTRATION":
                return f"='Balance by Category'!{get_column_letter(latest_col)}{row}"
    if "Seasonality" in workbook.sheetnames and category and finding.finding_type == "SEASONALITY":
        source = workbook["Seasonality"]
        row = _find_row(source, category, column=2, start_row=8)
        if row:
            header = _find_col(source, "YE vs Avg", row=7)
            if header:
                return f"='Seasonality'!{get_column_letter(header)}{row}"
    if "Flat File" in workbook.sheetnames and latest:
        flat = workbook["Flat File"]
        header_row, headers = _header_map(flat, "Amount")
        amount = headers.get("Amount")
        period = headers.get("Period")
        scope = headers.get("Scope")
        fdd_view = headers.get("FDD_View")
        management_view = headers.get("Management_View")
        normality = headers.get("Normality")
        if header_row and amount and period and scope and fdd_view and finding.finding_type == "DEBT_LIKE":
            return f'=SUMIFS(\'Flat File\'!${amount}:${amount},\'Flat File\'!${period}:${period},"{latest}",\'Flat File\'!${scope}:${scope},"IN_SCOPE",\'Flat File\'!${fdd_view}:${fdd_view},"debt_like")'
        if header_row and amount and period and scope and fdd_view and management_view and finding.finding_type == "DEBT_LIKE_GAP":
            return f'=SUMIFS(\'Flat File\'!${amount}:${amount},\'Flat File\'!${period}:${period},"{latest}",\'Flat File\'!${scope}:${scope},"IN_SCOPE",\'Flat File\'!${fdd_view}:${fdd_view},"debt_like")-SUMIFS(\'Flat File\'!${amount}:${amount},\'Flat File\'!${period}:${period},"{latest}",\'Flat File\'!${scope}:${scope},"IN_SCOPE",\'Flat File\'!${management_view}:${management_view},"debt_like")'
        if header_row and amount and period and scope and normality and finding.finding_type == "ONE_OFF":
            return f'=SUMIFS(\'Flat File\'!${amount}:${amount},\'Flat File\'!${period}:${period},"{latest}",\'Flat File\'!${scope}:${scope},"IN_SCOPE",\'Flat File\'!${normality}:${normality},"one_off")'
    return ""


def _balance_layout(sheet):
    for row in range(1, min(sheet.max_row, 12) + 1):
        for col in range(1, min(sheet.max_column, 5) + 1):
            if str(sheet.cell(row, col).value or "").strip().casefold() == "category":
                period_cols = [c for c in range(col + 1, sheet.max_column + 1) if sheet.cell(row, c).value not in (None, "")]
                return row, col, period_cols
    return None


def _header_map(sheet, required: str):
    for row in range(1, min(sheet.max_row, 10) + 1):
        headers = {str(sheet.cell(row, col).value): get_column_letter(col) for col in range(1, sheet.max_column + 1) if sheet.cell(row, col).value not in (None, "")}
        if required in headers:
            return row, headers
    return None, {}


def _find_row(sheet, value, *, column: int = 1, start_row: int = 1) -> int | None:
    if value in (None, ""):
        return None
    wanted = str(value).strip().casefold()
    for row in range(start_row, sheet.max_row + 1):
        if str(sheet.cell(row, column).value or "").strip().casefold() == wanted:
            return row
    return None


def _find_col(sheet, value, *, row: int = 1) -> int | None:
    if value in (None, ""):
        return None
    wanted = str(value).strip().casefold()
    for column in range(1, sheet.max_column + 1):
        if str(sheet.cell(row, column).value or "").strip().casefold() == wanted:
            return column
    return None


def _theme_for_finding(finding: Finding) -> str:
    return {
        "DEBT_LIKE": "Net debt & equity value",
        "DEBT_LIKE_GAP": "Net debt & equity value",
        "ONE_OFF": "Quality of earnings",
        "SEASONALITY": "Seasonality & phasing",
        "MONTHLY_VARIABILITY": "Seasonality & phasing",
        "STALE_BALANCE": "Working capital & balance validity",
        "NEW_ITEM": "Completeness & balance validity",
        "CLIFF": "Quality of earnings",
        "CATEGORY_MOVEMENT": "Balance movements",
        "TOTAL_CHANGE": "Balance movements",
        "CONCENTRATION": "Balance composition",
    }.get(finding.finding_type, "Other OCL matters")


def _deal_issue_title(finding: Finding) -> str:
    return {
        "DEBT_LIKE": "Debt-like items within OCL",
        "DEBT_LIKE_GAP": "FDD vs management debt-like classification gap",
        "ONE_OFF": "One-off / non-recurring OCL items",
        "SEASONALITY": "Year-end balance may not be representative",
        "MONTHLY_VARIABILITY": "Material monthly volatility",
        "STALE_BALANCE": "Potential stale accrual",
        "NEW_ITEM": "New closing obligation",
        "CLIFF": "Balance released or settled to nil",
    }.get(finding.finding_type, finding.title)


def _so_what(finding: Finding) -> str:
    return {
        "DEBT_LIKE": "Could affect net debt / equity value depending on the agreed transaction definition; factual settlement profile should be evidenced.",
        "DEBT_LIKE_GAP": "The difference between management and FDD treatment is a potential deal-value reconciliation item.",
        "ONE_OFF": "May not represent the recurring level of operating liabilities and should be considered when assessing normalized working capital.",
        "SEASONALITY": "A year-end working-capital peg based only on the closing month may not represent the normal in-year level.",
        "MONTHLY_VARIABILITY": "A single balance-sheet date may not represent the underlying run-rate and phasing should be understood.",
        "STALE_BALANCE": "An unchanged liability may indicate a genuinely long-dated obligation or an accrual requiring reassessment.",
        "NEW_ITEM": "The obligation is new versus the prior period and its origin and expected settlement should be understood.",
        "CLIFF": "The reduction to nil may reflect settlement, release or reversal and can affect recurring earnings / working-capital interpretation.",
        "CATEGORY_MOVEMENT": "The movement is sufficiently large to warrant understanding the operational or accounting driver.",
        "TOTAL_CHANGE": "The overall OCL movement is material and should be reconciled to the principal category-level drivers.",
        "CONCENTRATION": "A large share of OCL sits in one category, increasing the importance of understanding its composition and settlement timing.",
    }.get(finding.finding_type, "The item is material to the OCL analysis and requires evidence-led interpretation.")


def _ask_management(finding: Finding) -> str:
    metrics = finding.metrics
    if finding.finding_type == "DEBT_LIKE_GAP":
        return "Please explain the factual basis for the difference between management's and the FDD debt-like classification, including expected settlement timing."
    if finding.finding_type == "SEASONALITY":
        return f"Please explain the operational reason the {metrics.get('category')} year-end balance differs materially from its trailing 12-month average."
    if finding.finding_type == "STALE_BALANCE":
        return f"Please confirm whether the unchanged {metrics.get('source_label')} balance remains a valid outstanding obligation and when it is expected to settle."
    if finding.finding_type == "NEW_ITEM":
        return f"Please explain the event or calculation that gave rise to the new {metrics.get('source_label')} balance."
    if finding.finding_type == "CLIFF":
        return f"Please explain whether the prior {metrics.get('source_label')} balance was settled, released or reversed and the basis for that treatment."
    if finding.finding_type == "CATEGORY_MOVEMENT":
        return f"Please explain the principal driver of the movement in {metrics.get('category')} and whether the closing level is expected to recur."
    if finding.finding_type == "TOTAL_CHANGE":
        return "Please explain the principal operational or accounting drivers of the overall OCL movement."
    if finding.finding_type == "DEBT_LIKE":
        return "Please describe the underlying obligations and expected settlement dates for the items identified as debt-like in the FDD review."
    if finding.finding_type == "ONE_OFF":
        return "Please explain the specific events giving rise to the items identified as one-off / non-recurring."
    if finding.finding_type == "CONCENTRATION":
        return f"Please explain the principal components and settlement profile of {metrics.get('category')}."
    if finding.finding_type == "MONTHLY_VARIABILITY":
        return f"Please explain the primary operational driver of the monthly volatility in {metrics.get('category')}."
    return "Please explain the underlying driver and expected settlement profile of this item."


def _finish_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B8" if sheet.title not in {"Deal Issues"} else "A4"
    sheet.column_dimensions["A"].width = 5 if sheet.title != "Deal Issues" else 90
    for column in range(2, sheet.max_column + 1):
        width = 14
        for row in range(1, min(sheet.max_row, 150) + 1):
            value = sheet.cell(row, column).value
            if value is not None:
                width = max(width, min(55, len(str(value)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width
    if sheet.title == "Deal Issues":
        for row in sheet.iter_rows():
            alignment = copy(row[0].alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            row[0].alignment = alignment
    elif sheet.title in {"Key Findings", "Q&A"}:
        headers = {str(sheet.cell(7, column).value): column for column in range(1, sheet.max_column + 1)}
        narrative = {
            "FDD implication / So what", "Evidence", "Evidence limitation", "Fact to establish",
            "Question", "Why it matters", "Evidence trigger",
        }
        for header in narrative:
            column = headers.get(header)
            if not column:
                continue
            sheet.column_dimensions[get_column_letter(column)].width = 50
            for row in range(8, sheet.max_row + 1):
                cell = sheet.cell(row, column)
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment
