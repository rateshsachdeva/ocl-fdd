"""Render evidence-aware extended analysis without changing the base workbook logic."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ocl_agent.schemas import AnalysisResult, AnalysisTable

PROJECT_LABEL = "TargetCo - Other Current Liabilities"


def embed_extended_analysis(path: Path, result: AnalysisResult) -> None:
    workbook = load_workbook(path)
    _write_coverage_sheet(workbook, _table(result, "analysis_coverage"))
    _append_formula_linked_run_rate(workbook)
    _write_context_sheet(workbook, result)
    workbook.save(path)


def _write_coverage_sheet(workbook, table: AnalysisTable | None) -> None:
    if "Analysis Coverage" in workbook.sheetnames:
        del workbook["Analysis Coverage"]
    if table is None:
        return
    sheet = workbook.create_sheet("Analysis Coverage")
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = "Analysis Coverage"
    sheet["B6"] = "Which FDD analyses are supported by the evidence actually supplied"
    for col, header in enumerate(table.headers, start=2):
        sheet.cell(7, col, header)
        sheet.cell(7, col).font = Font(bold=True)
    for row_index, row in enumerate(table.rows, start=8):
        for col_index, value in enumerate(row, start=2):
            sheet.cell(row_index, col_index, value)
    _finish(sheet)


def _append_formula_linked_run_rate(workbook) -> None:
    if "Analysis Summary" not in workbook.sheetnames or "Monthly Balance" not in workbook.sheetnames:
        return
    source = workbook["Monthly Balance"]
    layout = _balance_layout(source)
    if layout is None:
        return
    header_row, category_col, period_cols = layout
    if len(period_cols) < 4:
        return

    summary = workbook["Analysis Summary"]
    title_row = summary.max_row + 3
    summary.cell(title_row, 2, "Year-end build, recurrence and normalization references")
    headers = [
        "Category",
        "Prior 3M Average",
        "Latest",
        "YE Build / (Unwind)",
        "YE Build %",
        "12M Average",
        "Latest vs 12M Avg %",
        "Non-zero Months %",
        "Pattern",
        "Interpretation",
    ]
    for col, value in enumerate(headers, start=2):
        summary.cell(title_row + 1, col, value)

    target_row = title_row + 2
    latest_col = period_cols[-1]
    prior_3 = period_cols[-4:-1]
    trailing_12 = period_cols[-12:] if len(period_cols) >= 12 else period_cols
    latest_letter = get_column_letter(latest_col)
    prior_first = get_column_letter(prior_3[0])
    prior_last = get_column_letter(prior_3[-1])
    trail_first = get_column_letter(trailing_12[0])
    trail_last = get_column_letter(trailing_12[-1])

    for source_row in range(header_row + 1, source.max_row + 1):
        label = source.cell(source_row, category_col).value
        if label in (None, ""):
            continue
        summary.cell(target_row, 2, f"='Monthly Balance'!{get_column_letter(category_col)}{source_row}")
        summary.cell(target_row, 3, f"=AVERAGE('Monthly Balance'!{prior_first}{source_row}:{prior_last}{source_row})")
        summary.cell(target_row, 4, f"='Monthly Balance'!{latest_letter}{source_row}")
        summary.cell(target_row, 5, f"=D{target_row}-C{target_row}")
        summary.cell(target_row, 6, f'=IFERROR(E{target_row}/ABS(C{target_row}),"")')
        if len(period_cols) >= 12:
            summary.cell(target_row, 7, f"=AVERAGE('Monthly Balance'!{trail_first}{source_row}:{trail_last}{source_row})")
            summary.cell(target_row, 8, f'=IFERROR(D{target_row}/ABS(G{target_row})-1,"")')
            summary.cell(target_row, 9, f'=COUNTIF(\'Monthly Balance\'!{trail_first}{source_row}:{trail_last}{source_row},"<>0")/{len(trailing_12)}')
            summary.cell(target_row, 10, f'=IF(I{target_row}>=75%,"PERSISTENT_BALANCE","INTERMITTENT_BALANCE")')
            summary.cell(target_row, 11, "Reference only - not an FDD adjustment")
        else:
            summary.cell(target_row, 7, "")
            summary.cell(target_row, 8, "")
            summary.cell(target_row, 9, "")
            summary.cell(target_row, 10, "INSUFFICIENT_12M_HISTORY")
            summary.cell(target_row, 11, "Year-end build supported; 12M normalization reference unavailable")
        target_row += 1


def _write_context_sheet(workbook, result: AnalysisResult) -> None:
    tables = [table for table in result.tables if table.key in {"context_ratios", "accrual_to_expense", "movement_patterns"}]
    if "Additional Analysis" in workbook.sheetnames:
        del workbook["Additional Analysis"]
    if not tables:
        return
    sheet = workbook.create_sheet("Additional Analysis")
    sheet["A1"] = PROJECT_LABEL
    sheet["A2"] = "Additional Analysis"
    row = 6
    for table in tables:
        sheet.cell(row, 2, table.title)
        row += 1
        for col, header in enumerate(table.headers, start=2):
            sheet.cell(row, col, header)
            sheet.cell(row, col).font = Font(bold=True)
        row += 1
        for values in table.rows:
            for col, value in enumerate(values, start=2):
                sheet.cell(row, col, value)
            row += 1
        row += 2
    _finish(sheet)


def _table(result: AnalysisResult, key: str) -> AnalysisTable | None:
    return next((table for table in result.tables if table.key == key), None)


def _balance_layout(sheet):
    for row in range(1, min(sheet.max_row, 12) + 1):
        for col in range(1, min(sheet.max_column, 5) + 1):
            if str(sheet.cell(row, col).value or "").strip().casefold() == "category":
                period_cols = [c for c in range(col + 1, sheet.max_column + 1) if sheet.cell(row, c).value not in (None, "")]
                return row, col, period_cols
    return None


def _finish(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B8"
    sheet.column_dimensions["A"].width = 5
    for column in range(2, sheet.max_column + 1):
        width = 14
        for row in range(1, min(sheet.max_row, 200) + 1):
            value = sheet.cell(row, column).value
            if value is not None:
                width = max(width, min(60, len(str(value)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width
