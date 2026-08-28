"""Render validated AI-host narrative into Deal Issues, Key Findings and Q&A.

All financial figures remain linked to deterministic workbook schedules. The AI
host supplies interpretation and question wording only after the Python analysis
is finalized.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from ocl_agent.part2_analysis.run import _analysis_sheet, _finish_sheet
from ocl_agent.project_title import resolve_project_title
from ocl_agent.schemas import AnalysisResult, ManagementQuestion


def apply_partner_interpretation(
    path: Path,
    result: AnalysisResult,
    interpretation: dict,
) -> tuple[ManagementQuestion, ...]:
    path = Path(path)
    workbook = load_workbook(path)
    for name in ("Deal Issues", "Key Findings", "Management Questions", "Q&A"):
        if name in workbook.sheetnames:
            del workbook[name]

    finding_by_id = {item.finding_id: item for item in result.findings}
    _write_deal_issues(workbook, finding_by_id, interpretation)
    _write_key_findings(workbook, finding_by_id, interpretation)
    questions = _write_qanda(workbook, interpretation)
    _finish_sheet(workbook["Deal Issues"])
    _finish_sheet(workbook["Key Findings"])
    _finish_sheet(workbook["Q&A"])
    workbook.save(path)
    return questions


def _write_deal_issues(workbook, finding_by_id: dict, interpretation: dict) -> None:
    sheet = workbook.create_sheet("Deal Issues")
    sheet["A1"] = resolve_project_title(workbook=workbook)
    sheet["A2"] = "Key deal issues — FDD partner interpretation of finalized analysis"
    issues = interpretation.get("deal_issues") or []
    if not issues:
        sheet["A4"] = "No material deal issue identified from the available evidence"
        sheet["A4"].font = Font(bold=True)
        sheet["A5"] = str(interpretation.get("overall_assessment") or "")
        return

    row = 4
    for issue in issues:
        priority = str(issue.get("priority") or "MEDIUM").upper()
        lens = str(issue.get("fdd_lens") or "")
        sheet.cell(row, 1, f"{priority} | {issue['title']} | {lens}")
        sheet.cell(row, 1).font = Font(bold=True)
        sheet.cell(row + 1, 1, f"FDD implication / So what: {issue['so_what']}")
        sheet.cell(row + 2, 1, f"Evidence: {issue['evidence']}")
        sheet.cell(row + 3, 1, f"Evidence limitation: {issue['evidence_limit']}")
        sheet.cell(row + 4, 1, f"Fact to establish: {issue['management_focus']}")
        row += 6


def _write_key_findings(workbook, finding_by_id: dict, interpretation: dict) -> None:
    sheet = _analysis_sheet(workbook, "Key Findings", "FDD partner interpretation of finalized Python OCL analysis")
    headers = [
        "ID",
        "FDD Lens",
        "Area",
        "Metric",
        "FY periods / Item",
        "FDD implication / So what",
        "Evidence",
        "Evidence limitation",
        "Fact to establish",
        "Materiality",
    ]
    for col, value in enumerate(headers, start=2):
        sheet.cell(7, col, value)

    for row, item in enumerate(interpretation.get("key_findings") or [], start=8):
        values = [
            item["id"],
            item["fdd_lens"],
            item["area"],
            item["metric"],
            item["period_item"],
            item["so_what"],
            item["evidence"],
            item["evidence_limit"],
            item["fact_to_establish"],
            item["materiality"],
        ]
        for col, value in enumerate(values, start=2):
            sheet.cell(row, col, value)


def _write_qanda(workbook, interpretation: dict) -> tuple[ManagementQuestion, ...]:
    sheet = _analysis_sheet(workbook, "Q&A", "FDD partner questions arising only from finalized Python OCL analysis")
    headers = ["#", "FDD Lens", "Theme", "Question", "Why it matters", "Evidence trigger", "Management Response"]
    for col, value in enumerate(headers, start=2):
        sheet.cell(7, col, value)

    output: list[ManagementQuestion] = []
    items = interpretation.get("management_questions") or []
    if not items:
        sheet.cell(8, 2, "-")
        sheet.cell(8, 3, "No material management question was identified from the available evidence.")
        sheet.cell(8, 4, str(interpretation.get("overall_assessment") or ""))
        return ()

    for number, item in enumerate(items, start=1):
        row = number + 7
        sheet.cell(row, 2, number)
        sheet.cell(row, 3, item["fdd_lens"])
        sheet.cell(row, 4, item["theme"])
        sheet.cell(row, 5, item["question"])
        sheet.cell(row, 6, item["why_it_matters"])
        sheet.cell(row, 7, item["evidence"])
        sheet.cell(row, 8, "")
        output.append(
            ManagementQuestion(
                question_id=str(item["id"]),
                question=str(item["question"]),
                rationale=f"{item['why_it_matters']} Evidence: {item['evidence']}",
                evidence_references=tuple(str(ref) for ref in item.get("evidence_refs") or []),
                linked_finding_id=(str(item["linked_finding_id"]) if item.get("linked_finding_id") else None),
                priority=str(item.get("priority") or "MEDIUM").upper(),
            )
        )
    return tuple(output)
