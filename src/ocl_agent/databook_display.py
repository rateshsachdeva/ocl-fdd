"""Final presentation preferences for the generated OCL databook.

This module is presentation-only. It may change visible formatting and hide
structurally redundant subtotal rows, but it must not change financial values,
formulas, scope, mappings, controls or FDD conclusions.
"""
from __future__ import annotations

import calendar
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel
from ocl_agent.workbook_hierarchy import apply_collapsed_detail_group

BLACK = "000000"
PERIOD_FORMAT = "mmmyy"
FORMULA_SUM_RANGE = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.IGNORECASE)


def apply_databook_display_preferences(path: Path, handoff: Any | None = None) -> Path:
    """Apply the user-facing display contract to a completed databook."""
    path = Path(path)
    workbook = load_workbook(path)
    apply_display_preferences_to_workbook(workbook, handoff)
    workbook.save(path)
    return path


def apply_display_preferences_to_workbook(workbook, handoff: Any | None = None) -> None:
    """Apply final display preferences to an already-open workbook."""
    annual_alignment = _annual_alignment(handoff)

    for sheet in workbook.worksheets:
        _format_period_headers(sheet, annual_alignment)

    for name in ("Balance by Category", "Monthly Balance", "Seasonality", "Analysis Summary"):
        if name not in workbook.sheetnames:
            continue
        sheet = workbook[name]
        _ensure_outline_groups(sheet)

    if "Item Monthly Charts" in workbook.sheetnames:
        _format_monthly_charts(workbook["Item Monthly Charts"])

def _annual_alignment(handoff: Any | None) -> dict[str, str]:
    result: dict[str, str] = {}
    if handoff is None:
        return result
    items = handoff.get("monthly_to_annual", ()) if isinstance(handoff, dict) else getattr(handoff, "monthly_to_annual", ())
    for item in items or ():
        annual_value = item.get("annual_period") if isinstance(item, dict) else getattr(item, "annual_period", None)
        monthly_value = item.get("monthly_period") if isinstance(item, dict) else getattr(item, "monthly_period", None)
        annual = str(annual_value or "").strip()
        monthly = str(monthly_value or "").strip()
        if annual and monthly:
            result[annual] = monthly
    return result


def _format_period_headers(sheet, annual_alignment: dict[str, str]) -> None:
    """Format period headings as Excel dates when a source-backed date is known."""
    for row in range(1, min(sheet.max_row, 12) + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            value = cell.value
            display_date = None
            if isinstance(value, datetime):
                display_date = value.date()
            elif isinstance(value, date):
                display_date = value
            elif isinstance(value, str):
                text = value.strip()
                mapped = annual_alignment.get(text, text)
                display_date = _parse_period_end(mapped)
            if display_date is None:
                continue
            cell.value = display_date
            cell.number_format = PERIOD_FORMAT
            font = copy(cell.font)
            font.color = BLACK
            cell.font = font


def _parse_period_end(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None

    month_match = re.fullmatch(r"(\d{4})-(\d{2})", text)
    if month_match:
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        if 1 <= month <= 12:
            return date(year, month, calendar.monthrange(year, month)[1])
        return None

    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _format_monthly_charts(sheet) -> None:
    """Keep the monthly balance/LTM legend visible and bars closely spaced."""
    for chart in getattr(sheet, "_charts", []):
        charts = getattr(chart, "_charts", None) or [chart]
        bar_chart = charts[0]
        if hasattr(bar_chart, "gapWidth"):
            bar_chart.gapWidth = 40
        if hasattr(chart, "gapWidth"):
            chart.gapWidth = 40

        # The combined chart uses the root legend. Recreate it explicitly so
        # Excel always renders it even when the component charts had no titles.
        chart.legend = Legend(legendPos="b", overlay=False)

        if getattr(bar_chart, "ser", None):
            bar_chart.ser[0].tx = SeriesLabel(v="Monthly balance")
        if len(charts) > 1 and getattr(charts[1], "ser", None):
            charts[1].ser[0].tx = SeriesLabel(v="LTM 12M average")


def _find_category_layout(sheet) -> tuple[int, int, int] | None:
    for row in range(1, min(sheet.max_row, 15) + 1):
        for column in range(1, min(sheet.max_column, 6) + 1):
            if str(sheet.cell(row, column).value or "").strip().casefold() == "category":
                total_row = None
                for candidate in range(row + 1, sheet.max_row + 1):
                    if str(sheet.cell(candidate, column).value or "").strip().casefold() == "total ocl":
                        total_row = candidate
                        break
                if total_row is not None:
                    return row, column, total_row
    return None


def _subtotal_range(sheet, row: int, first_value_col: int) -> tuple[int, int] | None:
    ranges: list[tuple[int, int]] = []
    for column in range(first_value_col, sheet.max_column + 1):
        value = sheet.cell(row, column).value
        if value in (None, ""):
            continue
        if not isinstance(value, str):
            return None
        match = FORMULA_SUM_RANGE.fullmatch(value.strip())
        if not match:
            return None
        if match.group(1).upper() != match.group(3).upper():
            return None
        ranges.append((int(match.group(2)), int(match.group(4))))
    if not ranges or len(set(ranges)) != 1:
        return None
    return ranges[0]


def _ensure_outline_groups(sheet) -> None:
    layout = _find_category_layout(sheet)
    if layout is None:
        return
    header_row, category_col, total_row = layout
    first_value_col = category_col + 1

    for row in range(header_row + 1, total_row):
        if sheet.cell(row, category_col).value in (None, ""):
            continue
        range_rows = _subtotal_range(sheet, row, first_value_col)
        if range_rows is not None and range_rows[0] > row:
            apply_collapsed_detail_group(sheet, row, range_rows[0], range_rows[1])
