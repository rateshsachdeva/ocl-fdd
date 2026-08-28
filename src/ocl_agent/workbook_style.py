"""Presentation-only styling for the final OCL databook.

The style layer may improve layout and review usability but must never create or
change categories, periods, financial values, controls or conclusions.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_NAVY = "002060"
KPMG_BLUE = "00338D"
LIGHT_GREY = "E1E4E2"
GRAND_TOTAL = "E5E5E5"
SOURCE_GREY = "808080"
NOTE_GREY = "7F7F7F"
INPUT_BLUE = "0000FF"
BLACK = "000000"
WHITE = "FFFFFF"
AMBER = "FFF2CC"
PASS_FILL = "C6EFCE"
PASS_FONT = "006100"
FAIL_FILL = "FFC7CE"
FAIL_FONT = "9C0006"
ACCOUNTING = '#,##0;[Red](#,##0);-'
PERCENT = '0.0%;[Red](0.0%);-'
MULTIPLE = '0.0x'
THIN_GREY = Side(style="thin", color="BFBFBF")
MEDIUM_BLUE = Side(style="medium", color=KPMG_BLUE)
WIDTH_SAMPLE_ROWS = 120


@lru_cache(maxsize=None)
def _font(color: str = BLACK, bold: bool = False, italic: bool = False, size: int = 8) -> Font:
    """Return one immutable common font instead of rebuilding it per cell."""
    return Font(name="Arial", size=size, color=color, bold=bold, italic=italic)


@lru_cache(maxsize=None)
def _alignment(
    vertical: str = "center",
    horizontal: str | None = None,
    wrap_text: bool = False,
    indent: int = 0,
) -> Alignment:
    """Return one immutable common alignment for a presentation role."""
    return Alignment(vertical=vertical, horizontal=horizontal, wrap_text=wrap_text, indent=indent)


HEADER_GREY_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
SECTION_BLUE_FILL = PatternFill("solid", fgColor=KPMG_BLUE)
TOTAL_GREY_FILL = PatternFill("solid", fgColor=GRAND_TOTAL)
AMBER_FILL = PatternFill("solid", fgColor=AMBER)
PASS_CELL_FILL = PatternFill("solid", fgColor=PASS_FILL)
FAIL_CELL_FILL = PatternFill("solid", fgColor=FAIL_FILL)
HEADER_GREY_BORDER = Border(bottom=THIN_GREY)
SECTION_BLUE_BORDER = Border(bottom=MEDIUM_BLUE)
TOTAL_BORDER = Border(top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)
PARENT_BORDER = Border(top=THIN_GREY)

FRONT_ORDER = [
    "Deal Issues", "Key Findings", "Q&A", "Checks", "Balance by Category",
    "Monthly Balance", "Roll-forward", "Seasonality", "Item Monthly Charts", "Analysis Summary",
]
SUPPORT_ORDER = ["Flat File", "Movements", "TB", "Monthly Flat", "Mapping", "UNMAPPED", "SCOPE_EXCLUDED"]


def apply_workbook_style(path: Path) -> Path:
    path = Path(path)
    workbook = load_workbook(path)
    style_workbook(workbook)
    workbook.save(path)
    return path


def style_workbook(workbook) -> None:
    """Apply the style contract to an already-open workbook."""
    for sheet in workbook.worksheets:
        _base_sheet(sheet)
        if sheet.title.startswith("SRC_"):
            _source_sheet(sheet)
        elif sheet.title in {"Flat File", "Monthly Flat"}:
            _flat_sheet(sheet)
        elif sheet.title in {"Balance by Category", "Monthly Balance"}:
            _balance_sheet(sheet)
        elif sheet.title == "Roll-forward":
            _rollforward_sheet(sheet)
        elif sheet.title == "Checks":
            _checks_sheet(sheet)
        elif sheet.title == "Mapping":
            _mapping_sheet(sheet)
        elif sheet.title == "UNMAPPED":
            _unmapped_sheet(sheet)
        elif sheet.title == "SCOPE_EXCLUDED":
            _scope_excluded_sheet(sheet)
        elif sheet.title in {"Analysis Summary", "Seasonality", "Item Monthly Charts", "Analysis Coverage", "Additional Analysis", "Key Findings", "Q&A"}:
            _analysis_sheet(sheet)
        elif sheet.title == "Deal Issues":
            _deal_issues_sheet(sheet)
        else:
            _generic_sheet(sheet)
    _reorder_sheets(workbook)
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass


def _base_sheet(sheet) -> None:
    """Apply sheet/page defaults without repainting every populated cell."""
    sheet.sheet_view.showGridLines = False
    sheet.print_options.gridLines = False
    try:
        sheet.print_options.gridLinesSet = True
    except AttributeError:
        pass
    sheet.sheet_view.zoomScale = 90
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    if sheet.max_row:
        sheet.row_dimensions[1].height = 18


def style_generated_support_cell(cell, *, role: str, accounting: bool = False) -> None:
    """Style a large generated support cell once, while it is being written.

    ``source`` cells are protected source hardcodes; ``linked`` cells are flat
    file lineage/model cells; ``model`` cells are deterministic support data.
    """
    value = cell.value
    is_formula = isinstance(value, str) and value.startswith("=")
    is_numeric = isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)
    is_date = isinstance(value, (date, datetime))
    if role == "source":
        color = BLACK if is_formula else INPUT_BLUE
    elif role == "linked":
        color = BLACK if is_formula or is_numeric or is_date else INPUT_BLUE
    else:
        color = BLACK
    cell.font = _font(color)
    cell.alignment = _alignment()
    if accounting:
        cell.number_format = ACCOUNTING


def _support_body_is_preformatted(
    sheet,
    start_row: int,
    role: str,
    amount_col: int | None = None,
    *,
    max_row: int | None = None,
    max_col: int | None = None,
) -> bool:
    """Check a bounded representative sample before any compatibility repaint."""
    max_row = sheet.max_row if max_row is None else max_row
    max_col = sheet.max_column if max_col is None else max_col
    if max_row < start_row or max_col < 1:
        return True
    candidates = {start_row, min(max_row, start_row + 1), max_row, max(start_row, (start_row + max_row) // 2)}
    for row in sorted(candidates):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            value = cell.value
            is_formula = isinstance(value, str) and value.startswith("=")
            is_numeric = isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)
            is_date = isinstance(value, (date, datetime))
            if role == "source":
                # A completed presentation pass converts source period values
                # in the bounded header area to black Excel dates. Freshly
                # rendered CSV strings remain protected source blue until then.
                expected_color = BLACK if is_formula or (row <= 12 and is_date) else INPUT_BLUE
            elif role == "linked":
                expected_color = BLACK if is_formula or is_numeric or is_date else INPUT_BLUE
            else:
                expected_color = BLACK
            color = cell.font.color
            rgb = str(color.rgb or "") if color and color.type == "rgb" else ""
            if cell.font.name != "Arial" or cell.font.sz != 8 or not rgb.upper().endswith(expected_color):
                return False
            if cell.alignment.vertical != "center" or bool(cell.alignment.wrap_text):
                return False
            if amount_col == col and value not in (None, "") and cell.number_format != ACCOUNTING:
                return False
    return True


def _ensure_generated_support_body(
    sheet,
    start_row: int,
    role: str,
    amount_col: int | None = None,
    *,
    max_row: int | None = None,
    max_col: int | None = None,
) -> None:
    max_row = sheet.max_row if max_row is None else max_row
    max_col = sheet.max_column if max_col is None else max_col
    if _support_body_is_preformatted(
        sheet,
        start_row,
        role,
        amount_col,
        max_row=max_row,
        max_col=max_col,
    ):
        return
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            style_generated_support_cell(
                sheet.cell(row, col),
                role=role,
                accounting=bool(amount_col == col and sheet.cell(row, col).value not in (None, "")),
            )


def _title(sheet, title_row: int = 1, subtitle_row: int | None = 2) -> None:
    if sheet.cell(title_row, 1).value not in (None, ""):
        sheet.cell(title_row, 1).font = _font(TITLE_NAVY, True, size=14)
    if subtitle_row and sheet.cell(subtitle_row, 1).value not in (None, ""):
        sheet.cell(subtitle_row, 1).font = _font(BLACK, True)


def _header_row(sheet, row: int, start_col: int = 1, end_col: int | None = None, *, blue: bool = False) -> None:
    end_col = end_col or sheet.max_column
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        if cell.value in (None, "") and not blue:
            continue
        cell.fill = SECTION_BLUE_FILL if blue else HEADER_GREY_FILL
        cell.font = _font(WHITE if blue else BLACK, True)
        cell.alignment = _alignment(horizontal="right" if col > start_col else "left", wrap_text=True)
        cell.border = SECTION_BLUE_BORDER if blue else HEADER_GREY_BORDER


def _section_row(sheet, row: int, start_col: int = 2, end_col: int | None = None) -> None:
    end_col = end_col or max(start_col, sheet.max_column)
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = SECTION_BLUE_FILL
        cell.font = _font(WHITE, True)
        cell.border = SECTION_BLUE_BORDER


def _source_sheet(sheet) -> None:
    _header_row(sheet, 1, blue=False)
    sheet.sheet_properties.tabColor = SOURCE_GREY
    sheet.freeze_panes = "A2"
    _ensure_generated_support_body(sheet, 2, "source", _source_amount_column(sheet))
    sheet.protection.sheet = True
    _reasonable_widths(sheet, 28)


def _flat_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    _title(sheet, 1, None)
    _header_row(sheet, 2, end_col=max_col)
    sheet.freeze_panes = "A3"
    sheet.sheet_properties.tabColor = KPMG_BLUE
    headers = _headers(sheet, 2)
    amount_col = headers.get("Amount")
    _ensure_generated_support_body(sheet, 3, "linked", amount_col, max_row=max_row, max_col=max_col)
    _reasonable_widths(sheet, 34)
    _set_width(sheet, headers.get("Source_Record_ID"), 26)
    _set_width(sheet, headers.get("Source_Label"), 30)


def _source_amount_column(sheet) -> int | None:
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(1, col).value or "").strip().casefold() in {"amount", "raw_amount", "signed_amount"}:
            return col
    return None


def _balance_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    _title(sheet)
    sheet.column_dimensions["A"].width = 5
    _section_row(sheet, 6, 2, max_col)
    _header_row(sheet, 7, 2, max_col)
    sheet.freeze_panes = "B8"
    sheet.sheet_properties.tabColor = KPMG_BLUE
    for row in range(8, max_row + 1):
        label = str(sheet.cell(row, 2).value or "")
        formula_cells = [sheet.cell(row, col) for col in range(3, max_col + 1) if isinstance(sheet.cell(row, col).value, str) and sheet.cell(row, col).value.startswith("=")]
        is_total = label.casefold() in {"total ocl", "total mapped ocl", "total"}
        is_parent = bool(formula_cells) and any(str(cell.value).upper().startswith("=SUM(") for cell in formula_cells) and not is_total
        if is_total:
            for col in range(2, max_col + 1):
                cell = sheet.cell(row, col)
                cell.fill = TOTAL_GREY_FILL
                cell.font = _font(BLACK, True)
                cell.border = TOTAL_BORDER
        elif is_parent:
            for col in range(2, max_col + 1):
                cell = sheet.cell(row, col)
                cell.font = _font(BLACK, True)
                cell.border = PARENT_BORDER
        else:
            sheet.cell(row, 2).font = _font(BLACK)
            sheet.cell(row, 2).alignment = _alignment(indent=1)
        for col in range(3, max_col + 1):
            cell = sheet.cell(row, col)
            cell.number_format = ACCOUNTING
            cell.alignment = _alignment(horizontal="right")
            cell.font = _font(BLACK, is_parent or is_total)
    sheet.column_dimensions["B"].width = 30
    for col in range(3, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13


def _rollforward_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    sheet.sheet_properties.tabColor = KPMG_BLUE
    _title(sheet)
    sheet.column_dimensions["A"].width = 5
    sheet.freeze_panes = "C8"
    header_rows: set[int] = set()
    for row in range(6, max_row + 1):
        label = sheet.cell(row, 2).value
        if label in (None, ""):
            continue
        if row in header_rows:
            continue
        other_values = [sheet.cell(row, col).value for col in range(3, max_col + 1)]
        if all(value in (None, "") for value in other_values):
            _section_row(sheet, row, 2, max_col)
            if row + 1 <= max_row:
                header_rows.add(row + 1)
                _header_row(sheet, row + 1, 2, max_col)
            continue
        for col in range(3, max_col + 1):
            cell = sheet.cell(row, col)
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = ACCOUNTING
                cell.alignment = _alignment(horizontal="right")
                cell.font = _font(BLACK)
        sheet.cell(row, 2).font = _font(BLACK)
    sheet.column_dimensions["B"].width = 22
    for col in range(3, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13


def _checks_sheet(sheet) -> None:
    _title(sheet, 1, 2)
    sheet.cell(2, 1).font = _font(NOTE_GREY)
    _header_row(sheet, 4)
    sheet.freeze_panes = "A5"
    sheet.sheet_properties.tabColor = KPMG_BLUE
    headers = _headers(sheet, 4)
    for row in range(5, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.font = _font(BLACK)
            cell.alignment = _alignment()
        for name in ("Python_Actual", "Python_Expected", "Python_Difference", "Workbook_Difference"):
            col = headers.get(name)
            if col:
                sheet.cell(row, col).number_format = ACCOUNTING
        for name in ("Python_Status", "Workbook_Status"):
            col = headers.get(name)
            if col:
                _status_cell(sheet.cell(row, col), str(sheet.cell(row, col).value or ""))
    _set_width(sheet, headers.get("Control_ID"), 26)
    _set_width(sheet, headers.get("Message"), 70)


def _mapping_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    _title(sheet, 1, None)
    _header_row(sheet, 2, end_col=max_col)
    sheet.freeze_panes = "A3"
    headers = _headers(sheet, 2)
    for row in range(3, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            value = cell.value
            black = (
                isinstance(value, str) and value.startswith("=")
                or isinstance(value, (int, float, Decimal, date, datetime)) and not isinstance(value, bool)
            )
            cell.font = _font(BLACK if black else INPUT_BLUE)
            cell.alignment = _alignment()
        status_col = headers.get("Review_Status")
        if status_col and str(sheet.cell(row, status_col).value or "").upper() != "REVIEWED":
            for col in range(1, max_col + 1):
                sheet.cell(row, col).fill = AMBER_FILL
    _reasonable_widths(sheet, 36)
    _set_width(sheet, headers.get("Reason"), 55)


def _unmapped_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    _title(sheet, 1, 2)
    sheet.cell(2, 1).font = _font(NOTE_GREY)
    _header_row(sheet, 3, end_col=max_col)
    sheet.freeze_panes = "A4"
    sheet.sheet_properties.tabColor = "C65911"
    headers = _headers(sheet, 3)
    amount_col = headers.get("Amount")
    source_col = headers.get("Source_Dataset")
    for row in range(4, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            cell.fill = AMBER_FILL
            cell.font = _font(NOTE_GREY if col == source_col else BLACK)
            cell.alignment = _alignment()
        if amount_col:
            sheet.cell(row, amount_col).number_format = ACCOUNTING
    _reasonable_widths(sheet, 38)


def _scope_excluded_sheet(sheet) -> None:
    _title(sheet, 1, None)
    _header_row(sheet, 2)
    sheet.freeze_panes = "A3"
    sheet.sheet_properties.tabColor = SOURCE_GREY
    headers = _headers(sheet, 2)
    amount_col = headers.get("Amount")
    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).font = _font(BLACK)
            sheet.cell(row, col).alignment = _alignment()
        if amount_col:
            sheet.cell(row, amount_col).number_format = ACCOUNTING
    _reasonable_widths(sheet, 36)


def _analysis_sheet(sheet) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    _title(sheet)
    sheet.column_dimensions["A"].width = 5
    sheet.sheet_properties.tabColor = KPMG_BLUE
    sheet.freeze_panes = "B8"
    section_rows = [row for row in range(3, max_row + 1) if sheet.cell(row, 2).value not in (None, "") and all(sheet.cell(row, col).value in (None, "") for col in range(3, min(max_col, 8) + 1))]
    for row in section_rows:
        if row + 1 <= max_row:
            _section_row(sheet, row, 2, max_col)
            _header_row(sheet, row + 1, 2, max_col)
    if sheet.title in {"Key Findings", "Q&A", "Analysis Summary", "Seasonality", "Item Monthly Charts"}:
        _section_row(sheet, 6, 2, max_col)
        _header_row(sheet, 7, 2, max_col)
    data_start = 8
    if sheet.title == "Analysis Summary" and any(sheet.merged_cells.ranges):
        _header_row(sheet, 8, 2, max_col)
        data_start = 9
        sheet.freeze_panes = "B9"
    for row in range(data_start, max_row + 1):
        wrap = sheet.title in {"Key Findings", "Q&A"}
        for col in range(2, max_col + 1):
            cell = sheet.cell(row, col)
            numeric = _numeric_like(cell)
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            fill_rgb = str(cell.fill.fgColor.rgb or "") if cell.fill.fgColor.type == "rgb" else ""
            if not is_formula and fill_rgb.upper().endswith(KPMG_BLUE):
                cell.font = _font(WHITE, True)
            else:
                cell.font = _font(BLACK) if is_formula else _font(BLACK, bool(cell.font.bold), bool(cell.font.italic))
            cell.alignment = _alignment(
                vertical="top" if wrap else "center",
                horizontal="right" if numeric else "left",
                wrap_text=wrap,
            )
            if numeric:
                header = str(sheet.cell(7, col).value or sheet.cell(8, col).value or "")
                if "%" in header or header == "YE vs Avg":
                    cell.number_format = PERCENT
                else:
                    cell.number_format = ACCOUNTING
        if wrap:
            sheet.row_dimensions[row].height = min(120, max(24, 12 + 12 * max(1, _wrapped_lines(sheet, row, max_col))))
    headers = _headers(sheet, 7)
    flag_col = headers.get("Flag") or headers.get("Review Flag")
    if flag_col:
        for row in range(data_start, max_row + 1):
            if sheet.cell(row, flag_col).value not in (None, ""):
                sheet.cell(row, flag_col).fill = PatternFill("solid", fgColor=AMBER)
    _analysis_widths(sheet)
    if sheet.title in {"Seasonality", "Analysis Summary"}:
        _style_analysis_hierarchy(sheet, data_start, max_row, max_col)


def _deal_issues_sheet(sheet) -> None:
    _title(sheet)
    sheet.sheet_properties.tabColor = KPMG_BLUE
    sheet.freeze_panes = "A4"
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        first = sheet.cell(row, 1)
        first.alignment = _alignment(vertical="top", wrap_text=True)
        if first.value and row >= 4 and (row - 4) % 6 == 0:
            first.font = _font(BLACK, True)
        elif first.value and row >= 4:
            first.font = _font(BLACK, bool(first.font.bold), bool(first.font.italic))
        if first.value and row >= 4:
            sheet.row_dimensions[row].height = max(15, min(90, 15 * (len(str(first.value)) // 100 + 1)))
    sheet.column_dimensions["A"].width = 90


def _generic_sheet(sheet) -> None:
    _header_row(sheet, 1)
    _ensure_generated_support_body(sheet, 2, "model")
    _reasonable_widths(sheet, 30)


def _status_cell(cell, status: str) -> None:
    status = status.upper()
    if status == "PASS":
        cell.fill = PASS_CELL_FILL
        cell.font = _font(PASS_FONT, True)
    elif status == "FAIL":
        cell.fill = FAIL_CELL_FILL
        cell.font = _font(FAIL_FONT, True)
    elif status == "REVIEW_REQUIRED":
        cell.fill = AMBER_FILL
        cell.font = _font(BLACK, True)
    elif status == "NOT_APPLICABLE":
        cell.fill = HEADER_GREY_FILL
        cell.font = _font(NOTE_GREY, True)


def _headers(sheet, row: int) -> dict[str, int]:
    return {str(sheet.cell(row, col).value): col for col in range(1, sheet.max_column + 1) if sheet.cell(row, col).value not in (None, "")}


def _find_header(sheet, value: str) -> int | None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    wanted = value.casefold()
    for row in range(1, min(max_row, 12) + 1):
        for col in range(1, min(max_col, 6) + 1):
            if str(sheet.cell(row, col).value or "").strip().casefold() == wanted:
                return row
    return None


def _numeric_like(cell) -> bool:
    if isinstance(cell.value, (int, float)):
        return True
    return cell.column > 2 and isinstance(cell.value, str) and cell.value.startswith("=")


def _wrapped_lines(sheet, row: int, max_col: int) -> int:
    total = 1
    for col in range(2, max_col + 1):
        text = str(sheet.cell(row, col).value or "")
        total = max(total, max(1, len(text) // 55 + 1))
    return total


def _analysis_widths(sheet) -> None:
    for col in range(2, sheet.max_column + 1):
        header = str(sheet.cell(7, col).value or "")
        width = 14
        if header in {"FDD implication / So what", "Evidence", "Evidence limitation", "Fact to establish", "Question", "Why it matters", "Evidence trigger"}:
            width = 50
        elif header in {"Theme", "Area", "Metric", "FY periods / Item"}:
            width = 22
        elif header in {"Category", "Analysis"}:
            width = 28
        sheet.column_dimensions[get_column_letter(col)].width = width


def _style_analysis_hierarchy(sheet, data_start: int, max_row: int, max_col: int) -> None:
    for row in range(data_start, max_row + 1):
        label = str(sheet.cell(row, 2).value or "")
        dimension = sheet.row_dimensions[row]
        is_total = label.casefold() == "total ocl" or "'!B" in label and row == max_row
        if is_total:
            for col in range(2, max_col + 1):
                cell = sheet.cell(row, col)
                cell.fill = TOTAL_GREY_FILL
                cell.font = _font(BLACK, True)
                cell.border = TOTAL_BORDER
        elif dimension.collapsed:
            for col in range(2, max_col + 1):
                sheet.cell(row, col).font = _font(BLACK, True)
        elif dimension.outlineLevel:
            sheet.cell(row, 2).alignment = _alignment(indent=1)


def _reasonable_widths(sheet, max_width: int = 40) -> None:
    """Set usable widths from a fixed 120-row sample, never the full data set."""
    max_row = sheet.max_row
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        width = 10
        for row in range(1, min(max_row, WIDTH_SAMPLE_ROWS) + 1):
            value = sheet.cell(row, col).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
        sheet.column_dimensions[get_column_letter(col)].width = width


def _set_width(sheet, column: int | None, width: int) -> None:
    if column:
        sheet.column_dimensions[get_column_letter(column)].width = width


def _reorder_sheets(workbook) -> None:
    ordered_names: list[str] = []
    for name in FRONT_ORDER + SUPPORT_ORDER:
        if name in workbook.sheetnames and name not in ordered_names:
            ordered_names.append(name)
    for name in workbook.sheetnames:
        if name.startswith("SRC_") and name not in ordered_names:
            ordered_names.append(name)
    for name in workbook.sheetnames:
        if name not in ordered_names:
            ordered_names.append(name)
    workbook._sheets = [workbook[name] for name in ordered_names]
