import inspect
import json
import shutil
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from ocl_agent import end_to_end
from ocl_agent.config import RepoPaths
from ocl_agent.output_versioning import PublishedDatabook
from ocl_agent.schemas import AnalysisResult, AnalysisTable
from ocl_agent.workflow_checkpoint import checkpoint_matches


def _case(tmp_path: Path):
    repo_root = Path(__file__).resolve().parents[1]
    package = tmp_path / "package"
    package.mkdir()
    (package / "standardized.csv").write_text("Category,Amount\nBonus,1\n", encoding="utf-8")
    (package / "execution_manifest.json").write_text(
        json.dumps({"execution_id": "PKG-001", "final_execution_status": "COMPLETED", "outputs_created": ["standardized.csv"]}),
        encoding="utf-8",
    )
    (package / "databook_metadata.json").write_text(json.dumps({"workflow_run_id": "PKG-001"}), encoding="utf-8")
    output = tmp_path / "output"
    config = tmp_path / "config"
    source = tmp_path / "references" / "source"
    other = tmp_path / "references" / "other"
    for folder in (output, config, source, other):
        folder.mkdir(parents=True, exist_ok=True)
    for name in ("judgment_scope.csv", "mapping.csv", "judgment_wc_debt.csv"):
        (config / name).write_text(f"header\n{name}\n", encoding="utf-8")
    paths = RepoPaths(repo_root, repo_root / "assets", config, output, source.parent, source, other)
    runtime_config = paths.work / "ocl_config" / "PKG-001"
    runtime_config.mkdir(parents=True)
    (runtime_config / "semantic_handoff.json").write_text(
        json.dumps({"handoff_version": "1.0", "status": "CONFIRMED", "package_id": "PKG-001", "datasets": [], "monthly_to_annual": []}),
        encoding="utf-8",
    )
    return paths, package, runtime_config


def _minimal_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    flat = workbook.active
    flat.title = "Flat File"
    flat.append(["OCL records"])
    flat.append(["Source_Record_ID", "Amount", "Scope", "Review_Status"])
    flat.append(["R1", 1, "IN_SCOPE", "REVIEWED"])
    checks = workbook.create_sheet("Checks")
    checks.cell(4, 1, "Control_ID")
    checks.cell(4, 2, "Python_Status")
    checks.cell(5, 1, "chk_categories_sum")
    checks.cell(5, 2, "PASS")
    workbook.create_sheet("Mapping")
    workbook.create_sheet("UNMAPPED")
    workbook.create_sheet("SCOPE_EXCLUDED")
    workbook.save(path)


def _analysis() -> AnalysisResult:
    return AnalysisResult(
        (),
        (AnalysisTable("monthly_review", "Monthly review", ("Category", "Amount"), (("Bonus", "1"),)),),
        annual_periods=("FY25",),
        monthly_periods=(),
        latest_annual_period="FY25",
    )


def _install_stage_spies(monkeypatch, counters):
    def fake_part1(_package, _config, _output, *, working_databook, support_dir):
        counters["part1"] += 1
        _minimal_workbook(working_databook)
        support_dir.mkdir(parents=True, exist_ok=True)
        (support_dir / "OCL_Stage2_Review.xlsx").write_bytes(b"review")
        return SimpleNamespace(
            state="DATABOOK_READY",
            databook=working_databook,
            build=SimpleNamespace(records=()),
            package=None,
            handoff=SimpleNamespace(monthly_to_annual=()),
            movement_build=None,
        )

    def fake_analysis(*_args, **_kwargs):
        counters["analysis"] += 1
        return _analysis()

    monkeypatch.setattr(end_to_end, "run_part1", fake_part1)
    monkeypatch.setattr(end_to_end, "run_analysis", fake_analysis)


def _write_interpretation(request_path: Path, *, evidence_hash: str | None = None) -> None:
    request = json.loads(request_path.read_text(encoding="utf-8"))
    payload = {
        "status": "COMPLETED",
        "evidence_hash": evidence_hash or request["evidence_hash"],
        "overall_assessment": "No material OCL deal issue was identified from the evidence supplied.",
        "deal_issues": [],
        "key_findings": [
            {
                "id": "KF_01",
                "fdd_lens": "Normalized working capital",
                "area": "Overall assessment",
                "metric": "Available evidence",
                "period_item": "FY25",
                "so_what": "No material issue was identified.",
                "evidence": "The deterministic monthly review was considered.",
                "evidence_limit": "Only supplied evidence was assessed.",
                "fact_to_establish": "No further fact required from current evidence.",
                "materiality": "NO_MATERIAL_ISSUE",
                "linked_finding_id": None,
                "evidence_refs": ["table:monthly_review:0"],
            }
        ],
        "management_questions": [],
    }
    request_path.with_name("analysis_interpretation.json").write_text(json.dumps(payload), encoding="utf-8")


def test_resume_skips_part1_and_analysis_then_publishes_databook_only(tmp_path: Path, monkeypatch):
    paths, package, _runtime_config = _case(tmp_path)
    counters = {"part1": 0, "analysis": 0}
    _install_stage_spies(monkeypatch, counters)

    first = end_to_end.run_end_to_end(paths, data_prep_output=package)
    assert first.state == "AWAITING_ANALYSIS_INTERPRETATION"
    assert first.working_databook == paths.ocl_runtime / "PKG-001" / "OCL_Databook_working.xlsx"
    assert first.working_databook.is_file()
    evidence = paths.work / "analysis" / "PKG-001" / "analysis_evidence.json"
    assert evidence.is_file()
    checkpoint = json.loads(first.checkpoint.read_text(encoding="utf-8"))
    assert checkpoint["completed_stage"] == "ANALYSIS_READY"
    _write_interpretation(evidence)
    (paths.output / "OCL_Databook_v1.xlsx").write_bytes(b"historical")

    second = end_to_end.run_end_to_end(paths, data_prep_output=package)

    assert second.state == "READY"
    assert counters == {"part1": 1, "analysis": 1}
    assert second.published_version == 2
    assert second.databook == paths.output / "OCL_Databook_v2.xlsx"
    assert (paths.output / "OCL_Databook_v1.xlsx").read_bytes() == b"historical"
    assert not list(paths.output.glob("*.pptx"))
    assert not (paths.output / "OCL_Databook.xlsx").exists()
    assert sorted(path.name for path in paths.output.iterdir() if path.is_file()) == ["OCL_Databook_v1.xlsx", "OCL_Databook_v2.xlsx"]
    assert (paths.support_working / "PKG-001" / "OCL_Stage2_Review.xlsx").is_file()
    assert second.checkpoint.parent == paths.work / "ocl_runtime" / "PKG-001"
    assert "part1" not in second.timings and "deterministic_analysis" not in second.timings
    assert {"data_prep_package_loading", "partner_interpretation_rendering", "final_presentation", "final_qa", "version_publication"} <= set(second.timings)

    third = end_to_end.run_end_to_end(paths, data_prep_output=package)
    assert third.state == "READY"
    assert third.published_version == 2
    assert not (paths.output / "OCL_Databook_v3.xlsx").exists()
    assert counters == {"part1": 1, "analysis": 1}


def test_ready_checkpoint_survives_real_finalization_lifecycle(tmp_path: Path, monkeypatch, capsys):
    paths, package, _runtime_config = _case(tmp_path)
    counters = {
        "part1": 0,
        "analysis": 0,
        "interpretation": 0,
        "presentation": 0,
        "qa": 0,
        "publication": 0,
    }
    _install_stage_spies(monkeypatch, counters)

    source_fingerprint = end_to_end.source_package_fingerprint(paths.source)

    def cached_data_prep(*_args, **_kwargs):
        return SimpleNamespace(
            source_fingerprint=source_fingerprint,
            warnings=(),
            ready=True,
            standardized_output=package,
        )

    def render_interpretation(path, _analysis_result, _interpretation):
        counters["interpretation"] += 1
        workbook = load_workbook(path)
        workbook["Flat File"]["F1"] = "partner interpretation rendered"
        workbook.save(path)
        return []

    def final_presentation(path, _handoff):
        counters["presentation"] += 1
        workbook = load_workbook(path)
        workbook["Mapping"]["A1"] = "final presentation applied"
        workbook.save(path)
        return path

    def final_qa(path, qa_path):
        counters["qa"] += 1
        workbook = load_workbook(path, read_only=True)
        assert workbook["Mapping"]["A1"].value == "final presentation applied"
        workbook.close()
        qa_path.write_text(json.dumps({"status": "PASS"}), encoding="utf-8")
        return {"status": "PASS"}

    real_publish = end_to_end.publish_versioned_databook

    def publish(path, output, **kwargs):
        counters["publication"] += 1
        return real_publish(path, output, **kwargs)

    monkeypatch.setattr(end_to_end, "run_full_data_preparation", cached_data_prep)
    monkeypatch.setattr(end_to_end, "apply_partner_interpretation", render_interpretation)
    monkeypatch.setattr(end_to_end, "apply_final_workbook_presentation", final_presentation)
    monkeypatch.setattr(end_to_end, "validate_final_databook", final_qa)
    monkeypatch.setattr(end_to_end, "publish_versioned_databook", publish)

    awaiting = end_to_end.run_end_to_end(paths)
    evidence = paths.work / "analysis" / "PKG-001" / "analysis_evidence.json"
    assert awaiting.state == "AWAITING_ANALYSIS_INTERPRETATION"
    _write_interpretation(evidence)

    completed = end_to_end.run_end_to_end(paths)
    assert completed.state == "READY"
    assert completed.published_version == 1
    checkpoint = json.loads(completed.checkpoint.read_text(encoding="utf-8"))
    assert checkpoint["qa_databook_hash"] == checkpoint["published_databook_hash"]
    assert checkpoint["qa_databook_hash"] == checkpoint["working_databook_hash"]
    assert counters == {
        "part1": 1,
        "analysis": 1,
        "interpretation": 1,
        "presentation": 1,
        "qa": 1,
        "publication": 1,
    }

    capsys.readouterr()
    resumed = end_to_end.run_end_to_end(paths)
    output = capsys.readouterr().out

    assert resumed.state == "READY"
    assert resumed.databook == completed.databook
    assert resumed.published_version == completed.published_version
    assert "Checkpoint resume: ACCEPTED — READY" in output
    assert counters == {
        "part1": 1,
        "analysis": 1,
        "interpretation": 1,
        "presentation": 1,
        "qa": 1,
        "publication": 1,
    }
    assert not (paths.output / "OCL_Databook_v2.xlsx").exists()

    # An explicit reference to the exact same package has no published raw-
    # source fingerprint, but remains safely bound by its full package hash.
    explicit_resume = end_to_end.run_end_to_end(paths, data_prep_output=package)
    assert explicit_resume.state == "READY"
    assert explicit_resume.databook == completed.databook
    assert counters["publication"] == 1


def test_known_source_fingerprint_change_still_rejects_checkpoint():
    checkpoint = {
        "package_id": "PKG-001",
        "source_fingerprint": "source-a",
        "package_fingerprint": "package",
        "semantic_handoff_hash": "semantic",
        "judgment_config_hash": "judgments",
    }
    expected = {**checkpoint, "source_fingerprint": "source-b"}

    resumable, reason = checkpoint_matches(checkpoint, expected)

    assert resumable is False
    assert reason == "source_fingerprint changed"


@pytest.mark.parametrize("changed_input", ["semantic", "judgment", "package"])
def test_checkpoint_invalidates_when_bound_input_changes(tmp_path: Path, monkeypatch, capsys, changed_input: str):
    paths, package, runtime_config = _case(tmp_path)
    counters = {"part1": 0, "analysis": 0}
    _install_stage_spies(monkeypatch, counters)
    first = end_to_end.run_end_to_end(paths, data_prep_output=package)
    assert first.state == "AWAITING_ANALYSIS_INTERPRETATION"

    if changed_input == "semantic":
        semantic = runtime_config / "semantic_handoff.json"
        payload = json.loads(semantic.read_text(encoding="utf-8"))
        payload["unresolved_matters"] = ["changed"]
        semantic.write_text(json.dumps(payload), encoding="utf-8")
    elif changed_input == "judgment":
        (runtime_config / "mapping.csv").write_text("header\nchanged\n", encoding="utf-8")
    else:
        (package / "standardized.csv").write_text("Category,Amount\nBonus,2\n", encoding="utf-8")

    capsys.readouterr()
    second = end_to_end.run_end_to_end(paths, data_prep_output=package)
    output = capsys.readouterr().out
    assert second.state == "AWAITING_ANALYSIS_INTERPRETATION"
    assert counters == {"part1": 2, "analysis": 2}
    expected_reason = {
        "semantic": "semantic_handoff_hash changed",
        "judgment": "judgment_config_hash changed",
        "package": "package_fingerprint changed",
    }[changed_input]
    assert f"Checkpoint resume: REJECTED — {expected_reason}" in output


def test_stale_interpretation_is_rejected_without_rebuilding_deterministic_stages(tmp_path: Path, monkeypatch):
    paths, package, _runtime_config = _case(tmp_path)
    counters = {"part1": 0, "analysis": 0}
    _install_stage_spies(monkeypatch, counters)
    first = end_to_end.run_end_to_end(paths, data_prep_output=package)
    evidence = paths.work / "analysis" / "PKG-001" / "analysis_evidence.json"
    _write_interpretation(evidence, evidence_hash="STALE")

    second = end_to_end.run_end_to_end(paths, data_prep_output=package)

    assert second.state == "AWAITING_ANALYSIS_INTERPRETATION"
    assert counters == {"part1": 1, "analysis": 1}
    assert "evidence_hash does not match" in second.coordination["validation_error"]


def test_final_presentation_qa_and_publication_order(tmp_path: Path, monkeypatch):
    paths, package, _runtime_config = _case(tmp_path)
    counters = {"part1": 0, "analysis": 0}
    _install_stage_spies(monkeypatch, counters)
    first = end_to_end.run_end_to_end(paths, data_prep_output=package)
    evidence = paths.work / "analysis" / "PKG-001" / "analysis_evidence.json"
    _write_interpretation(evidence)
    events = []

    def presentation(path, _handoff):
        events.append("presentation")
        return path

    def qa(_path, qa_path):
        events.append("qa")
        qa_path.write_text(json.dumps({"status": "PASS"}), encoding="utf-8")
        return {"status": "PASS"}

    def publish(working, output, *, expected_sha256=None):
        events.append("publication")
        target = output / "OCL_Databook_v1.xlsx"
        output.mkdir(parents=True, exist_ok=True)
        shutil.copy2(working, target)
        assert expected_sha256 == end_to_end.sha256_file(target)
        return PublishedDatabook(1, target)

    monkeypatch.setattr(end_to_end, "apply_final_workbook_presentation", presentation)
    monkeypatch.setattr(end_to_end, "validate_final_databook", qa)
    monkeypatch.setattr(end_to_end, "publish_versioned_databook", publish)
    result = end_to_end.run_end_to_end(paths, data_prep_output=package)

    assert result.state == "READY"
    assert events == ["presentation", "qa", "publication"]


def test_active_orchestration_has_no_report_dependency():
    source = inspect.getsource(end_to_end)
    assert "part4_report" not in source
    assert "run_report" not in source
    launcher = (Path(__file__).resolve().parents[1] / "run_all.py").read_text(encoding="utf-8")
    assert "--skip-report" not in launcher
    assert "OCL_Report" not in launcher
