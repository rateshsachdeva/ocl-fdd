import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from ocl_agent.part1_databook.run import run_part1


def _write_csv(path: Path, headers, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _build_case(root: Path, *, unknown_type: bool = False):
    package = root / "package"
    config = root / "config"
    output = root / "output"
    package.mkdir(parents=True)
    config.mkdir(parents=True)
    annual = [
        ["A1", "FY24", "Bonus accrual", "2100", "Entity A", 300],
        ["A2", "FY25", "Bonus accrual", "2100", "Entity A", 450],
    ]
    _write_csv(package / "ocl_annual.csv", ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Amount"], annual)
    movement_type = "Mystery" if unknown_type else "Utilisation"
    movements = [
        ["M1", "FY25", "Bonus accrual", "2100", "Entity A", "Opening", 300],
        ["M2", "FY25", "Bonus accrual", "2100", "Entity A", "Additions", 200],
        ["M3", "FY25", "Bonus accrual", "2100", "Entity A", movement_type, 50],
        ["M4", "FY25", "Bonus accrual", "2100", "Entity A", "Closing", 450],
    ]
    _write_csv(package / "movement.csv", ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Movement_Type", "Amount"], movements)
    outputs = ["ocl_annual.csv", "movement.csv"]
    (package / "execution_manifest.json").write_text(json.dumps({"execution_id": "MOV-001", "final_execution_status": "COMPLETED", "outputs_created": outputs}), encoding="utf-8")
    (package / "databook_metadata.json").write_text(json.dumps({"workflow_run_id": "MOV-001"}), encoding="utf-8")
    _write_csv(config / "judgment_scope.csv", ["source_label", "source_code", "entity", "scope", "review_status", "reason"], [["Bonus accrual", "2100", "Entity A", "IN_SCOPE", "REVIEWED", ""]])
    _write_csv(config / "mapping.csv", ["source_label", "source_code", "entity", "category", "parent_category", "review_status", "reason"], [["Bonus accrual", "2100", "Entity A", "Bonus accrual", "Employee accruals", "REVIEWED", ""]])
    _write_csv(config / "judgment_wc_debt.csv", ["source_label", "source_code", "entity", "management_view", "fdd_view", "normality", "review_status", "reason"], [["Bonus accrual", "2100", "Entity A", "working_capital", "working_capital", "normal", "REVIEWED", ""]])
    handoff = {
        "handoff_version": "1.0",
        "status": "CONFIRMED",
        "package_id": "MOV-001",
        "datasets": [
            {"file": "ocl_annual.csv", "usages": ["OCL_RECORDS"], "fields": {"source_record_id": "Source_Record_ID", "period": "Period", "amount": "Amount", "source_label": "Source_Label", "source_code": "Source_Code", "entity": "Entity"}, "dimensions": []},
            {"file": "movement.csv", "usages": ["MOVEMENT_RECORDS"], "fields": {"source_record_id": "Source_Record_ID", "period": "Period", "amount": "Amount", "source_label": "Source_Label", "source_code": "Source_Code", "entity": "Entity", "movement_type": "Movement_Type"}, "dimensions": [], "movement_rules": {
                "Opening": {"role": "OPENING", "multiplier": 1},
                "Additions": {"role": "FLOW", "multiplier": 1},
                "Utilisation": {"role": "FLOW", "multiplier": -1},
                "Closing": {"role": "CLOSING", "multiplier": 1}
            }}
        ],
        "movement_to_annual": [{"movement_period": "FY25", "annual_period": "FY25"}],
        "monthly_to_annual": [],
        "controls": [],
        "unresolved_matters": []
    }
    (config / "semantic_handoff.json").write_text(json.dumps(handoff), encoding="utf-8")
    return package, config, output


def test_explicit_movement_rules_produce_rollforward(tmp_path: Path):
    package, config, output = _build_case(tmp_path)
    result = run_part1(package, config, output)
    assert result.state == "DATABOOK_READY"
    rollforward = next(control for control in result.controls if control.control_id == "chk_rollforward")
    assert rollforward.status.value == "PASS"
    workbook = load_workbook(result.databook, data_only=False, read_only=True)
    assert "Roll-forward" in workbook.sheetnames
    sheet = workbook["Roll-forward"]
    assert sheet["F2"].value == "=C2+D2"
    assert sheet["G2"].value == "=F2-E2"
    workbook.close()


def test_unknown_movement_type_stops_instead_of_guessing(tmp_path: Path):
    package, config, output = _build_case(tmp_path, unknown_type=True)
    result = run_part1(package, config, output)
    assert result.state == "AWAITING_CONTROL_ALIGNMENT"
    control = next(item for item in result.controls if item.control_id == "chk_rollforward")
    assert control.status.value == "REVIEW_REQUIRED"
    assert result.databook is None
