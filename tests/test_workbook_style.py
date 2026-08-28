from __future__ import annotations

from collections import Counter, defaultdict

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ocl_agent import workbook_style
from ocl_agent.databook_display import apply_display_preferences_to_workbook
from ocl_agent.workbook_style import (
    _analysis_sheet,
    _base_sheet,
    _flat_sheet,
    _reasonable_widths,
    style_generated_support_cell,
)


def _flat_sheet_with_rows(row_count: int) -> Worksheet:
    sheet = Workbook().active
    sheet.title = "Flat File"
    sheet.cell(1, 1, "Flat File")
    headers = ["Source_Record_ID", "Source_Label", "Category", "Period", "Amount"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(2, col, header)
    for row in range(3, row_count + 3):
        sheet.cell(row, 1, f"SRC-{row}")
        sheet.cell(row, 2, "Dataset A")
        sheet.cell(row, 3, "Bonus")
        sheet.cell(row, 4, "FY2025")
        sheet.cell(row, 5, row)
    return sheet


def _analysis_sheet_with_rows(row_count: int) -> Worksheet:
    sheet = Workbook().active
    sheet.title = "Key Findings"
    sheet.cell(1, 1, "Key Findings")
    sheet.cell(6, 2, "Key Findings")
    headers = ["Area", "Metric", "FY periods / Item", "So what", "Evidence", "Ask management"]
    for col, header in enumerate(headers, start=2):
        sheet.cell(7, col, header)
    for row in range(8, row_count + 8):
        for col, value in enumerate(("Area", "Metric", "FY2025", "Implication", "Evidence", ""), start=2):
            sheet.cell(row, col, value)
    return sheet


@pytest.mark.parametrize(
    ("styler", "builder"),
    [
        (_flat_sheet, _flat_sheet_with_rows),
        (_analysis_sheet, _analysis_sheet_with_rows),
    ],
)
def test_dimension_property_reads_do_not_scale_with_data_rows(monkeypatch, styler, builder) -> None:
    small_sheet = builder(2)
    large_sheet = builder(250)
    reads: defaultdict[int, Counter[str]] = defaultdict(Counter)
    original_max_row = Worksheet.max_row
    original_max_column = Worksheet.max_column

    def counted_max_row(sheet: Worksheet) -> int:
        reads[id(sheet)]["max_row"] += 1
        return original_max_row.fget(sheet)

    def counted_max_column(sheet: Worksheet) -> int:
        reads[id(sheet)]["max_column"] += 1
        return original_max_column.fget(sheet)

    monkeypatch.setattr(Worksheet, "max_row", property(counted_max_row))
    monkeypatch.setattr(Worksheet, "max_column", property(counted_max_column))

    styler(small_sheet)
    styler(large_sheet)

    assert reads[id(large_sheet)] == reads[id(small_sheet)]
    assert 0 < reads[id(large_sheet)]["max_row"] <= 3
    assert 0 < reads[id(large_sheet)]["max_column"] <= 4


def test_base_sheet_does_not_traverse_cells(monkeypatch) -> None:
    sheet = _flat_sheet_with_rows(500)

    def fail(*_args, **_kwargs):
        raise AssertionError("_base_sheet must not traverse worksheet cells")

    monkeypatch.setattr(sheet, "iter_rows", fail)
    _base_sheet(sheet)

    assert sheet.sheet_view.showGridLines is False
    assert sheet.print_options.gridLines is False


def test_preformatted_large_flat_sheet_skips_compatibility_repaint(monkeypatch) -> None:
    sheet = _flat_sheet_with_rows(250)
    for row in range(3, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            style_generated_support_cell(sheet.cell(row, col), role="linked", accounting=col == 5)

    def fail(*_args, **_kwargs):
        raise AssertionError("preformatted generated support body must not be repainted")

    monkeypatch.setattr(workbook_style, "style_generated_support_cell", fail)
    _flat_sheet(sheet)


def test_large_sheet_width_sampling_is_bounded(monkeypatch) -> None:
    sheet = Workbook().active
    sheet.cell(1, 1, "Header")
    sheet.cell(10000, 1, "Tail value")
    original_cell = sheet.cell
    rows_read: list[int] = []

    def counted_cell(row: int, column: int, *args, **kwargs):
        rows_read.append(row)
        return original_cell(row, column, *args, **kwargs)

    monkeypatch.setattr(sheet, "cell", counted_cell)
    _reasonable_widths(sheet)

    assert rows_read
    assert max(rows_read) <= workbook_style.WIDTH_SAMPLE_ROWS
    assert len(rows_read) == workbook_style.WIDTH_SAMPLE_ROWS


def test_display_preferences_do_not_add_second_full_sheet_traversal(monkeypatch) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movements"
    sheet.cell(10000, 9, "=G10000*H10000")

    def fail(*_args, **_kwargs):
        raise AssertionError("display preferences must not traverse every worksheet cell")

    monkeypatch.setattr(Worksheet, "iter_rows", fail)
    apply_display_preferences_to_workbook(workbook, {})


def test_generated_support_cells_reuse_common_style_components() -> None:
    sheet = Workbook().active
    first = sheet.cell(1, 1, "A")
    second = sheet.cell(2, 1, "B")

    style_generated_support_cell(first, role="model")
    style_generated_support_cell(second, role="model")

    assert first._style == second._style
