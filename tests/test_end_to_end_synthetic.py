import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from pptx import Presentation

from ocl_agent.part1_databook.run import run_part1
from ocl_agent.part2_analysis.run import run_analysis
from ocl_agent.part3_qanda.run import run_qanda
from ocl_agent.part4_report.run import run_report


def _write_csv(path: Path, headers, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _source_id(sheet: str, row: int) -> str:
    return json.dumps({"source_id": "Synthetic_OCL_Source.xlsx", "worksheet_name": sheet, "row": row}, separators=(",", ":"))


def _build_package(root: Path) -> Path:
    root.mkdir(parents=True)
    annual_rows = [
        (_source_id("Annual", 2), "FY24", "Bonus accrual", "2100", "Entity A", 300),
        (_source_id("Annual", 3), "FY24", "Holiday pay", "2110", "Entity A", 200),
        (_source_id("Annual", 4), "FY24", "VAT payable", "2120", "Entity A", 100),
        (_source_id("Annual", 5), "FY24", "Trade payables", "2000", "Entity A", 400),
        (_source_id("Annual", 6), "FY25", "Bonus accrual", "2100", "Entity A", 450),
        (_source_id("Annual", 7), "FY25", "Holiday pay", "2110", "Entity A", 250),
        (_source_id("Annual", 8), "FY25", "VAT payable", "2120", "Entity A", 150),
        (_source_id("Annual", 9), "FY25", "Trade payables", "2000", "Entity A", 500),
    ]
    _write_csv(root / "ocl_annual.csv", ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Amount"], annual_rows)

    monthly_rows = []
    monthly_values = {
        "Bonus accrual": [300, 250, 275, 330, 290, 310, 360, 340, 375, 390, 410, 450],
        "Holiday pay": [200, 180, 190, 205, 210, 220, 215, 225, 235, 240, 245, 250],
        "VAT payable": [100, 90, 105, 95, 110, 125, 120, 130, 140, 135, 145, 150],
    }
    codes = {"Bonus accrual": "2100", "Holiday pay": "2110", "VAT payable": "2120"}
    # Include the prior year-end month so both annual closing alignments can be tested.
    for label, amount in (("Bonus accrual", 300), ("Holiday pay", 200), ("VAT payable", 100)):
        monthly_rows.append((_source_id("Monthly", len(monthly_rows) + 2), "2024-12", label, codes[label], "Entity A", amount))
    for month in range(1, 13):
        period = f"2025-{month:02d}"
        for label, values in monthly_values.items():
            monthly_rows.append((_source_id("Monthly", len(monthly_rows) + 2), period, label, codes[label], "Entity A", values[month - 1]))
    _write_csv(root / "ocl_monthly.csv", ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Amount"], monthly_rows)

    _write_csv(root / "tb_control.csv", ["Period", "Control", "Amount"], [["FY24", "OCL", 600], ["FY25", "OCL", 850]])
    _write_csv(root / "revenue_context.csv", ["Period", "Revenue"], [["FY24", 5000], ["FY25", 5600]])
    _write_csv(root / "payroll_context.csv", ["Period", "Payroll"], [["FY24", 1200], ["FY25", 1400]])
    outputs = ["ocl_annual.csv", "ocl_monthly.csv", "tb_control.csv", "revenue_context.csv", "payroll_context.csv"]
    (root / "execution_manifest.json").write_text(json.dumps({"execution_id": "SYNTH-001", "final_execution_status": "COMPLETED", "outputs_created": outputs}), encoding="utf-8")
    (root / "databook_metadata.json").write_text(json.dumps({"workflow_run_id": "SYNTH-001", "description": "Synthetic package based on OCL databook control rules"}), encoding="utf-8")
    return root


def _build_config(root: Path) -> Path:
    root.mkdir(parents=True)
    _write_csv(root / "judgment_scope.csv", ["source_label", "source_code", "entity", "scope", "review_status", "reason"], [
        ["Bonus accrual", "2100", "Entity A", "IN_SCOPE", "REVIEWED", "Operating accrual"],
        ["Holiday pay", "2110", "Entity A", "IN_SCOPE", "REVIEWED", "Employee accrual"],
        ["VAT payable", "2120", "Entity A", "IN_SCOPE", "REVIEWED", "Operating tax payable"],
        ["Trade payables", "2000", "Entity A", "TRADE_PAYABLE", "REVIEWED", "Separate AP population"],
    ])
    _write_csv(root / "mapping.csv", ["source_label", "source_code", "entity", "category", "parent_category", "review_status", "reason"], [
        ["Bonus accrual", "2100", "Entity A", "Bonus accrual", "Employee accruals", "REVIEWED", ""],
        ["Holiday pay", "2110", "Entity A", "Holiday pay", "Employee accruals", "REVIEWED", ""],
        ["VAT payable", "2120", "Entity A", "VAT payable", "Taxes and statutory", "REVIEWED", ""],
        ["Trade payables", "2000", "Entity A", "", "", "REVIEWED", ""],
    ])
    _write_csv(root / "judgment_wc_debt.csv", ["source_label", "source_code", "entity", "management_view", "fdd_view", "normality", "review_status", "reason"], [
        ["Bonus accrual", "2100", "Entity A", "Working capital", "Working capital", "Normal", "REVIEWED", ""],
        ["Holiday pay", "2110", "Entity A", "Working capital", "Working capital", "Normal", "REVIEWED", ""],
        ["VAT payable", "2120", "Entity A", "Working capital", "Debt-like", "Normal", "REVIEWED", "Statutory balance assessed as debt-like for synthetic test"],
        ["Trade payables", "2000", "Entity A", "", "", "", "REVIEWED", ""],
    ])
    handoff = {
        "handoff_version": "1.0",
        "status": "CONFIRMED",
        "package_id": "SYNTH-001",
        "datasets": [
            {"file": "ocl_annual.csv", "usages": ["OCL_RECORDS"], "fields": {"source_record_id": "Source_Record_ID", "period": "Period", "amount": "Amount", "source_label": "Source_Label", "source_code": "Source_Code", "entity": "Entity"}, "dimensions": [], "notes": "Synthetic annual OCL candidate listing"},
            {"file": "ocl_monthly.csv", "usages": ["MONTHLY_RECORDS"], "fields": {"source_record_id": "Source_Record_ID", "period": "Period", "amount": "Amount", "source_label": "Source_Label", "source_code": "Source_Code", "entity": "Entity"}, "dimensions": [], "notes": "Synthetic monthly OCL listing"},
            {"file": "tb_control.csv", "usages": ["TB_CONTROL"], "fields": {}, "dimensions": [], "notes": "Exact OCL TB control"},
            {"file": "revenue_context.csv", "usages": ["REVENUE_CONTEXT"], "fields": {}, "dimensions": [], "notes": "Optional context"},
            {"file": "payroll_context.csv", "usages": ["PAYROLL_CONTEXT"], "fields": {}, "dimensions": [], "notes": "Optional context"},
        ],
        "unresolved_matters": [],
        "monthly_to_annual": [
            {"annual_period": "FY24", "monthly_period": "2024-12"},
            {"annual_period": "FY25", "monthly_period": "2025-12"},
        ],
        "controls": [
            {"control_id": "chk_listing_vs_tb", "dataset_file": "tb_control.csv", "period_field": "Period", "amount_field": "Amount", "filters": {"Control": ["OCL"]}}
        ],
    }
    (root / "semantic_handoff.json").write_text(json.dumps(handoff, indent=2), encoding="utf-8")
    return root


def test_complete_synthetic_workflow(tmp_path: Path):
    package = _build_package(tmp_path / "package")
    config = _build_config(tmp_path / "config")
    output = tmp_path / "output"

    part1 = run_part1(package, config, output)
    assert part1.state == "DATABOOK_READY"
    assert part1.databook is not None
    assert all(control.status.value in {"PASS", "NOT_APPLICABLE"} for control in part1.controls)

    analysis = run_analysis(part1.build.records, part1.databook)
    assert any(finding.finding_type == "TOTAL_CHANGE" for finding in analysis.findings)
    assert any(finding.finding_type == "CONCENTRATION" for finding in analysis.findings)
    assert any(finding.finding_type == "DEBT_LIKE" for finding in analysis.findings)

    questions = run_qanda(analysis, part1.databook)
    assert questions
    report = run_report(analysis, questions, output)
    assert report.exists()

    workbook = load_workbook(part1.databook, read_only=True, data_only=False)
    assert "Analysis Summary" in workbook.sheetnames
    assert "Key Findings" in workbook.sheetnames
    assert "Management Questions" in workbook.sheetnames
    assert "Monthly Balance" in workbook.sheetnames
    assert "SRC_ocl_annual" in workbook.sheetnames
    workbook.close()

    presentation = Presentation(report)
    assert len(presentation.slides) >= 4


def test_annual_only_degrades_gracefully(tmp_path: Path):
    package = _build_package(tmp_path / "package")
    (package / "ocl_monthly.csv").unlink()
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].remove("ocl_monthly.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_config(tmp_path / "config")
    handoff = json.loads((config / "semantic_handoff.json").read_text(encoding="utf-8"))
    handoff["datasets"] = [item for item in handoff["datasets"] if item["file"] != "ocl_monthly.csv"]
    handoff["monthly_to_annual"] = []
    (config / "semantic_handoff.json").write_text(json.dumps(handoff), encoding="utf-8")

    output = tmp_path / "output"
    part1 = run_part1(package, config, output)
    assert part1.state == "DATABOOK_READY"
    analysis = run_analysis(part1.build.records, part1.databook)
    questions = run_qanda(analysis, part1.databook)
    report = run_report(analysis, questions, output)
    workbook = load_workbook(part1.databook, read_only=True)
    assert "Monthly Balance" not in workbook.sheetnames
    workbook.close()
    assert report.exists()
