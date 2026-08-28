import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from ocl_agent.auto_semantics import ensure_semantic_handoff
from ocl_agent.part1_databook.run import run_part1
from ocl_agent.part2_analysis.extended import extended_analysis


def _write_csv(path: Path, headers, rows) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _build_canonical_package(root: Path) -> Path:
    root.mkdir(parents=True)
    headers = ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Amount"]
    rows = [
        ["SRC-1", "FY24", "Bonus accrual", "2100", "Entity A", "300000"],
        ["SRC-2", "FY25", "Bonus accrual", "2100", "Entity A", "450000"],
    ]
    _write_csv(root / "ocl_annual.csv", headers, rows)
    _write_csv(root / "revenue_context.csv", ["Period", "Revenue"], [["FY24", "5000000"], ["FY25", "5600000"]])
    outputs = ["ocl_annual.csv", "revenue_context.csv"]
    (root / "execution_manifest.json").write_text(
        json.dumps({"execution_id": "AUTO-SEM-001", "final_execution_status": "COMPLETED", "outputs_created": outputs}),
        encoding="utf-8",
    )
    (root / "databook_metadata.json").write_text(
        json.dumps({"workflow_run_id": "AUTO-SEM-001"}),
        encoding="utf-8",
    )
    return root


def _build_reviewed_config(root: Path) -> Path:
    root.mkdir(parents=True)
    _write_csv(
        root / "judgment_scope.csv",
        ["source_label", "source_code", "entity", "scope", "review_status", "reason"],
        [["Bonus accrual", "2100", "Entity A", "IN_SCOPE", "REVIEWED", "Operating accrual"]],
    )
    _write_csv(
        root / "mapping.csv",
        ["source_label", "source_code", "entity", "category", "parent_category", "review_status", "reason"],
        [["Bonus accrual", "2100", "Entity A", "Bonus accrual", "Employee accruals", "REVIEWED", ""]],
    )
    _write_csv(
        root / "judgment_wc_debt.csv",
        ["source_label", "source_code", "entity", "management_view", "fdd_view", "normality", "review_status", "reason"],
        [["Bonus accrual", "2100", "Entity A", "working_capital", "working_capital", "normal", "REVIEWED", ""]],
    )
    return root


def test_canonical_publication_skips_second_ai_semantic_checkpoint(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    config = _build_reviewed_config(tmp_path / "config")
    output = tmp_path / "output"

    result = run_part1(package, config, output)

    assert result.state == "DATABOOK_READY"
    assert result.handoff is not None
    payload = json.loads((config / "semantic_handoff.json").read_text(encoding="utf-8"))
    assert payload["status"] == "CONFIRMED"
    assert payload["confirmed_by"] == "INTEGRATED_CANONICAL_CONTRACT"
    annual = next(item for item in payload["datasets"] if item["file"] == "ocl_annual.csv")
    assert annual["usages"] == ["OCL_RECORDS"]
    assert "currency" not in annual["fields"]
    revenue = next(item for item in payload["datasets"] if item["file"] == "revenue_context.csv")
    assert revenue["fields"] == {"period": "Period", "amount": "Revenue"}


def test_noncanonical_publication_keeps_explicit_semantic_review_fallback(tmp_path: Path):
    package = tmp_path / "package"
    package.mkdir()
    _write_csv(
        package / "custom_output.csv",
        ["Source_Record_ID", "Period", "Source_Label", "Amount"],
        [["SRC-1", "FY25", "Something", "100"]],
    )
    (package / "execution_manifest.json").write_text(
        json.dumps({"execution_id": "CUSTOM-001", "final_execution_status": "COMPLETED", "outputs_created": ["custom_output.csv"]}),
        encoding="utf-8",
    )
    config = tmp_path / "config"

    assert ensure_semantic_handoff(package, config) is None
    assert not (config / "semantic_handoff.json").exists()


def test_canonical_extra_output_is_preserved_as_supporting_evidence(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(package / "settlement_support.csv", ["Source_Record_ID", "Period", "Amount", "Evidence_Type"], [["E1", "FY25", 100, "Settlement"]])
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("settlement_support.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = tmp_path / "config"

    handoff_path = ensure_semantic_handoff(package, config)
    payload = json.loads(handoff_path.read_text(encoding="utf-8"))
    support = next(item for item in payload["datasets"] if item["file"] == "settlement_support.csv")

    assert support["usages"] == ["SUPPORTING_EVIDENCE"]
    assert support["fields"] == {"source_record_id": "Source_Record_ID", "period": "Period", "amount": "Amount"}
    assert support["dimensions"] == ["Evidence_Type"]


def test_canonical_movement_roles_and_row_multipliers_activate_rollforward(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(
        package / "ocl_movements.csv",
        ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Movement_Type", "Movement_Multiplier", "Amount"],
        [
            ["M1", "FY24", "Bonus accrual", "2100", "Entity A", "OPENING", 1, 0],
            ["M2", "FY24", "Bonus accrual", "2100", "Entity A", "FLOW", 1, 300000],
            ["M3", "FY24", "Bonus accrual", "2100", "Entity A", "CLOSING", 1, 300000],
            ["M4", "FY25", "Bonus accrual", "2100", "Entity A", "OPENING", 1, 300000],
            ["M5", "FY25", "Bonus accrual", "2100", "Entity A", "FLOW", 1, 200000],
            ["M6", "FY25", "Bonus accrual", "2100", "Entity A", "FLOW", -1, 50000],
            ["M7", "FY25", "Bonus accrual", "2100", "Entity A", "CLOSING", 1, 450000],
        ],
    )
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("ocl_movements.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "DATABOOK_READY"
    movement = next(item for item in result.handoff.datasets if item.file == "ocl_movements.csv")
    assert movement.fields.movement_multiplier == "Movement_Multiplier"
    assert {record.movement_role for record in result.movement_build.records} == {"OPENING", "FLOW", "CLOSING"}
    assert {record.multiplier for record in result.movement_build.records} == {1, -1}
    assert next(control for control in result.controls if control.control_id == "chk_rollforward").status.value == "PASS"


def test_canonical_partial_movements_reach_records_analysis_and_rollforward_without_claiming_completeness(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(
        package / "ocl_movements.csv",
        ["Source_Record_ID", "Period", "Source_Label", "Source_Code", "Entity", "Movement_Type", "Movement_Multiplier", "Amount"],
        [
            ["M1", "FY24", "Bonus accrual", "2100", "Entity A", "OPENING", 1, 0],
            ["M2", "FY24", "Bonus accrual", "2100", "Entity A", "FLOW", 1, 300000],
            ["M3", "FY24", "Bonus accrual", "2100", "Entity A", "CLOSING", 1, 300000],
            ["M4", "FY25", "Bonus accrual", "2100", "Entity A", "OPENING", 1, 300000],
            ["M5", "FY25", "Bonus accrual", "2100", "Entity A", "FLOW", 1, 200000],
            ["M6", "FY25", "Bonus accrual", "2100", "Entity A", "FLOW", -1, 50000],
            ["M7", "FY25", "Bonus accrual", "2100", "Entity A", "CLOSING", 1, 450000],
        ],
    )
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("ocl_movements.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    metadata = json.loads((package / "databook_metadata.json").read_text(encoding="utf-8"))
    metadata["logical_datasets"] = [{
        "metadata": [{
            "metadata_type": "DATASET_PURPOSE",
            "status": "EVIDENCED",
            "confidence": "HIGH",
            "value": "Selected liability movements",
            "evidence": "The source states that only selected movement-subledger accounts are included.",
        }]
    }]
    (package / "databook_metadata.json").write_text(json.dumps(metadata), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "DATABOOK_READY"
    payload = json.loads((config / "semantic_handoff.json").read_text(encoding="utf-8"))
    movement_binding = next(item for item in payload["datasets"] if item["file"] == "ocl_movements.csv")
    assert movement_binding["usages"] == ["MOVEMENT_RECORDS", "SUPPORTING_EVIDENCE"]
    assert movement_binding["population_coverage"] == "PARTIAL"
    assert payload["movement_to_annual"] == []
    assert result.movement_build is not None
    assert result.movement_build.population_coverage == "PARTIAL"
    assert len(result.movement_build.records) == 7
    assert {record.movement_role for record in result.movement_build.records} == {"OPENING", "FLOW", "CLOSING"}
    assert all(record.dimensions["population_coverage"] == "PARTIAL" for record in result.movement_build.records)

    control = next(item for item in result.controls if item.control_id == "chk_rollforward")
    assert control.status.value == "PASS"
    assert control.evidence["population_coverage"] == "PARTIAL"
    assert control.evidence["population_completeness_assessed"] is False
    assert "full OCL population completeness" in control.message

    workbook = load_workbook(result.databook, read_only=True, data_only=False)
    assert "Roll-forward" in workbook.sheetnames
    assert "Movements" in workbook.sheetnames
    assert "selected movement records only" in workbook["Roll-forward"]["A3"].value
    workbook.close()

    _findings, tables = extended_analysis(result.build.records, movements=result.movement_build.records)
    movement_table = next(table for table in tables if table.key == "movement_patterns")
    assert "selected population only" in movement_table.title
    coverage = next(table for table in tables if table.key == "analysis_coverage")
    coverage_by_analysis = {row[0]: row for row in coverage.rows}
    assert coverage_by_analysis["Utilisation"][1] == "PARTIAL"
    assert coverage_by_analysis["Reversal patterns"][1] == "PARTIAL"
    assert "does not establish full OCL population completeness" in coverage_by_analysis["Utilisation"][3]


def test_canonical_whole_dataset_tb_control_needs_no_literal_ocl_value(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(package / "tb_control.csv", ["Period", "Amount"], [["FY24", 300000], ["FY25", 450000]])
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("tb_control.csv")
    manifest["control_bindings"] = [{
        "control_id": "chk_listing_vs_tb",
        "dataset_file": "tb_control.csv",
        "period_field": "Period",
        "amount_field": "Amount",
        "whole_dataset": True,
    }]
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "DATABOOK_READY"
    binding = next(control for control in result.handoff.controls if control.control_id == "chk_listing_vs_tb")
    assert binding.whole_dataset is True
    assert binding.filters == {}
    assert next(control for control in result.controls if control.control_id == "chk_listing_vs_tb").status.value == "PASS"


def test_canonical_filtered_tb_control_regenerates_exact_source_backed_filter(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(
        package / "tb_control.csv",
        ["Period", "Population_Code", "Amount"],
        [
            ["FY24", "ACC", 300000],
            ["FY24", "OTHER", 900000],
            ["FY25", "ACC", 450000],
            ["FY25", "OTHER", 1100000],
        ],
    )
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("tb_control.csv")
    manifest["control_bindings"] = [{
        "control_id": "chk_listing_vs_tb",
        "dataset_file": "tb_control.csv",
        "period_field": "Period",
        "amount_field": "Amount",
        "filters": {"Population_Code": ["ACC"]},
    }]
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "DATABOOK_READY"
    binding = next(control for control in result.handoff.controls if control.control_id == "chk_listing_vs_tb")
    assert binding.dataset_file == "tb_control.csv"
    assert binding.filters == {"Population_Code": ("ACC",)}
    assert binding.whole_dataset is False
    tb_dataset = next(item for item in result.handoff.datasets if item.file == "tb_control.csv")
    assert tb_dataset.fields.period == "Period"
    assert tb_dataset.fields.amount == "Amount"
    assert "OCL" not in json.dumps(manifest["control_bindings"])
    assert next(control for control in result.controls if control.control_id == "chk_listing_vs_tb").status.value == "PASS"


def test_ambiguous_valid_tb_dataset_does_not_auto_bind_without_explicit_manifest_semantics(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(
        package / "tb_control.csv",
        ["Period", "Population_Code", "Amount"],
        [["FY24", "ACC", 300000], ["FY24", "OTHER", 900000], ["FY25", "ACC", 450000]],
    )
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("tb_control.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "AWAITING_CONTROL_ALIGNMENT"
    assert not result.handoff.controls
    assert next(control for control in result.controls if control.control_id == "chk_listing_vs_tb").status.value == "REVIEW_REQUIRED"


def test_invalid_canonical_tb_population_remains_unresolved(tmp_path: Path):
    package = _build_canonical_package(tmp_path / "package")
    _write_csv(package / "tb_control.csv", ["Period", "Amount"], [["not-a-business-period", "not-numeric"]])
    manifest = json.loads((package / "execution_manifest.json").read_text(encoding="utf-8"))
    manifest["outputs_created"].append("tb_control.csv")
    (package / "execution_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    config = _build_reviewed_config(tmp_path / "config")

    result = run_part1(package, config, tmp_path / "output")

    assert result.state == "AWAITING_CONTROL_ALIGNMENT"
    assert not result.handoff.controls
    assert next(control for control in result.controls if control.control_id == "chk_listing_vs_tb").status.value == "REVIEW_REQUIRED"
