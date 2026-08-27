from __future__ import annotations
import csv, json
from pathlib import Path
from openpyxl import load_workbook
import pytest

from ocl_agent.part1_databook.input_contract import InputContractError, discover_standardized_package, profile_package
from ocl_agent.part1_databook.judgments import load_judgments
from ocl_agent.part1_databook.record_builder import build_ocl_records
from ocl_agent.part1_databook.review_workbook import write_semantic_review
from ocl_agent.part1_databook.run import run_stage2
from ocl_agent.part1_databook.semantic_handoff import SemanticHandoffError, load_semantic_handoff, write_semantic_handoff_draft


def _package(tmp_path: Path, rows=None):
    root=tmp_path/'latest'; root.mkdir()
    rows=rows or [
        {'Source_Record_ID':json.dumps({'source_id':'SRC_1','worksheet_name':'TB','region_id':'R1','physical_row':6,'generated_dimensions':{'Period':'FY24'}},sort_keys=True,separators=(',',':')),'Account':'Accrued payroll','Period':'FY24','Amount':'100','Entity':'A'},
        {'Source_Record_ID':json.dumps({'source_id':'SRC_1','worksheet_name':'TB','region_id':'R1','physical_row':7,'generated_dimensions':{'Period':'FY24'}},sort_keys=True,separators=(',',':')),'Account':'Broken amount','Period':'FY24','Amount':'abc','Entity':'A'},
    ]
    with (root/'records.csv').open('w',newline='',encoding='utf-8') as handle:
        writer=csv.DictWriter(handle,fieldnames=list(rows[0])); writer.writeheader(); writer.writerows(rows)
    (root/'execution_manifest.json').write_text(json.dumps({'execution_id':'EXEC_1','final_execution_status':'COMPLETED','outputs_created':['records.csv']}),encoding='utf-8')
    (root/'databook_metadata.json').write_text(json.dumps({'workflow_run_id':'RUN_1','logical_datasets':[{'logical_dataset_id':'tb','name':'Trial balance','role':'PRIMARY_DATA','dataset_grain':'Account x period','metadata':[]}]}),encoding='utf-8')
    return root


def _handoff(path: Path, *, usage_filters=None):
    dataset = {'file':'records.csv','usages':['OCL_RECORDS'],'fields':{'source_record_id':'Source_Record_ID','period':'Period','amount':'Amount','source_label':'Account','entity':'Entity'},'dimensions':[],'notes':'confirmed'}
    if usage_filters is not None:
        dataset['usage_filters'] = usage_filters
    path.write_text(json.dumps({'handoff_version':'1.0','status':'CONFIRMED','package_id':'RUN_1','datasets':[dataset],'unresolved_matters':[]}),encoding='utf-8')


def _config(path: Path):
    path.mkdir()
    (path/'judgment_scope.csv').write_text('source_label,scope,review_status,reason\nAccrued payroll,IN_SCOPE,REVIEWED,reviewed\n',encoding='utf-8')
    (path/'mapping.csv').write_text('source_label,category,parent_category,review_status,reason\nAccrued payroll,Payroll accruals,Employee accruals,REVIEWED,\n',encoding='utf-8')
    (path/'judgment_wc_debt.csv').write_text('source_label,management_view,fdd_view,normality,review_status,reason\n',encoding='utf-8')


def test_package_validates_manifest_and_profiles_once(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package,sample_limit=1)
    assert package.metadata is not None
    assert profiles[0].row_count==2 and len(profiles[0].sample_rows)==1
    assert profiles[0].columns==('Source_Record_ID','Account','Period','Amount','Entity')


def test_unpublishable_upstream_is_blocked(tmp_path):
    root=_package(tmp_path)
    (root/'execution_manifest.json').write_text(json.dumps({'final_execution_status':'FAILED_VALIDATION','outputs_created':['records.csv']}),encoding='utf-8')
    with pytest.raises(InputContractError,match='not publishable'):
        discover_standardized_package(root)


def test_draft_does_not_guess_account_meaning(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); draft=tmp_path/'draft.json'
    write_semantic_handoff_draft(package,profiles,draft); payload=json.loads(draft.read_text()); fields=payload['datasets'][0]['fields']
    assert fields['source_record_id']=='Source_Record_ID' and fields['period']=='Period' and fields['amount']=='Amount'
    assert fields['source_label'] is None
    assert payload['status']=='DRAFT'


def test_handoff_is_bound_to_current_package_and_columns(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); handoff=tmp_path/'semantic.json'; _handoff(handoff)
    parsed=load_semantic_handoff(handoff,package,profiles); assert parsed.record_bindings()[0].fields.source_label=='Account'
    payload=json.loads(handoff.read_text()); payload['package_id']='OLD_RUN'; handoff.write_text(json.dumps(payload))
    with pytest.raises(SemanticHandoffError,match='belongs to package'):
        load_semantic_handoff(handoff,package,profiles)


def test_handoff_without_usage_filters_is_backward_compatible(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); handoff=tmp_path/'semantic.json'; _handoff(handoff)
    parsed=load_semantic_handoff(handoff,package,profiles)
    assert parsed.record_bindings()[0].usage_filters == {}


@pytest.mark.parametrize(
    ("usage_filters", "message"),
    [
        ({"OCL_RECORDS": {"Missing_Column": ["Closing Balance"]}}, "missing column"),
        ({"MONTHLY_RECORDS": {"Period": ["FY24"]}}, "not present in dataset usages"),
        ({"OCL_RECORDS": {"Period": []}}, "has no values"),
        ({"UNSUPPORTED": {"Period": ["FY24"]}}, "unsupported usage"),
    ],
)
def test_invalid_usage_filters_are_rejected(tmp_path, usage_filters, message):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); handoff=tmp_path/'semantic.json'; _handoff(handoff,usage_filters=usage_filters)
    with pytest.raises(SemanticHandoffError,match=message):
        load_semantic_handoff(handoff,package,profiles)


def test_record_usage_filter_excludes_nonmatching_rows_without_build_errors(tmp_path):
    rows = [
        {'Source_Record_ID':'1','Account':'Accrued payroll','Period':'FY24','Amount':'100','Entity':'A','Balance_Type':'Closing Balance'},
        {'Source_Record_ID':'2','Account':'Accrued payroll','Period':'FY24','Amount':'abc','Entity':'A','Balance_Type':'Prior Year Balance'},
        {'Source_Record_ID':'3','Account':'Accrued payroll','Period':'FY23','Amount':'90','Entity':'A','Balance_Type':'Prior Year Balance'},
    ]
    package=discover_standardized_package(_package(tmp_path,rows)); profiles=profile_package(package); config=tmp_path/'config'; _config(config)
    _handoff(config/'semantic_handoff.json',usage_filters={'OCL_RECORDS':{'Balance_Type':'Closing Balance'}})
    semantic=load_semantic_handoff(config/'semantic_handoff.json',package,profiles)
    result=build_ocl_records(package,semantic,load_judgments(config))
    assert [row.source.source_record_id for row in result.records] == ['1']
    assert result.issues == ()
    assert result.input_rows_by_dataset == {'records.csv':3}
    assert result.excluded_rows_by_dataset == {'records.csv':2}


def test_record_build_preserves_lineage_and_surfaces_invalid_rows(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); config=tmp_path/'config'; _config(config); _handoff(config/'semantic_handoff.json')
    result=build_ocl_records(package,load_semantic_handoff(config/'semantic_handoff.json',package,profiles),load_judgments(config))
    assert len(result.records)==1 and len(result.issues)==1
    record=result.records[0]
    assert str(record.amount)=='100' and record.source.source_file=='SRC_1' and record.source.source_sheet=='TB'
    assert record.judgment.category=='Payroll accruals' and result.issues[0].issue_type=='INVALID_AMOUNT'
    assert len(result.records)+len(result.issues)==2


def test_semantic_review_has_required_review_sheets(tmp_path):
    package=discover_standardized_package(_package(tmp_path)); profiles=profile_package(package); config=tmp_path/'config'; _config(config); _handoff(config/'semantic_handoff.json')
    semantic=load_semantic_handoff(config/'semantic_handoff.json',package,profiles); build=build_ocl_records(package,semantic,load_judgments(config))
    workbook=load_workbook(write_semantic_review(package,profiles,semantic,build,tmp_path/'review.xlsx'),data_only=False)
    assert workbook.sheetnames==['Input_Datasets','Semantic_Handoff','Economic_Judgment_Review','OCL_Scope_Review','Mapping_Review','WC_Debt_Review','Unresolved_Items','Checks']
    economic = workbook['Economic_Judgment_Review']
    headers = [cell.value for cell in economic[1]]
    assert headers[:7] == ['Source_Label','Source_Code','Recommended_Config_Entity','Represented_Entities','Represented_Datasets','Represented_Record_Usages','Technical_Key_Count']
    assert economic['A2'].value == 'Accrued payroll'
    checks={row[0].value:row[1].value for row in workbook['Checks'].iter_rows(min_row=2)}
    assert checks['record_row_coverage']=='PASS' and checks['semantic_build_issues']=='REVIEW_REQUIRED'


def test_first_run_generates_review_and_draft_without_overwriting_config(tmp_path):
    root=_package(tmp_path); config=tmp_path/'config'; _config(config); original=(config/'mapping.csv').read_text()
    result=run_stage2(root,config,tmp_path/'output')
    assert result.state=='AWAITING_SEMANTIC_HANDOFF'
    assert result.input_review.exists() and result.handoff_draft.exists()
    assert not (config/'semantic_handoff.json').exists() and (config/'mapping.csv').read_text()==original


def test_confirmed_handoff_stops_at_review_when_judgments_incomplete(tmp_path):
    root=_package(tmp_path); config=tmp_path/'config'; _config(config); _handoff(config/'semantic_handoff.json')
    result=run_stage2(root,config,tmp_path/'output')
    assert result.state=='AWAITING_JUDGMENT_REVIEW'
    assert result.semantic_review and result.semantic_review.exists()
    assert result.review_context and result.review_context.exists()
    assert result.build and result.build.input_rows_by_dataset=={'records.csv':2}
