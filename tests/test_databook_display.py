from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font

from ocl_agent.databook_display import apply_databook_display_preferences
from ocl_agent.workbook_style import apply_workbook_style


def _handoff():
    return SimpleNamespace(
        monthly_to_annual=(
            SimpleNamespace(annual_period="FY25", monthly_period="2025-12"),
        )
    )


def _build_balance_sheet(workbook, title: str, period: str):
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = True
    sheet["B7"] = "Category"
    sheet["C7"] = period
    sheet["B8"] = "Employee accruals"
    sheet["C8"] = "=SUM(C9:C10)"
    sheet["C8"].font = Font(name="Arial", size=8, color="008000")
    sheet["B9"] = "Bonus accrual"
    sheet["C9"] = "=100"
    sheet["C9"].font = Font(name="Arial", size=8, color="008000")
    sheet["B10"] = "Holiday pay"
    sheet["C10"] = "=200"
    sheet["C10"].font = Font(name="Arial", size=8, color="008000")
    sheet["B11"] = "Total OCL"
    sheet["C11"] = "=C8+C9"
    sheet["C11"].font = Font(name="Arial", size=8, color="008000")
    return sheet


def _as_date(value):
    return value.date() if isinstance(value, datetime) else value


def test_display_preferences_black_dates_no_gridlines_total_fill_and_redundant_subtotal(tmp_path: Path):
    path = tmp_path / "OCL_Databook.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_balance_sheet(workbook, "Balance by Category", "FY25")
    _build_balance_sheet(workbook, "Monthly Balance", "2025-12")
    workbook.save(path)

    apply_workbook_style(path)
    apply_databook_display_preferences(path, _handoff())

    workbook = load_workbook(path)
    for sheet_name in ("Balance by Category", "Monthly Balance"):
        sheet = workbook[sheet_name]
        assert sheet.sheet_view.showGridLines is False
        assert sheet.print_options.gridLines is False
        assert _as_date(sheet["C7"].value) == date(2025, 12, 31)
        assert sheet["C7"].number_format == "mmmyy"
        assert sheet["C8"].font.color.type == "rgb"
        assert sheet["C8"].font.color.rgb.endswith("000000")
        assert sheet.row_dimensions[9].hidden is True
        assert sheet.row_dimensions[9].outlineLevel == 1
        assert sheet.row_dimensions[10].hidden is True
        assert sheet.row_dimensions[10].outlineLevel == 1
        assert sheet.row_dimensions[8].collapsed is True
        assert sheet["B11"].fill.fgColor.rgb.endswith("E5E5E5")
        assert sheet["C11"].fill.fgColor.rgb.endswith("E5E5E5")
        assert sheet["C11"].font.color.rgb.endswith("000000")
    workbook.close()


def test_nonredundant_parent_subtotals_remain_visible(tmp_path: Path):
    path = tmp_path / "OCL_Databook.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("Balance by Category")
    sheet["B7"] = "Category"
    sheet["C7"] = "FY25"
    sheet["B8"] = "Employee accruals"
    sheet["C8"] = "=SUM(C9:C10)"
    sheet["B9"] = "Bonus accrual"
    sheet["C9"] = "=100"
    sheet["B10"] = "Holiday pay"
    sheet["C10"] = "=200"
    sheet["B11"] = "VAT payable"
    sheet["C11"] = "=150"
    sheet["B12"] = "Total OCL"
    sheet["C12"] = "=C8+C9+C11"
    workbook.save(path)

    apply_workbook_style(path)
    apply_databook_display_preferences(path, _handoff())

    workbook = load_workbook(path)
    sheet = workbook["Balance by Category"]
    assert sheet.row_dimensions[8].hidden is False
    assert sheet.row_dimensions[8].collapsed is True
    assert sheet["B12"].fill.fgColor.rgb.endswith("E5E5E5")
    workbook.close()


def test_monthly_chart_legend_is_named_and_bar_gap_is_40_percent(tmp_path: Path):
    path = tmp_path / "OCL_Databook.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Item Monthly Charts"
    for column in range(2, 14):
        sheet.cell(7, column, column)
        sheet.cell(8, column, column * 10)
        sheet.cell(9, column, column * 8)

    bar = BarChart()
    bar.add_data(Reference(sheet, min_col=2, max_col=13, min_row=8, max_row=8), from_rows=True)
    bar.set_categories(Reference(sheet, min_col=2, max_col=13, min_row=7))
    line = LineChart()
    line.add_data(Reference(sheet, min_col=2, max_col=13, min_row=9, max_row=9), from_rows=True)
    bar += line
    sheet.add_chart(bar, "B12")
    workbook.save(path)

    apply_workbook_style(path)
    apply_databook_display_preferences(path, _handoff())

    workbook = load_workbook(path)
    chart = workbook["Item Monthly Charts"]._charts[0]
    assert chart.gapWidth == 40
    assert chart.legend is not None
    assert chart.legend.position == "b"
    assert chart._charts[0].ser[0].tx.v == "Monthly balance"
    assert chart._charts[1].ser[0].tx.v == "LTM 12M average"
    workbook.close()
