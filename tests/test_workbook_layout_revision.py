from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from ocl_agent.part1_databook.reconciliation import category_sum_control
from ocl_agent.part1_databook.movements import embed_rollforward
from ocl_agent.part1_databook.renderer import render_workbook
from ocl_agent.part1_databook.workbook_blueprint import build_blueprint
from ocl_agent.part2_analysis.extended_render import embed_extended_analysis
from ocl_agent.part2_analysis.run import (
    _analysis_sheet,
    _write_formula_linked_monthly_stats,
    _write_seasonality,
)
from ocl_agent.schemas import AnalysisResult, AnalysisTable, MovementRecord, OCLJudgment, OCLRecord, ReviewStatus, Scope, SourceReference
from ocl_agent.workbook_hierarchy import apply_collapsed_detail_group
from ocl_agent.workbook_style import ACCOUNTING, GRAND_TOTAL, PERCENT, apply_workbook_style


def _record(identifier: str, period: str, amount: int, category: str, parent: str, usage: str) -> OCLRecord:
    judgment = OCLJudgment(
        category,
        Scope.IN_SCOPE,
        category,
        parent,
        "working_capital",
        "working_capital",
        "normal",
        ReviewStatus.REVIEWED,
    )
    return OCLRecord(SourceReference(identifier), period, Decimal(amount), category, judgment, {"record_usage": usage})


def test_balance_sheets_sort_latest_groups_and_preserve_collapsed_detail(tmp_path: Path):
    records = (
        _record("A1", "FY24", 20, "Low detail", "Low parent", "OCL_RECORDS"),
        _record("A2", "FY25", 30, "Low detail", "Low parent", "OCL_RECORDS"),
        _record("B1", "FY24", 90, "High detail", "High parent", "OCL_RECORDS"),
        _record("B2", "FY25", 120, "High detail", "High parent", "OCL_RECORDS"),
        _record("MA", "2025-11", 20, "Low detail", "Low parent", "MONTHLY_RECORDS"),
        _record("MB", "2025-11", 110, "High detail", "High parent", "MONTHLY_RECORDS"),
        _record("MC", "2025-12", 25, "Low detail", "Low parent", "MONTHLY_RECORDS"),
        _record("MD", "2025-12", 130, "High detail", "High parent", "MONTHLY_RECORDS"),
    )
    blueprint = build_blueprint(records)
    path = render_workbook(blueprint, records, [category_sum_control(records)], tmp_path / "book.xlsx")
    workbook = load_workbook(path, data_only=False)
    for name in ("Balance by Category", "Monthly Balance"):
        sheet = workbook[name]
        assert [sheet.cell(row, 2).value for row in range(8, 13)] == [
            "High parent", "High detail", "Low parent", "Low detail", "Total OCL"
        ]
        assert sheet.row_dimensions[9].outlineLevel == 1
        assert sheet.row_dimensions[9].hidden is True
        assert sheet.row_dimensions[8].collapsed is True
        assert sheet.row_dimensions[11].outlineLevel == 1
        assert sheet.row_dimensions[11].hidden is True
        assert sheet.cell(12, 2).value == "Total OCL"
    workbook.close()


def test_rollforward_renders_one_formula_table_per_dynamic_category(tmp_path: Path):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.title = "Movements"
    workbook.create_sheet("Roll-forward")
    workbook.save(path)
    movements = []
    for category in ("Bonus accrual", "Energy accrual"):
        judgment = OCLJudgment(category, Scope.IN_SCOPE, category, review_status=ReviewStatus.REVIEWED)
        for period, opening, flow, closing in (("2025-11", 10, 2, 12), ("2025-12", 12, 3, 15)):
            movements.extend(
                (
                    MovementRecord(SourceReference(f"{category}-{period}-O"), period, Decimal(opening), category, "OPENING", Decimal(1), judgment),
                    MovementRecord(SourceReference(f"{category}-{period}-F"), period, Decimal(flow), category, "FLOW", Decimal(1), judgment),
                    MovementRecord(SourceReference(f"{category}-{period}-C"), period, Decimal(closing), category, "CLOSING", Decimal(1), judgment),
                )
            )
    embed_rollforward(path, tuple(movements))
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Roll-forward"]
    title_rows = [row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value in {"Bonus accrual", "Energy accrual"}]
    assert title_rows == [6, 14]
    for title_row in title_rows:
        assert [sheet.cell(title_row + offset, 2).value for offset in range(2, 6)] == ["Opening", "Net movement", "Closing", "Calculated closing"]
        assert sheet.cell(title_row + 5, 3).value == f"=C{title_row + 2}+C{title_row + 3}"
        assert sheet.cell(title_row + 2, 4).value == f"=C{title_row + 4}"
    workbook.close()


def _monthly_source(workbook):
    sheet = workbook.create_sheet("Monthly Balance")
    sheet["A1"] = "TargetCo - Other Current Liabilities"
    sheet["A2"] = "Monthly Balance"
    sheet["B6"] = "Monthly balance by category"
    sheet["B7"] = "Category"
    column = 3
    for year in range(2022, 2026):
        for month in range(1, 13):
            sheet.cell(7, column, f"{year}-{month:02d}")
            sheet.cell(8, column, f"=SUM({sheet.cell(9, column).coordinate}:{sheet.cell(9, column).coordinate})")
            sheet.cell(9, column, year * 100 + month)
            sheet.cell(10, column, f"={sheet.cell(9, column).coordinate}")
            column += 1
    sheet["B8"] = "Employee accruals"
    sheet["B9"] = "Bonus accrual"
    sheet["B10"] = "Total OCL"
    apply_collapsed_detail_group(sheet, 8, 9, 9)
    return sheet


def test_summary_and_seasonality_use_formula_linked_three_year_hierarchy(tmp_path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    source = _monthly_source(workbook)
    annual = workbook.create_sheet("Balance by Category")
    annual["B7"] = "Category"
    annual["C7"] = "FY25"
    annual["B8"] = "Employee accruals"
    annual["C8"] = "=SUM(C9:C9)"
    annual["B9"] = "Bonus accrual"
    annual["C9"] = 1
    annual["B10"] = "Total OCL"
    annual["C10"] = "=C9"
    apply_collapsed_detail_group(annual, 8, 9, 9)

    summary = _analysis_sheet(workbook, "Analysis Summary", "Formula-linked monthly OCL statistics")
    handoff = SimpleNamespace(
        monthly_to_annual=tuple(
            SimpleNamespace(annual_period=f"FY{str(year)[-2:]}", monthly_period=f"{year}-12")
            for year in range(2022, 2026)
        )
    )
    _write_formula_linked_monthly_stats(summary, workbook, 6, handoff)
    _write_seasonality(workbook)
    path = tmp_path / "book.xlsx"
    workbook.save(path)
    apply_workbook_style(path)

    workbook = load_workbook(path, data_only=False)
    summary = workbook["Analysis Summary"]
    assert [summary.cell(7, column).value for column in (3, 7, 11)] == ["FY23", "FY24", "FY25"]
    assert "FY22" not in [cell.value for cell in summary[7]]
    assert [summary.cell(8, column).value for column in range(3, 7)] == ["Average", "Minimum", "Maximum", "Latest"]
    assert "Std Dev" not in [cell.value for row in summary.iter_rows() for cell in row]
    assert summary["C9"].value.startswith("=AVERAGE('Monthly Balance'!")
    assert summary["F9"].value.startswith("='Monthly Balance'!")
    assert summary["F9"].number_format == ACCOUNTING
    assert summary.row_dimensions[10].outlineLevel == 1
    assert summary.row_dimensions[10].hidden is True

    seasonality = workbook["Seasonality"]
    headers = {seasonality.cell(7, column).value: column for column in range(2, seasonality.max_column + 1)}
    assert headers["Peak Month"] + 1 == headers["Flag"]
    assert seasonality.cell(8, headers["YE vs Avg"]).number_format == PERCENT
    assert seasonality.row_dimensions[9].outlineLevel == 1
    assert seasonality.row_dimensions[9].hidden is True
    assert seasonality.cell(10, 2).fill.fgColor.rgb.endswith(GRAND_TOTAL)
    assert workbook["Balance by Category"].cell(10, 2).fill.fgColor.rgb.endswith(GRAND_TOTAL)
    workbook.close()


def test_internal_extended_analysis_remains_in_result_but_tabs_are_removed(tmp_path: Path):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.title = "Analysis Coverage"
    workbook.create_sheet("Additional Analysis")
    workbook.create_sheet("Checks")
    workbook.save(path)
    result = AnalysisResult(
        (),
        (
            AnalysisTable("analysis_coverage", "Coverage", ("Analysis", "Status"), (("Aging", "UNSUPPORTED"),)),
            AnalysisTable("mix_shift", "Mix", ("Category",), (("Bonus",),)),
        ),
    )
    embed_extended_analysis(path, result)
    workbook = load_workbook(path)
    assert "Analysis Coverage" not in workbook.sheetnames
    assert "Additional Analysis" not in workbook.sheetnames
    assert {table.key for table in result.tables} == {"analysis_coverage", "mix_shift"}
    workbook.close()


def test_tab_order_places_monthly_balance_after_annual_balance(tmp_path: Path):
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.title = "Flat File"
    for name in ("Analysis Summary", "Monthly Balance", "Balance by Category", "Checks", "Q&A", "Key Findings", "Deal Issues", "Roll-forward", "Seasonality", "Item Monthly Charts", "Movements"):
        workbook.create_sheet(name)
    workbook.save(path)
    apply_workbook_style(path)
    workbook = load_workbook(path)
    assert workbook.sheetnames[:10] == [
        "Deal Issues", "Key Findings", "Q&A", "Checks", "Balance by Category",
        "Monthly Balance", "Roll-forward", "Seasonality", "Item Monthly Charts", "Analysis Summary",
    ]
    workbook.close()
