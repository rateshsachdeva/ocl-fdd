from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ocl_agent.part1_databook.reconciliation import category_sum_control
from ocl_agent.part1_databook.renderer import render_workbook
from ocl_agent.part1_databook.workbook_blueprint import build_blueprint
from ocl_agent.schemas import OCLJudgment, OCLRecord, ReviewStatus, Scope, SourceReference


def test_renderer_creates_only_blueprinted_sheets(tmp_path: Path):
    row = OCLRecord(
        SourceReference("SRC:1", "source.xlsx", "TB", "D10"),
        "FY25",
        Decimal("100"),
        "Bonus accrual",
        OCLJudgment("Bonus accrual", Scope.IN_SCOPE, "Bonus", "Employee accruals", review_status=ReviewStatus.REVIEWED),
    )
    blueprint = build_blueprint([row], supported_analyses=[])
    path = render_workbook(blueprint, [row], [category_sum_control([row])], tmp_path / "OCL_Databook.xlsx")
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == [sheet.title for sheet in blueprint.sheets]
    assert "Aging" not in workbook.sheetnames
    assert "Monthly OCL" not in workbook.sheetnames
