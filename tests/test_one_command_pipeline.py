from pathlib import Path

from openpyxl import Workbook, load_workbook

from ocl_agent.config import RepoPaths
from ocl_agent.end_to_end import run_end_to_end


def _build_raw_source(path: Path) -> None:
    wb = Workbook()
    annual = wb.active
    annual.title = "Annual OCL"
    annual.append(["GL Account", "Account description", "FY24", "FY25"])
    annual.append(["2100", "Bonus accrual", 300, 450])
    annual.append(["2110", "Holiday pay accrual", 200, 250])
    annual.append(["2120", "VAT payable", 100, 150])
    annual.append(["2130", "Professional fee accrual", 130, 120])
    annual.append(["2000", "Trade payables", 400, 500])
    annual.append(["2200", "Current bank loan", 500, 400])

    monthly = wb.create_sheet("Monthly OCL")
    months = [f"2025-{month:02d}" for month in range(1, 13)]
    monthly.append(["GL Account", "Account description", *months])
    monthly.append(["2100", "Bonus accrual", 300, 250, 275, 330, 290, 310, 360, 340, 375, 390, 410, 450])
    monthly.append(["2110", "Holiday pay accrual", 200, 180, 190, 205, 210, 220, 215, 225, 235, 240, 245, 250])
    monthly.append(["2120", "VAT payable", 100, 90, 105, 95, 110, 125, 120, 130, 140, 135, 145, 150])
    monthly.append(["2130", "Professional fee accrual", 130, 115, 120, 125, 110, 105, 100, 115, 125, 130, 118, 120])

    movements = wb.create_sheet("Movements FY25")
    movements.append(["GL Account", "Account description", "Movement type", "Amount"])
    for code, label, opening, additions, utilisation, closing in (
        ("2100", "Bonus accrual", 300, 520, 370, 450),
        ("2110", "Holiday pay accrual", 200, 190, 140, 250),
        ("2120", "VAT payable", 100, 280, 230, 150),
        ("2130", "Professional fee accrual", 130, 310, 320, 120),
    ):
        movements.append([code, label, "Opening", opening])
        movements.append([code, label, "Additions", additions])
        movements.append([code, label, "Utilisation", utilisation])
        movements.append([code, label, "Closing", closing])

    tb = wb.create_sheet("TB Control")
    tb.append(["Period", "Control line", "Amount"])
    tb.append(["FY24", "OCL", 730])
    tb.append(["FY25", "OCL", 970])
    tb.append(["FY24", "Current liabilities including AP and financing", 1630])
    tb.append(["FY25", "Current liabilities including AP and financing", 1870])

    context = wb.create_sheet("Context")
    context.append(["Period", "Revenue", "Payroll"])
    context.append(["FY24", 5200, 1250])
    context.append(["FY25", 5700, 1420])
    wb.save(path)


def test_raw_source_runs_to_final_databook(tmp_path: Path):
    repo_root = Path(__file__).resolve().parents[1]
    references = tmp_path / "references"
    source = references / "source"
    other = references / "other"
    config = tmp_path / "config"
    output = tmp_path / "output"
    for folder in (source, other, config, output):
        folder.mkdir(parents=True, exist_ok=True)
    raw = source / "Synthetic_Client_OCL_Source.xlsx"
    _build_raw_source(raw)
    source_bytes = raw.read_bytes()

    paths = RepoPaths(repo_root, repo_root / "assets", config, output, references, source, other)
    result = run_end_to_end(paths, skip_report=False)

    assert result.state == "READY"
    assert result.databook == output / "OCL_Databook.xlsx"
    assert result.databook.exists()
    assert result.report == output / "OCL_Report.pptx"
    assert result.report.exists()
    assert result.qa and result.qa["status"] == "PASS"
    assert raw.read_bytes() == source_bytes

    prepared = tmp_path / "work" / "data_prep" / "latest"
    assert (prepared / "execution_manifest.json").exists()
    assert (prepared / "databook_metadata.json").exists()
    assert (prepared / "lineage.csv").exists()
    assert (prepared / "ocl_annual.csv").exists()
    assert (prepared / "ocl_monthly.csv").exists()
    assert (prepared / "ocl_movements.csv").exists()
    assert (prepared / "tb_control.csv").exists()

    wb = load_workbook(result.databook, data_only=False)
    expected = {
        "Flat File", "Balance by Category", "Monthly Flat", "Monthly Balance", "Checks", "Mapping",
        "UNMAPPED", "SCOPE_EXCLUDED", "Roll-forward", "Analysis Summary", "Key Findings", "Management Questions",
    }
    assert expected.issubset(set(wb.sheetnames))
    assert all(wb[name].protection.sheet for name in wb.sheetnames if name.startswith("SRC_"))
    checks = wb["Checks"]
    statuses = [checks.cell(row, 2).value for row in range(2, checks.max_row + 1)]
    assert "FAIL" not in statuses
    assert "REVIEW_REQUIRED" not in statuses
    mapping = wb["Mapping"]
    rows = list(mapping.iter_rows(min_row=2, values_only=True))
    by_label = {row[0]: row for row in rows}
    assert by_label["Trade payables"][3] == "TRADE_PAYABLE"
    assert by_label["Current bank loan"][3] == "FINANCING"
    wb.close()
