import csv, json
from pathlib import Path
from openpyxl import load_workbook
from ocl_agent.part1_databook.run import run_part1


def _write_config(config: Path):
    config.mkdir()
    (config/'judgment_scope.csv').write_text('source_label,source_code,entity,scope,review_status,reason\nBonus accrual,2100,,IN_SCOPE,REVIEWED,\n')
    (config/'mapping.csv').write_text('source_label,source_code,entity,category,parent_category,review_status,reason\nBonus accrual,2100,,Bonus,Employee accruals,REVIEWED,\n')
    (config/'judgment_wc_debt.csv').write_text('source_label,source_code,entity,management_view,fdd_view,normality,review_status,reason\nBonus accrual,2100,,working_capital,working_capital,normal,REVIEWED,\n')


def test_part1_renders_source_linked_dynamic_databook(tmp_path: Path):
    root = tmp_path/'latest'; root.mkdir()
    rid = json.dumps({'source_id':'SRC1','worksheet_name':'TB','region_id':'R','physical_row':4},sort_keys=True,separators=(',',':'))
    with (root/'annual.csv').open('w',newline='') as handle:
        writer=csv.DictWriter(handle,fieldnames=['Source_Record_ID','Account','Code','Period','Amount']); writer.writeheader(); writer.writerow({'Source_Record_ID':rid,'Account':'Bonus accrual','Code':'2100','Period':'FY25','Amount':'123.45'})
    (root/'execution_manifest.json').write_text(json.dumps({'execution_id':'E1','final_execution_status':'COMPLETED','outputs_created':['annual.csv']}))
    (root/'databook_metadata.json').write_text(json.dumps({'workflow_run_id':'RUN1','logical_datasets':[]}))
    config=tmp_path/'config'; _write_config(config)
    (config/'semantic_handoff.json').write_text(json.dumps({'handoff_version':'1.0','status':'CONFIRMED','package_id':'RUN1','datasets':[{'file':'annual.csv','usages':['OCL_RECORDS'],'fields':{'source_record_id':'Source_Record_ID','period':'Period','amount':'Amount','source_label':'Account','source_code':'Code'},'dimensions':[]}]}))
    output = tmp_path/'output'
    working = tmp_path/'work'/'ocl_runtime'/'RUN1'/'OCL_Databook_working.xlsx'
    support = output/'support working'/'RUN1'
    result=run_part1(root,config,output,working_databook=working,support_dir=support)
    assert result.state=='DATABOOK_READY'
    assert result.databook == working
    assert not (output/'OCL_Databook.xlsx').exists()
    assert (support/'OCL_Input_Review.xlsx').exists()
    assert (support/'OCL_Stage2_Review.xlsx').exists()
    assert result.review_context.parent == working.parent
    workbook=load_workbook(result.databook,data_only=False)
    source=[name for name in workbook.sheetnames if name.startswith('SRC_')][0]
    assert workbook[source].protection.sheet is True
    assert workbook['Flat File']['N3'].value==f"='{source}'!E2"
    assert workbook['Flat File']['A2'].value=='Source_Dataset'
    assert workbook['Balance by Category']['B7'].value=='Category'
    assert workbook['Balance by Category']['B8'].value == 'Employee accruals'
    assert workbook['Balance by Category']['C8'].value == '=SUM(C9:C9)'
    assert workbook['Balance by Category']['C9'].value.startswith('=SUMIFS(')
    assert workbook['Balance by Category'].row_dimensions[9].outlineLevel == 1
    assert workbook['Balance by Category'].row_dimensions[9].hidden is True


def test_part1_uses_actual_annual_and_monthly_periods_and_tb_control(tmp_path: Path):
    root=tmp_path/'latest'; root.mkdir()
    rid='{"source_id":"SRC","worksheet_name":"Sheet1"}'
    for filename,period in [('annual.csv','FY2026'),('monthly.csv','Jun-26')]:
        with (root/filename).open('w',newline='') as handle:
            writer=csv.DictWriter(handle,fieldnames=['Source_Record_ID','Account','Code','Period','Amount']); writer.writeheader(); writer.writerow({'Source_Record_ID':rid+filename,'Account':'Bonus accrual','Code':'2100','Period':period,'Amount':'125'})
    with (root/'tb.csv').open('w',newline='') as handle:
        writer=csv.DictWriter(handle,fieldnames=['Source_Record_ID','Period','Line','Amount']); writer.writeheader(); writer.writerow({'Source_Record_ID':'TB1','Period':'FY2026','Line':'OCL control','Amount':'125'})
    (root/'execution_manifest.json').write_text(json.dumps({'execution_id':'E1','final_execution_status':'COMPLETED','outputs_created':['annual.csv','monthly.csv','tb.csv']}))
    (root/'databook_metadata.json').write_text(json.dumps({'workflow_run_id':'RUN1','logical_datasets':[]}))
    config=tmp_path/'config'; _write_config(config)
    (config/'semantic_handoff.json').write_text(json.dumps({'handoff_version':'1.0','status':'CONFIRMED','package_id':'RUN1','datasets':[{'file':'annual.csv','usages':['OCL_RECORDS'],'fields':{'source_record_id':'Source_Record_ID','period':'Period','amount':'Amount','source_label':'Account','source_code':'Code'},'dimensions':[]},{'file':'monthly.csv','usages':['MONTHLY_RECORDS'],'fields':{'source_record_id':'Source_Record_ID','period':'Period','amount':'Amount','source_label':'Account','source_code':'Code'},'dimensions':[]},{'file':'tb.csv','usages':['TB_CONTROL'],'fields':{},'dimensions':[]}],'monthly_to_annual':[{'annual_period':'FY2026','monthly_period':'Jun-26'}],'controls':[{'control_id':'chk_listing_vs_tb','dataset_file':'tb.csv','period_field':'Period','amount_field':'Amount','filters':{'Line':'OCL control'}}]}))
    result=run_part1(root,config,tmp_path/'output')
    assert result.state=='DATABOOK_READY'
    assert result.blueprint.periods==('FY2026',)
    assert result.blueprint.monthly_periods==('Jun-26',)
    statuses={control.control_id:control.status.value for control in result.controls}
    assert statuses['chk_listing_vs_tb']=='PASS'
    assert statuses['chk_monthly_to_annual']=='PASS'
    workbook=load_workbook(result.databook,data_only=False)
    assert 'Monthly Flat' in workbook.sheetnames and 'Monthly Balance' in workbook.sheetnames
