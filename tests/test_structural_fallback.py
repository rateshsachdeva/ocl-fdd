from pathlib import Path

from openpyxl import Workbook


def test_one_period_per_sheet_fallback(tmp_path: Path):
    repo_root = Path(__file__).resolve().parents[1]
    embedded = repo_root / "fdd-data-preparation" / "src"
    import sys
    if str(embedded) not in sys.path:
        sys.path.insert(0, str(embedded))
    from fdd_data import prepare_source_package

    source = tmp_path / "source"
    output = tmp_path / "latest"
    source.mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "TB_FY2024"
    ws["A1"] = "Trial balance"
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Company Code", "G/L Account", "G/L Account Long Text", "Currency", "Closing Balance"])
    ws.append(["1000", "2100", "Bonus accrual", "EUR", 300])
    ws.append(["1000", "2110", "Holiday pay accrual", "EUR", 200])

    ws2 = wb.create_sheet("TB_FY2025")
    ws2["A1"] = "Trial balance"
    for _ in range(3):
        ws2.append([])
    ws2.append(["Company Code", "G/L Account", "G/L Account Long Text", "Currency", "Closing Balance"])
    ws2.append(["1000", "2100", "Bonus accrual", "EUR", 450])
    ws2.append(["1000", "2110", "Holiday pay accrual", "EUR", 250])
    wb.save(source / "client_tb.xlsx")

    result = prepare_source_package(source, output)
    assert (output / "ocl_annual.csv").exists()
    text = (output / "ocl_annual.csv").read_text(encoding="utf-8")
    assert "FY2024" in text
    assert "FY2025" in text
    assert "Bonus accrual" in text
    assert "450" in text
    assert result.warnings
    assert (tmp_path / "source_diagnostic.json").exists()
