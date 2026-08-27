import json
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ocl_agent.part2_analysis.ai_interpretation import (
    load_analysis_interpretation,
    write_analysis_request,
)
from ocl_agent.part2_analysis.ai_render import apply_partner_interpretation
from ocl_agent.schemas import AnalysisResult, AnalysisTable


def _analysis() -> AnalysisResult:
    return AnalysisResult(
        findings=(),
        tables=(
            AnalysisTable(
                "movement_review",
                "Latest annual movement review",
                ("Category", "FY24", "FY25", "Movement", "Movement %", "Databook Review", "Finding Eligible"),
                (("Bonus accrual", Decimal("300000"), Decimal("360000"), Decimal("60000"), 20.0, "REVIEW", ""),),
            ),
        ),
        annual_periods=("FY24", "FY25"),
        monthly_periods=(),
        latest_annual_period="FY25",
    )


def _key_finding():
    return {
        "id": "KF_01",
        "fdd_lens": "Normalized working capital",
        "area": "Working capital & balance validity",
        "metric": "Annual movement",
        "period_item": "FY24-FY25 / Bonus accrual",
        "so_what": "The movement is below the headline threshold but should be understood before concluding on the normal closing level.",
        "evidence": "Bonus accrual increased from 300,000 to 360,000, a 20.0% movement.",
        "evidence_limit": "Annual movement alone does not establish whether the closing balance is representative of the normal run-rate.",
        "fact_to_establish": "Whether the increase reflects normal recurring trading and the expected settlement timing.",
        "materiality": "NOTABLE",
        "ask_management": "Please explain the operational driver of the increase and expected settlement timing.",
        "linked_finding_id": None,
        "evidence_refs": ["table:movement_review:0"],
    }


def test_ai_partner_interpretation_uses_only_finalized_python_analysis(tmp_path: Path):
    analysis = _analysis()
    request_path = tmp_path / "analysis_evidence.json"
    interpretation_path = tmp_path / "analysis_interpretation.json"
    instruction_path = tmp_path / "FDD_PARTNER_ANALYSIS.md"
    instruction_path.write_text("test", encoding="utf-8")
    write_analysis_request(
        analysis,
        request_path,
        required_artifact=interpretation_path,
        instruction_path=instruction_path,
    )
    request = json.loads(request_path.read_text(encoding="utf-8"))
    assert request["analysis_status"] == "FINALIZED"
    assert request["source_scope"] == "PYTHON_ANALYSIS_ONLY"
    assert "table:movement_review:0" in request["valid_evidence_refs"]

    interpretation = {
        "status": "COMPLETED",
        "evidence_hash": request["evidence_hash"],
        "overall_assessment": "The available annual data shows no headline movement, although one category warrants follow-up for closing-balance representativeness.",
        "deal_issues": [],
        "key_findings": [_key_finding()],
        "management_questions": [
            {
                "id": "Q_01",
                "fdd_lens": "Normalized working capital",
                "theme": "Working capital & balance validity",
                "question": "Please explain the operational driver of the increase in bonus accrual and expected settlement timing.",
                "why_it_matters": "This would help determine whether the closing accrual is representative of normal working capital.",
                "evidence": "Bonus accrual increased by 20.0% between FY24 and FY25.",
                "priority": "MEDIUM",
                "linked_finding_id": None,
                "evidence_refs": ["table:movement_review:0"],
            }
        ],
    }
    interpretation_path.write_text(json.dumps(interpretation), encoding="utf-8")
    loaded = load_analysis_interpretation(interpretation_path, request_path)
    assert loaded["key_findings"][0]["fdd_lens"] == "Normalized working capital"


def test_ai_partner_render_surfaces_partner_context_and_never_leaves_tabs_blank(tmp_path: Path):
    analysis = _analysis()
    workbook_path = tmp_path / "OCL_Databook.xlsx"
    wb = Workbook()
    wb.active.title = "Flat File"
    wb.save(workbook_path)

    item = _key_finding()
    item.update(
        {
            "area": "Overall assessment",
            "metric": "Material deal issues",
            "period_item": "FY24-FY25",
            "so_what": "No material OCL deal issue was identified from the supplied evidence.",
            "evidence": "The available evidence did not support a headline OCL issue.",
            "evidence_limit": "Only the analyses included in the evidence package were assessed.",
            "fact_to_establish": "No further fact required from current evidence",
            "materiality": "NO_MATERIAL_ISSUE",
            "ask_management": "",
        }
    )
    interpretation = {
        "overall_assessment": "No material deal issue was identified from the available evidence.",
        "deal_issues": [],
        "key_findings": [item],
        "management_questions": [],
    }
    questions = apply_partner_interpretation(workbook_path, analysis, interpretation)
    assert questions == ()

    wb = load_workbook(workbook_path, read_only=False)
    assert wb["Deal Issues"]["A4"].value == "No material deal issue identified from the available evidence"
    assert wb["Deal Issues"]["A5"].value
    assert wb["Key Findings"]["B8"].value == "KF_01"
    assert wb["Key Findings"]["C8"].value == "Normalized working capital"
    assert wb["Key Findings"]["I8"].value == "Only the analyses included in the evidence package were assessed."
    headers = [wb["Key Findings"].cell(7, column).value for column in range(2, 12)]
    assert "Movement" not in headers and "Magnitude" not in headers and "Ask management" not in headers
    for header in ("FDD implication / So what", "Evidence", "Evidence limitation", "Fact to establish"):
        column = headers.index(header) + 2
        assert wb["Key Findings"].column_dimensions[wb["Key Findings"].cell(7, column).column_letter].width == 50
        assert wb["Key Findings"].cell(8, column).alignment.wrap_text is True
    assert not wb["Deal Issues"].merged_cells.ranges
    assert wb["Deal Issues"].column_dimensions["A"].width == 90
    assert wb["Deal Issues"]["A5"].alignment.wrap_text is True
    qa_headers = {wb["Q&A"].cell(7, column).value: column for column in range(2, 9)}
    for header in ("Question", "Why it matters", "Evidence trigger"):
        column = qa_headers[header]
        assert wb["Q&A"].column_dimensions[wb["Q&A"].cell(7, column).column_letter].width == 50
    assert "No material management question" in wb["Q&A"]["C8"].value
    wb.close()


def test_deal_issue_layout_is_single_column_without_figure_row(tmp_path: Path):
    workbook_path = tmp_path / "OCL_Databook.xlsx"
    workbook = Workbook()
    workbook.active.title = "Flat File"
    workbook.save(workbook_path)
    interpretation = {
        "overall_assessment": "One issue requires follow-up.",
        "deal_issues": [
            {
                "title": "Closing balance representativeness",
                "priority": "HIGH",
                "fdd_lens": "Normalized working capital",
                "so_what": "The closing balance may not represent the in-year level.",
                "evidence": "Monthly balances vary through the year.",
                "evidence_limit": "The schedule does not establish the operational cause.",
                "management_focus": "The operational driver and expected settlement timing.",
                "linked_finding_id": None,
            }
        ],
        "key_findings": [],
        "management_questions": [],
    }
    apply_partner_interpretation(workbook_path, _analysis(), interpretation)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Deal Issues"]
    assert not sheet.merged_cells.ranges
    assert sheet.column_dimensions["A"].width == 90
    assert all(sheet.cell(row, 1).alignment.wrap_text is True for row in range(4, 9))
    assert all(sheet.cell(row, column).value in (None, "") for row in range(1, sheet.max_row + 1) for column in range(2, sheet.max_column + 1))
    assert "Figure" not in [cell.value for row in sheet.iter_rows() for cell in row]
    workbook.close()
